import os
import re
import time
from pathlib import Path

from openpyxl import load_workbook


class DataExtract:
    def __init__(self):
        self.excel_path = r"C:\Users\admin\Desktop\tiktok-竞品信息抓取.xlsx"

    # 关键词(读取文件-提取关键词-返回关键词)
    def keywords(self):
        excel_path = getattr(self, 'excel_path', None) or os.environ.get('EXCEL_PATH')
        if excel_path:
            excel_path = Path(excel_path)
        else:
            excel_files = [
                p for p in Path.cwd().glob('*.xls*')
                if not p.name.startswith('~$')
            ]
            if not excel_files:
                raise FileNotFoundError('当前目录下未找到 Excel 表格文件')
            excel_path = max(excel_files, key=lambda p: p.stat().st_mtime)

        wb = load_workbook(excel_path, data_only=True)
        sheet_name = '关键词表格'
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'表格中未找到工作表：{sheet_name}')

        ws = wb[sheet_name]
        keyword_col = None
        header_row = None
        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value).strip() == '关键词':
                    keyword_col = cell.column
                    header_row = cell.row
                    break
            if keyword_col:
                break

        if not keyword_col:
            raise ValueError('关键词表格中未找到“关键词”列')

        result = []
        seen = set()
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row=row, column=keyword_col).value
            if value is None:
                continue
            keyword = str(value).strip()
            if not keyword or keyword in seen:
                continue
            result.append(keyword)
            seen.add(keyword)
            # print(result)
        return result

    def new_excel_shell(self):
        pass

if __name__ == '__main__':
    excel = DataExtract()
    # excel.keywords()
    excel.new_excel_shell()