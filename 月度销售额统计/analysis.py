
import pandas as pd
import os
from openpyxl import load_workbook
class ExcelUtil:
    def __init__(self):
        pass
    def get_video_sales_data(self,file_path):
        """
        读取Excel并返回二维数组：
        [Creator ID, Time, Video items sold, Shoppable video attributed GMV ($)]
        只保留 Video items sold > 0 的数据
        """

        # 读取Excel（表头在第3行）
        df = pd.read_excel(file_path, header=2)

        # 需要的字段
        cols = [
            "Creator name",
            "Time",
            "Video items sold",
            "Video indirect GMV ($)"
        ]

        # 筛选字段
        df = df[cols]

        # 过滤 Video items sold > 0
        df = df[df["Video items sold"] > 0]

        # 返回二维数组
        result = df.values.tolist()

        return result

    def get_file_paths(self,dir_path):
        file_paths = []

        for name in os.listdir(dir_path):
            full_path = os.path.join(dir_path, name)
            if os.path.isfile(full_path):
                file_paths.append(full_path)

        return file_paths

    def append_data_to_excel(self,file_path, data_list, sheet_name=None):
        """
        追加数据到特殊结构Excel

        参数
        file_path: excel路径
        data_list: [
            ['达人名','日期','销量','GMV'],
            ...
        ]
        sheet_name: 可选 sheet名
        """

        wb = load_workbook(file_path)

        ws = wb[sheet_name] if sheet_name else wb.active

        # 获取表头位置
        header_map = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(1, col).value
            if v:
                header_map[v] = col

        for name, date, sales, gmv in data_list:

            if name not in header_map:
                print(f"未找到达人: {name}")
                continue

            start_col = header_map[name]

            # 判断是否第一个达人
            if start_col == 1:
                date_col = start_col + 1
                gmv_col = start_col + 2
                sales_col = start_col + 3
            else:
                date_col = start_col
                gmv_col = start_col + 1
                sales_col = start_col + 2

            # 找空行
            row = 3
            while ws.cell(row=row, column=date_col).value:
                row += 1

            # ws.cell(row=row, column=date_col, value=date)
            # ws.cell(row=row, column=gmv_col, value=gmv)
            # ws.cell(row=row, column=sales_col, value=sales)
            ws.cell(row=row, column=date_col, value=date)  # 日期写入日期列
            ws.cell(row=row, column=gmv_col, value=gmv)  # GMV写入GMV列 ✅
            ws.cell(row=row, column=sales_col, value=sales)  # 销量写入销量列 ✅

        wb.save(file_path)

    def MergeData(self,file_url, SummaryTable):
        print(file_url, SummaryTable)
        files = self.get_file_paths(SummaryTable)
        for file in files:
            data = self.get_video_sales_data(file)
            self.append_data_to_excel(file_url, data)




if __name__ == '__main__':
    excel_util = ExcelUtil()
    SummaryTable = r'C:\RPA流程\月度销售额统计\汇总表.xlsx'
    file_url = r'C:\RPA流程\月度销售额统计\flie'
    excel_util.MergeData(SummaryTable, file_url)











