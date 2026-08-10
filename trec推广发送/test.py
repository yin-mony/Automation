"""TREC 推广发送测试流程。

本文件只放测试专用逻辑。正式流程没有明确授权前，不同步到 main.py。

功能：
1. 读取上游「trec公司+个人合作推广」子项目导出的公司/个人搜索匹配结果表。
2. 从每条记录中提取邮箱和电话。
3. 按“邮箱优先，没有邮箱再用电话”的规则生成待发送推广清单。
"""

from __future__ import annotations

import argparse
import csv
import html
import os
import re
import smtplib
from dataclasses import dataclass
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = CURRENT_DIR.parent
UPSTREAM_PROJECT_DIR = WORKSPACE_DIR / "trec公司+个人合作推广"
UPSTREAM_OUTPUT_DIR = UPSTREAM_PROJECT_DIR / "file"
LOCAL_OUTPUT_DIR = CURRENT_DIR / "file"

COMPANY_FILE_NAME = "已完成搜索匹配的公司联系信息数据.xlsx"
PERSON_FILE_NAME = "已完成搜索匹配的个人联系信息数据.xlsx"
SEND_QUEUE_FILE_NAME = "邮件发送记录.xlsx"
DEFAULT_COMPANY_FILE = UPSTREAM_OUTPUT_DIR / COMPANY_FILE_NAME
DEFAULT_PERSON_FILE = UPSTREAM_OUTPUT_DIR / PERSON_FILE_NAME

EMAIL_SUBJECT = "Partner with us on agent CE renewals"
EMAIL_MAIN_BODY_TEXT = """Hi,
Your agents' renewal season is coming up. We offer a full 18-hour TREC-approved CE package at $49.99 -- probably the lowest price they'll find. TREC Provider #11011-CEP.

Here's the deal: for every agent in your office who uses our package, I'll send you 20% back as a referral fee. No complicated setup -- just a straight split.

Your agents get a solid course at a great price. You get an easy way to help your team save while putting something back in your pocket.

Want to see the quality first? Let me know if you're open to a quick chat.

Best,
Qian Yi
Ho"""
EMAIL_SIGNATURE_TEXT = """Time2renew Support Team

Website: time2renew"""
EMAIL_BODY_TEXT = EMAIL_MAIN_BODY_TEXT + "\n\n\n" + EMAIL_SIGNATURE_TEXT
EMAIL_LOGO_PATH = CURRENT_DIR / "assets" / "time2renew-logo.png"
EMAIL_LOGO_CID = "time2renew-logo"
SEND_STATUS_PENDING = "待发送"
DEFAULT_TEST_SENDER = "info@time2renew.com"
DEFAULT_TEST_RECEIVER = "18280194086@163.com"
DEFAULT_SMTP_SERVER = "smtp.qiye.aliyun.com"
DEFAULT_SMTP_PORT = 465
SMTP_PASSWORD_ENV_KEYS = ("SMTP_AUTH_CODE", "TIME2RENEW_SMTP_AUTH_CODE")

EMAIL_PATTERN = re.compile(
    r"(?<![A-Za-z0-9._%+-])"
    r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    r"(?![A-Za-z0-9._%+-])"
)
PHONE_PATTERN = re.compile(
    r"(?<!\d)"
    r"(?:\+?1[\s.\-]*)?"
    r"(?:\([2-9]\d{2}\)|[2-9]\d{2})"
    r"[\s.\-]*[2-9]\d{2}[\s.\-]*\d{4}"
    r"(?!\d)"
)


@dataclass(frozen=True)
class SourceSpec:
    """描述一个上游搜索匹配结果文件。"""

    kind: str
    expected_file_name: str
    name_headers: tuple[str, ...]
    license_headers: tuple[str, ...]
    source_headers: tuple[str, ...]


@dataclass
class SourceRow:
    """上游 Excel/CSV 中的一条原始数据。"""

    kind: str
    path: Path
    row_number: int
    values: dict[str, str]


@dataclass
class SendCandidate:
    """整理后的待发送对象。"""

    kind: str
    name: str
    license_code: str
    channel: str
    contact_type: str
    contact_value: str
    emails: list[str]
    phones: list[str]
    subject: str
    body: str
    collect_status: str
    search_status: str
    source_urls: str
    source_file: str
    source_row_number: int
    send_status: str = SEND_STATUS_PENDING
    send_result: str = ""
    send_error: str = ""
    remark: str = ""

    def to_excel_row(self) -> list[str | int]:
        return [
            self.kind,
            self.name,
            self.license_code,
            self.channel,
            self.contact_type,
            self.contact_value,
            "; ".join(self.emails),
            self.subject,
            self.body,
            self.collect_status,
            self.search_status,
            self.source_urls,
            self.source_file,
            self.source_row_number,
            self.send_status,
            self.send_result,
            self.send_error,
            self.remark,
        ]


SOURCE_SPECS = [
    SourceSpec(
        kind="公司",
        expected_file_name=COMPANY_FILE_NAME,
        name_headers=("公司名称", "company_name"),
        license_headers=("公司许可证号", "许可证号", "code"),
        source_headers=("联系方式来源链接", "source_urls"),
    ),
    SourceSpec(
        kind="个人",
        expected_file_name=PERSON_FILE_NAME,
        name_headers=("姓名", "name"),
        license_headers=("许可证号", "code"),
        source_headers=("HAR来源链接", "Google来源链接", "har_source_urls", "google_source_urls"),
    ),
]

EMAIL_HEADERS = ("邮箱", "emails", "email", "office email", "office_email")
PHONE_HEADERS = ("电话", "phones", "phone", "mobile", "tel")
PREFERRED_HEADERS = ("优先联系方式", "preferred_contact_value")
COLLECT_STATUS_HEADERS = ("采集状态", "contact_status")
SEARCH_STATUS_HEADERS = ("搜索状态", "search_status")

OUTPUT_HEADERS = [
    "来源类型",
    "对象名称",
    "许可证号",
    "发送渠道",
    "联系方式类型",
    "联系方式",
    "邮箱列表",
    "邮件主题",
    "邮件正文",
    "采集状态",
    "搜索状态",
    "来源链接",
    "来源文件",
    "来源行号",
    "发送状态",
    "发送结果",
    "失败原因",
    "备注",
]


def clean_text(value) -> str:
    """把 Excel/CSV 单元格值统一成干净字符串。"""

    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_header(value: str) -> str:
    """表头匹配时忽略大小写、空格、下划线和常见标点。"""

    text = clean_text(value).lower()
    for char in (" ", "_", "-", "：", ":", "（", "）", "(", ")"):
        text = text.replace(char, "")
    return text


def first_existing_value(row: dict[str, str], headers: Iterable[str]) -> str:
    for header in headers:
        value = row.get(header, "")
        if value:
            return value
    return ""


def combined_value(row: dict[str, str], headers: Iterable[str]) -> str:
    values = []
    for header in headers:
        value = row.get(header, "")
        if value:
            values.append(value)
    return "; ".join(values)


def unique_keep_order(values: Iterable[str]) -> list[str]:
    result = []
    seen = set()
    for value in values:
        key = value.lower()
        if key and key not in seen:
            seen.add(key)
            result.append(value)
    return result


def extract_emails(text: str) -> list[str]:
    matches = []
    for match in EMAIL_PATTERN.findall(text or ""):
        email = match.lower().strip("._-")
        if ".." not in email:
            matches.append(email)
    return unique_keep_order(matches)


def normalize_phone(raw_phone: str) -> str:
    digits = re.sub(r"\D", "", raw_phone or "")
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return ""
    if digits[0] in "01" or digits[3] in "01":
        return ""
    return digits


def extract_phones(text: str) -> list[str]:
    phones = []
    for match in PHONE_PATTERN.findall(text or ""):
        phone = normalize_phone(match)
        if phone:
            phones.append(phone)
    return unique_keep_order(phones)


def resolve_headers(headers: Iterable[str]) -> dict[str, str]:
    """生成规范表头到真实表头的映射，兼容中文表头和内部英文字段。"""

    resolved = {}
    for header in headers:
        header_text = clean_text(header)
        if not header_text:
            continue
        resolved[normalize_header(header_text)] = header_text
    return resolved


def row_value(row: dict[str, str], aliases: Iterable[str]) -> str:
    for alias in aliases:
        value = row.get(alias, "")
        if value:
            return value
    return ""


def read_xlsx_rows(path: Path, kind: str) -> list[SourceRow]:
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        row_iterator = worksheet.iter_rows(values_only=True)
        raw_headers = next(row_iterator, None)
        if not raw_headers:
            return []

        header_map = resolve_headers(clean_text(header) for header in raw_headers)
        headers = [
            header_map.get(normalize_header(clean_text(header)), clean_text(header))
            for header in raw_headers
        ]

        rows = []
        for row_number, values in enumerate(row_iterator, start=2):
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else ""
                row[header] = clean_text(value)
            rows.append(SourceRow(kind=kind, path=path, row_number=row_number, values=row))
        return rows
    finally:
        workbook.close()


def read_csv_rows(path: Path, kind: str) -> list[SourceRow]:
    rows = []
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row_number, row in enumerate(reader, start=2):
            rows.append(
                SourceRow(
                    kind=kind,
                    path=path,
                    row_number=row_number,
                    values={clean_text(key): clean_text(value) for key, value in row.items()},
                )
            )
    return rows


def read_source_rows(path: Path, kind: str) -> list[SourceRow]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx_rows(path, kind)
    if suffix == ".csv":
        return read_csv_rows(path, kind)
    raise ValueError(f"不支持的文件类型: {path}")


def locate_result_file(output_dir: Path, spec: SourceSpec) -> Path | None:
    exact_path = output_dir / spec.expected_file_name
    if exact_path.exists():
        return exact_path

    patterns = [
        f"*搜索匹配*{spec.kind}*联系信息*.xlsx",
        f"*{spec.kind}*联系信息*.xlsx",
        f"*搜索匹配*{spec.kind}*.csv",
        f"*{spec.kind}*联系信息*.csv",
    ]
    matches = []
    for pattern in patterns:
        matches.extend(output_dir.glob(pattern))
    if not matches:
        return None
    return max(matches, key=lambda item: item.stat().st_mtime)


def collect_contacts(row: dict[str, str], include_preferred: bool = True) -> tuple[list[str], list[str]]:
    preferred_headers = PREFERRED_HEADERS if include_preferred else ()
    email_text = combined_value(row, EMAIL_HEADERS + preferred_headers)
    phone_text = combined_value(row, PHONE_HEADERS + preferred_headers)
    emails = extract_emails(email_text)
    phones = extract_phones(phone_text)
    return emails, phones


def build_candidate(source_row: SourceRow, spec: SourceSpec) -> SendCandidate | None:
    row = source_row.values
    emails, phones = collect_contacts(row)
    if not emails:
        return None

    source_urls = combined_value(row, spec.source_headers)
    return SendCandidate(
        kind=source_row.kind,
        name=row_value(row, spec.name_headers),
        license_code=row_value(row, spec.license_headers),
        channel="邮件",
        contact_type="邮箱",
        contact_value=emails[0],
        emails=emails,
        phones=phones,
        subject=EMAIL_SUBJECT,
        body=EMAIL_BODY_TEXT,
        collect_status=first_existing_value(row, COLLECT_STATUS_HEADERS),
        search_status=first_existing_value(row, SEARCH_STATUS_HEADERS),
        source_urls=source_urls,
        source_file=source_row.path.name,
        source_row_number=source_row.row_number,
    )


def infer_source_spec(path: Path, kind: str = "") -> SourceSpec:
    kind_text = clean_text(kind)
    if kind_text:
        for spec in SOURCE_SPECS:
            if spec.kind == kind_text:
                return spec

    path_text = path.name
    for spec in SOURCE_SPECS:
        if spec.kind in path_text:
            return spec
    return SOURCE_SPECS[0]


def build_channel_candidate(
    source_row: SourceRow,
    spec: SourceSpec,
    contact_value: str,
    emails: list[str],
    phones: list[str],
) -> SendCandidate:
    row = source_row.values
    return SendCandidate(
        kind=source_row.kind,
        name=row_value(row, spec.name_headers),
        license_code=row_value(row, spec.license_headers),
        channel="邮件",
        contact_type="邮箱",
        contact_value=contact_value,
        emails=emails,
        phones=phones,
        subject=EMAIL_SUBJECT,
        body=EMAIL_BODY_TEXT,
        collect_status=first_existing_value(row, COLLECT_STATUS_HEADERS),
        search_status=first_existing_value(row, SEARCH_STATUS_HEADERS),
        source_urls=combined_value(row, spec.source_headers),
        source_file=source_row.path.name,
        source_row_number=source_row.row_number,
    )


def build_candidates_from_source_files(
    source_files: Iterable[tuple[str, str | Path]],
    use_email: bool = True,
) -> tuple[list[SendCandidate], list[str]]:
    candidates = []
    warnings = []

    if not use_email:
        use_email = True

    for kind, file_path in source_files:
        path = Path(file_path)
        spec = infer_source_spec(path, kind)
        if not path.exists():
            warnings.append(f"文件不存在，已跳过: {path}")
            continue

        rows = read_source_rows(path, spec.kind)
        for source_row in rows:
            emails, phones = collect_contacts(source_row.values, include_preferred=False)
            if use_email:
                for email in emails:
                    candidates.append(
                        build_channel_candidate(source_row, spec, email, emails, phones)
                    )

    return candidates, warnings


def build_send_candidates(output_dir: Path) -> tuple[list[SendCandidate], list[str]]:
    candidates = []
    warnings = []
    seen = set()

    for spec in SOURCE_SPECS:
        source_path = locate_result_file(output_dir, spec)
        if not source_path:
            warnings.append(f"未找到{spec.kind}搜索匹配结果文件: {spec.expected_file_name}")
            continue

        rows = read_source_rows(source_path, spec.kind)
        for source_row in rows:
            candidate = build_candidate(source_row, spec)
            if not candidate:
                continue

            # 同一个对象只进入一次待发送队列，符合“一人只发一件”的业务要求。
            dedupe_key = (
                candidate.kind,
                candidate.name.lower(),
                candidate.license_code.lower(),
                candidate.contact_value.lower(),
            )
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            candidates.append(candidate)

    return candidates, warnings


def save_candidates(candidates: list[SendCandidate], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Promotion Send Queue"
    worksheet.append(OUTPUT_HEADERS)

    for candidate in candidates:
        worksheet.append(candidate.to_excel_row())

    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(str(output_path))


def first_env_value(keys: Iterable[str] | str) -> tuple[str, str]:
    if isinstance(keys, str):
        key_or_password = keys.strip()
        if not key_or_password:
            return "", ""
        env_value = os.getenv(key_or_password)
        if env_value:
            return key_or_password, env_value
        return "SMTP_PASSWORD_ENV_KEYS", key_or_password

    for key in keys:
        value = os.getenv(key)
        if value:
            return key, value
    return "", ""


def text_to_html_paragraphs(text: str) -> str:
    paragraphs = []
    for paragraph in text.strip().split("\n\n"):
        lines = [html.escape(line) for line in paragraph.splitlines()]
        paragraphs.append("<p style=\"margin:0 0 14px 0;\">" + "<br>".join(lines) + "</p>")
    return "\n".join(paragraphs)


def build_email_html() -> str:
    main_html = text_to_html_paragraphs(EMAIL_MAIN_BODY_TEXT)
    website_url = "https://time2renew.com"
    return f"""<!doctype html>
<html>
<body style="font-family:Arial, Helvetica, sans-serif; font-size:14px; color:#111; line-height:1.45;">
{main_html}
<div style="margin-top:28px;">
  <p style="font-size:18px; font-weight:700; margin:0 0 22px 0;">Time2renew Support Team</p>
  <p style="margin:0 0 22px 0;"><strong>Website:</strong> <a href="{website_url}" style="color:#111; text-decoration:underline;">time2renew</a></p>
  <img src="cid:{EMAIL_LOGO_CID}" alt="Time2Renew" width="180" style="display:block; width:180px; height:auto; border:0;">
</div>
</body>
</html>"""


def build_email_message(sender: str, receiver: str) -> MIMEMultipart:
    if not EMAIL_LOGO_PATH.exists():
        raise FileNotFoundError(f"邮件 Logo 图片不存在: {EMAIL_LOGO_PATH}")

    message = MIMEMultipart("related")
    message["From"] = formataddr((str(Header("Time2renew Support Team", "utf-8")), sender))
    message["To"] = receiver
    message["Subject"] = Header(EMAIL_SUBJECT, "utf-8").encode()

    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(EMAIL_BODY_TEXT, "plain", "utf-8"))
    alternative.attach(MIMEText(build_email_html(), "html", "utf-8"))
    message.attach(alternative)

    with EMAIL_LOGO_PATH.open("rb") as image_file:
        image = MIMEImage(image_file.read(), _subtype="png")
    image.add_header("Content-ID", f"<{EMAIL_LOGO_CID}>")
    image.add_header("Content-Disposition", "inline", filename=EMAIL_LOGO_PATH.name)
    message.attach(image)
    return message


def send_test_email(
    sender: str,
    receiver: str,
    smtp_server: str,
    smtp_port: int,
    smtp_user: str = "",
    password_env: str = "",
) -> bool:
    """发送一封真实测试邮件。密码只从环境变量读取，不写入代码。"""

    sender = clean_text(sender)
    receiver = clean_text(receiver)
    smtp_server = clean_text(smtp_server)
    smtp_user = clean_text(smtp_user) or sender
    env_keys = (password_env,) if password_env else SMTP_PASSWORD_ENV_KEYS
    password_key, smtp_password = first_env_value(env_keys)

    if not sender:
        raise ValueError("测试发件人不能为空")
    if not receiver:
        raise ValueError("测试收件人不能为空")
    if not smtp_server:
        raise ValueError("SMTP 服务器不能为空")
    if not smtp_password:
        env_hint = password_env or (
            "SMTP_PASSWORD_ENV_KEYS"
            if isinstance(SMTP_PASSWORD_ENV_KEYS, str)
            else " / ".join(SMTP_PASSWORD_ENV_KEYS)
        )
        raise ValueError(f"未找到 SMTP 密码/授权码环境变量: {env_hint}")

    message = build_email_message(sender, receiver)

    server = smtplib.SMTP_SSL(smtp_server, int(smtp_port), timeout=30)
    try:
        server.login(smtp_user, smtp_password)
        server.sendmail(sender, [receiver], message.as_bytes())
    finally:
        server.quit()

    print("测试邮件发送成功")
    print("发件人:", sender)
    print("收件人:", receiver)
    print("SMTP 服务器:", f"{smtp_server}:{smtp_port}")
    print("SMTP 用户:", smtp_user)
    print("密码来源环境变量:", password_key)
    return True


def log_line(log_callback: Callable[[str], None] | None, message: str) -> None:
    if log_callback:
        log_callback(message)
    else:
        print(message)


def dispatch_candidates(
    candidates: list[SendCandidate],
    execute_send: bool = False,
    sender: str = DEFAULT_TEST_SENDER,
    smtp_server: str = DEFAULT_SMTP_SERVER,
    smtp_port: int = DEFAULT_SMTP_PORT,
    smtp_user: str = "",
    password_env: str = "",
    log_callback: Callable[[str], None] | None = None,
) -> dict[str, int]:
    summary = {
        "total": len(candidates),
        "email_total": sum(1 for item in candidates if item.channel == "邮件"),
        "email_sent": 0,
        "email_failed": 0,
    }

    if not execute_send:
        for candidate in candidates:
            candidate.send_status = "待发送"
            candidate.send_result = "已生成邮件记录，未执行发送"
        return summary

    email_candidates = [item for item in candidates if item.channel == "邮件"]

    if email_candidates:
        smtp_user = clean_text(smtp_user) or clean_text(sender)
        env_keys = (password_env,) if password_env else SMTP_PASSWORD_ENV_KEYS
        _, smtp_password = first_env_value(env_keys)
        if not smtp_password:
            for candidate in email_candidates:
                candidate.send_status = "发送失败"
                candidate.send_result = "邮件未发送"
                candidate.send_error = "未配置 SMTP 密码/授权码"
            summary["email_failed"] = len(email_candidates)
            log_line(log_callback, "邮件发送失败: 未配置 SMTP 密码/授权码")
        else:
            server = smtplib.SMTP_SSL(smtp_server, int(smtp_port), timeout=30)
            try:
                server.login(smtp_user, smtp_password)
                for candidate in email_candidates:
                    try:
                        message = build_email_message(sender, candidate.contact_value)
                        server.sendmail(sender, [candidate.contact_value], message.as_bytes())
                        candidate.send_status = "已发送"
                        candidate.send_result = "邮件发送成功"
                        summary["email_sent"] += 1
                        log_line(log_callback, f"邮件发送成功: {candidate.contact_value}")
                    except Exception as error:
                        candidate.send_status = "发送失败"
                        candidate.send_result = "邮件发送失败"
                        candidate.send_error = str(error)
                        summary["email_failed"] += 1
                        log_line(log_callback, f"邮件发送失败: {candidate.contact_value} - {error}")
            except Exception as error:
                for candidate in email_candidates:
                    candidate.send_status = "发送失败"
                    candidate.send_result = "邮件登录或连接失败"
                    candidate.send_error = str(error)
                summary["email_failed"] = len(email_candidates)
                log_line(log_callback, f"邮件登录或连接失败: {error}")
            finally:
                try:
                    server.quit()
                except Exception:
                    pass

    return summary


def process_source_files(
    source_files: Iterable[tuple[str, str | Path]],
    use_email: bool = True,
    execute_send: bool = False,
    output_file: str | Path | None = None,
    sender: str = DEFAULT_TEST_SENDER,
    smtp_server: str = DEFAULT_SMTP_SERVER,
    smtp_port: int = DEFAULT_SMTP_PORT,
    smtp_user: str = "",
    password_env: str = "",
    log_callback: Callable[[str], None] | None = None,
) -> dict[str, object]:
    output_path = Path(output_file) if output_file else (LOCAL_OUTPUT_DIR / SEND_QUEUE_FILE_NAME)
    candidates, warnings = build_candidates_from_source_files(
        source_files=source_files,
        use_email=use_email,
    )
    for warning in warnings:
        log_line(log_callback, f"提示: {warning}")

    summary = dispatch_candidates(
        candidates=candidates,
        execute_send=execute_send,
        sender=sender,
        smtp_server=smtp_server,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        password_env=password_env,
        log_callback=log_callback,
    )
    save_candidates(candidates, output_path)
    log_line(log_callback, f"已保存邮件发送记录: {output_path}")
    log_line(log_callback, f"邮件任务数量: {len(candidates)}")
    return {
        "output_file": str(output_path),
        "candidates": candidates,
        "warnings": warnings,
        "summary": summary,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成 TREC 推广待发送联系清单")
    parser.add_argument(
        "--upstream-file-dir",
        default=str(UPSTREAM_OUTPUT_DIR),
        help="trec公司+个人合作推广 子项目 file 目录",
    )
    parser.add_argument(
        "--file-file",
        default=str(LOCAL_OUTPUT_DIR / SEND_QUEUE_FILE_NAME),
        help="邮件发送记录输出文件",
    )
    parser.add_argument(
        "--company-file",
        default=str(DEFAULT_COMPANY_FILE),
        help="公司推广搜索匹配结果文件",
    )
    parser.add_argument(
        "--person-file",
        default=str(DEFAULT_PERSON_FILE),
        help="个人推广搜索匹配结果文件",
    )
    parser.add_argument(
        "--include-company",
        action="store_true",
        help="处理公司推广文件",
    )
    parser.add_argument(
        "--include-person",
        action="store_true",
        help="处理个人推广文件",
    )
    parser.add_argument(
        "--use-email",
        action="store_true",
        help="生成/执行邮件发送任务",
    )
    parser.add_argument(
        "--execute-send",
        action="store_true",
        help="真实执行发送；不加时只生成邮件发送记录",
    )
    parser.add_argument(
        "--send-test-email",
        action="store_true",
        help="发送一封真实测试邮件，只测试邮件模板，不生成待发送清单",
    )
    parser.add_argument(
        "--sender",
        default=DEFAULT_TEST_SENDER,
        help="测试邮件发件人",
    )
    parser.add_argument(
        "--receiver",
        default=DEFAULT_TEST_RECEIVER,
        help="测试邮件收件人",
    )
    parser.add_argument(
        "--smtp-server",
        default=DEFAULT_SMTP_SERVER,
        help="SMTP 服务器",
    )
    parser.add_argument(
        "--smtp-port",
        type=int,
        default=DEFAULT_SMTP_PORT,
        help="SMTP SSL 端口",
    )
    parser.add_argument(
        "--smtp-user",
        default="",
        help="SMTP 登录用户，默认使用发件人邮箱",
    )
    parser.add_argument(
        "--smtp-password-env",
        default="",
        help="指定读取 SMTP 密码/授权码的环境变量名",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.send_test_email:
        try:
            send_test_email(
                sender=args.sender,
                receiver=args.receiver,
                smtp_server=args.smtp_server,
                smtp_port=args.smtp_port,
                smtp_user=args.smtp_user,
                password_env=args.smtp_password_env,
            )
            return 0
        except Exception as error:
            print("测试邮件发送失败:", error)
            return 1

    use_selected_flow = (
        args.include_company
        or args.include_person
        or args.use_email
        or args.execute_send
    )
    if use_selected_flow:
        source_files = []
        if args.include_company:
            source_files.append(("公司", args.company_file))
        if args.include_person:
            source_files.append(("个人", args.person_file))
        if not source_files:
            source_files = [("公司", args.company_file), ("个人", args.person_file)]

        process_source_files(
            source_files=source_files,
            use_email=True,
            execute_send=args.execute_send,
            output_file=args.output_file,
            sender=args.sender,
            smtp_server=args.smtp_server,
            smtp_port=args.smtp_port,
            smtp_user=args.smtp_user,
            password_env=args.smtp_password_env,
        )
        return 0

    upstream_output_dir = Path(args.upstream_output_dir).resolve()
    output_file = Path(args.output_file).resolve()

    if not upstream_output_dir.exists():
        print("上游 file 目录不存在:", upstream_output_dir)
        save_candidates([], output_file)
        print("已生成空待发送清单:", output_file)
        return 0

    candidates, warnings = build_send_candidates(upstream_output_dir)
    save_candidates(candidates, output_file)

    for warning in warnings:
        print("提示:", warning)
    print("已生成待发送推广联系清单:", output_file)
    print("待发送数量:", len(candidates))
    print("邮箱数量:", sum(1 for item in candidates if item.contact_type == "邮箱"))
    print("电话数量:", sum(1 for item in candidates if item.contact_type == "电话"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
