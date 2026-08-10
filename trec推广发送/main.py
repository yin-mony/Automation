"""TREC 推广邮件正式流程。

main.py 只负责正式业务逻辑：读取上游公司/个人搜索匹配结果表，提取“邮箱”列，
生成并发送固定推广邮件，同时把后台发送记录保存到本子项目 file 目录。

测试逻辑保留在 test.py，GUI 窗口入口保留在 run.py。
"""

from __future__ import annotations

import csv
import html
import re
import smtplib
import time
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from pathlib import Path

from openpyxl import Workbook, load_workbook


class Main:
    """TREC 公司/个人推广邮件发送主流程。"""

    @staticmethod
    def defaultConfig():
        """返回 main.py 和 run.py 共用的默认配置。"""
        baseDir = Path(__file__).resolve().parent
        workspaceDir = baseDir.parent
        upstreamOutputDir = workspaceDir / "trec公司+个人合作推广" / "file"

        return {
            # GUI 窗口标题：run.py 使用该值设置窗口名称。
            "windowTitle": "TREC 推广邮件发送",

            # GUI 默认窗口尺寸：格式为 宽x高。
            "windowSize": "980x620",

            # GUI 最小窗口宽度：避免窗口过窄导致文件路径输入框不可用。
            "windowMinWidth": 860,

            # GUI 最小窗口高度：保证日志框和发送按钮有足够显示空间。
            "windowMinHeight": 540,

            # 运行环境标记：False 表示本机调试，True 表示线上/服务器运行。
            "isOnline": False,

            # 公司推广结果文件：默认读取上游公司搜索匹配后的固定测试文件。
            "companyFile": str(upstreamOutputDir / "已完成搜索匹配的公司联系信息数据.xlsx"),

            # 个人推广结果文件：默认读取上游个人搜索匹配后的固定测试文件。
            "personFile": str(upstreamOutputDir / "已完成搜索匹配的个人联系信息数据.xlsx"),

            # 默认是否处理公司文件：GUI 初始勾选公司推广数据。
            "includeCompany": True,

            # 默认是否处理个人文件：GUI 初始勾选个人推广数据。
            "includePerson": True,

            # 后台发送记录文件：记录每封邮件的发送状态，不在 GUI 中展示任务清单。
            "outputFile": str(baseDir / "file" / "邮件发送记录.xlsx"),

            # 发件邮箱地址：固定系统参数，不交给 GUI 或外部配置修改。
            "senderEmail": "info@time2renew.com",

            # SMTP 服务器地址：固定使用 time2renew.com 的阿里企业邮箱 SMTP。
            "smtpServer": "smtp.qiye.aliyun.com",

            # SMTP SSL 端口：固定使用阿里企业邮箱 SSL 发送端口。
            "smtpPort": 465,

            # SMTP 登录用户名：固定使用发件邮箱登录，不交给 GUI 修改。
            "smtpUser": "info@time2renew.com",

            # SMTP 授权码/密码：固定系统参数，GUI 只读掩码展示，不允许修改。
            "smtpPassword": "__TIME2RENEW_SMTP_AUTH_CODE__",

            # 邮件主题：写入邮件客户端 Subject 字段，不放进正文。
            "emailSubject": "Partner with us on agent CE renewals",

            # 邮件正文主体：不包含底部团队签名和 Logo，姓名按 Qian Yi / Ho 分两行。
            "emailMainBody": (
                "Hi,\n"
                "Your agents' renewal season is coming up. We offer a full 18-hour "
                "TREC-approved CE package at $49.99 -- probably the lowest price they'll "
                "find. TREC Provider #11011-CEP.\n\n"
                "Here's the deal: for every agent in your office who uses our package, "
                "I'll send you 20% back as a referral fee. No complicated setup -- just "
                "a straight split.\n\n"
                "Your agents get a solid course at a great price. You get an easy way "
                "to help your team save while putting something back in your pocket.\n\n"
                "Want to see the quality first? Let me know if you're open to a quick chat.\n\n"
                "Best,\n"
                "Qian Yi\n"
                "Ho"
            ),

            # 邮件签名文字：默认追加到正文末尾、Logo 图片上方。
            "emailSignature": "Time2renew Support Team\n\nWebsite: time2renew",

            # 邮件官网链接：HTML 邮件中 time2renew 文本点击后打开该网址。
            "websiteUrl": "https://time2renew.com",

            # 邮件 Logo 图片路径：HTML 邮件以 CID 内嵌图片方式放到签名下方。
            "logoPath": str(baseDir / "assets" / "time2renew-logo.png"),

            # 邮件 Logo CID：HTML 正文通过 cid 引用内嵌图片。
            "logoCid": "time2renew-logo",

            # 邮件 Logo 展示宽度：控制收件箱中签名图标视觉尺寸。
            "logoWidth": 180,

            # 邮件发送间隔秒数：默认不等待，如遇 SMTP 限流可调大。
            "sendWaitSecond": 0.0,
        }

    @staticmethod
    def configHelpText():
        """返回 GUI 可使用的配置说明文案。"""
        return {
            "isOnline": "运行环境选择。本机表示电脑调试；线上表示正式或服务器环境运行。",
            "companyFile": "公司模式搜索完成后的结果表，默认读取上游 file 中的固定文件。",
            "personFile": "个人模式搜索完成后的结果表，默认读取上游 file 中的固定文件。",
            "includeCompany": "是否读取公司推广数据文件。",
            "includePerson": "是否读取个人推广数据文件。",
            "senderEmail": "固定发件邮箱地址，不在 GUI 中开放修改。",
            "smtpServer": "固定 SMTP 服务器地址，不在 GUI 中开放修改。",
            "smtpPort": "固定 SMTP SSL 端口，不在 GUI 中开放修改。",
            "smtpPassword": "固定 SMTP 授权码，GUI 只读掩码展示，不允许修改。",
            "outputFile": "后台邮件发送记录文件，GUI 不展示任务清单。",
        }

    def __init__(self, config=None, logCallback=None):
        """初始化正式流程配置、列名规则、邮件模板和日志回调。"""
        if config is None:
            config = {}

        # 配置合并规则：调用方只传需要覆盖的非 SMTP 值，其余统一使用 Main.defaultConfig()。
        defaultConfig = self.defaultConfig()
        fixedSmtpKeys = {"senderEmail", "smtpServer", "smtpPort", "smtpUser", "smtpPassword"}
        for key, value in config.items():
            if key not in fixedSmtpKeys:
                defaultConfig[key] = value
        self.config = defaultConfig
        self.logCallback = logCallback

        # GUI 窗口配置：run.py 直接读取这些默认值创建界面。
        self.windowTitle = str(self.config.get("windowTitle"))
        self.windowSize = str(self.config.get("windowSize"))
        self.windowMinWidth = int(self.config.get("windowMinWidth"))
        self.windowMinHeight = int(self.config.get("windowMinHeight"))
        self.isOnline = bool(self.config.get("isOnline"))

        # 输入文件配置：只读取上游搜索匹配结果表，不重新执行搜索流程。
        self.companyFile = Path(str(self.config.get("companyFile")))
        self.personFile = Path(str(self.config.get("personFile")))
        self.includeCompany = bool(self.config.get("includeCompany"))
        self.includePerson = bool(self.config.get("includePerson"))

        # 输出记录配置：发送结果只写后台记录，不在 GUI 中显示任务清单。
        self.outputFile = Path(str(self.config.get("outputFile")))

        # SMTP 固定配置：这些值只允许在 main.py 默认配置中维护，GUI 和外部 config 不可覆盖。
        self.senderEmail = str(self.config.get("senderEmail") or "").strip()
        self.smtpServer = str(self.config.get("smtpServer") or "").strip()
        self.smtpPort = int(self.config.get("smtpPort") or 465)
        self.smtpUser = str(self.config.get("smtpUser") or "").strip()
        self.smtpPassword = str(self.config.get("smtpPassword") or "").strip()

        # 邮件模板配置：主题进入 Subject，正文和签名进入邮件 Body。
        self.emailSubject = str(self.config.get("emailSubject") or "").strip()
        self.emailMainBody = str(self.config.get("emailMainBody") or "").strip()
        self.emailSignature = str(self.config.get("emailSignature") or "").strip()
        self.websiteUrl = str(self.config.get("websiteUrl") or "https://time2renew.com").strip()
        self.logoPath = Path(str(self.config.get("logoPath") or ""))
        self.logoCid = str(self.config.get("logoCid") or "time2renew-logo").strip()
        self.logoWidth = int(self.config.get("logoWidth") or 180)
        self.sendWaitSecond = float(self.config.get("sendWaitSecond") or 0)

        # 固定列名配置：正式发送流程只从“邮箱”列提取收件地址。
        self.emailHeaders = ("邮箱", "email", "emails", "office email", "office_email")
        self.collectStatusHeaders = ("采集状态", "contact_status")
        self.searchStatusHeaders = ("搜索状态", "search_status")
        self.companyNameHeaders = ("公司名称", "company_name")
        self.companyLicenseHeaders = ("公司许可证号", "许可证号", "code")
        self.companySourceHeaders = ("联系方式来源链接", "source_urls")
        self.personNameHeaders = ("姓名", "name")
        self.personLicenseHeaders = ("许可证号", "code")
        self.personSourceHeaders = ("HAR来源链接", "Google来源链接", "har_source_urls", "google_source_urls")

        # 邮箱正则：只识别标准邮箱地址，避免把无效文本当成收件人。
        self.emailPattern = re.compile(
            r"(?<![A-Za-z0-9._%+-])"
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
            r"(?![A-Za-z0-9._%+-])"
        )

        # 后台发送记录表头：用于追踪每封邮件的来源行和发送结果。
        self.outputHeaders = [
            "来源类型",
            "运行环境",
            "对象名称",
            "许可证号",
            "邮箱",
            "邮件主题",
            "邮件正文",
            "采集状态",
            "搜索状态",
            "来源链接",
            "来源文件",
            "来源行号",
            "发送状态",
            "发送结果",
            "失败原因",
            "备注",
        ]

    def log(self, message):
        """输出日志到 GUI 或命令行。"""
        if self.logCallback:
            self.logCallback(str(message))
        else:
            print(message)

    def cleanText(self, value):
        """把单元格值转换成干净字符串。"""
        if value is None:
            return ""
        text = str(value).strip()
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    def normalizeHeader(self, value):
        """表头匹配时忽略大小写、空格、下划线和常见标点。"""
        text = self.cleanText(value).lower()
        for char in (" ", "_", "-", "：", ":", "（", "）", "(", ")"):
            text = text.replace(char, "")
        return text

    def uniqueKeepOrder(self, values):
        """按出现顺序去重，避免同一单元格内重复邮箱发送两次。"""
        result = []
        seen = set()
        for value in values:
            key = str(value).lower()
            if key and key not in seen:
                seen.add(key)
                result.append(value)
        return result

    def extractEmails(self, text):
        """从邮箱列文本中提取一个或多个邮箱地址。"""
        emails = []
        for match in self.emailPattern.findall(text or ""):
            email = match.lower().strip("._-")
            if email and ".." not in email:
                emails.append(email)
        return self.uniqueKeepOrder(emails)

    def getRowValue(self, rowData, headers):
        """按多个候选表头读取第一个非空值。"""
        normalizedHeaders = {self.normalizeHeader(header) for header in headers}
        for key, value in rowData.items():
            if self.normalizeHeader(key) in normalizedHeaders and value:
                return value
        return ""

    def combineValues(self, rowData, headers):
        """按多个候选表头合并所有非空值。"""
        normalizedHeaders = {self.normalizeHeader(header) for header in headers}
        values = []
        for key, value in rowData.items():
            if self.normalizeHeader(key) in normalizedHeaders and value:
                values.append(value)
        return "; ".join(values)

    def readExcelRows(self, filePath, sourceType):
        """读取 Excel 文件并返回带来源行号的数据。"""
        workbook = load_workbook(str(filePath), read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rowIterator = worksheet.iter_rows(values_only=True)
            rawHeaders = next(rowIterator, None)
            if not rawHeaders:
                return []

            headers = [self.cleanText(header) for header in rawHeaders]
            rows = []
            for rowNumber, values in enumerate(rowIterator, start=2):
                rowData = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[index] if index < len(values) else ""
                    rowData[header] = self.cleanText(value)
                rows.append({
                    "sourceType": sourceType,
                    "filePath": filePath,
                    "rowNumber": rowNumber,
                    "rowData": rowData,
                })
            return rows
        finally:
            workbook.close()

    def readCsvRows(self, filePath, sourceType):
        """读取 CSV 文件并返回带来源行号的数据。"""
        rows = []
        with filePath.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            for rowNumber, row in enumerate(reader, start=2):
                rows.append({
                    "sourceType": sourceType,
                    "filePath": filePath,
                    "rowNumber": rowNumber,
                    "rowData": {
                        self.cleanText(key): self.cleanText(value)
                        for key, value in row.items()
                        if self.cleanText(key)
                    },
                })
        return rows

    def readSourceRows(self, filePath, sourceType):
        """根据文件后缀读取 Excel 或 CSV 数据。"""
        suffix = filePath.suffix.lower()
        if suffix == ".xlsx":
            return self.readExcelRows(filePath, sourceType)
        if suffix == ".csv":
            return self.readCsvRows(filePath, sourceType)
        raise ValueError(f"不支持的文件类型: {filePath}")

    def selectedSourceFiles(self):
        """根据配置返回本次要处理的公司/个人文件。"""
        sourceFiles = []
        if self.includeCompany:
            sourceFiles.append(("公司", self.companyFile))
        if self.includePerson:
            sourceFiles.append(("个人", self.personFile))
        return sourceFiles

    def sourceRule(self, sourceType):
        """返回公司或个人文件的字段读取规则。"""
        if sourceType == "个人":
            return {
                "nameHeaders": self.personNameHeaders,
                "licenseHeaders": self.personLicenseHeaders,
                "sourceHeaders": self.personSourceHeaders,
            }
        return {
            "nameHeaders": self.companyNameHeaders,
            "licenseHeaders": self.companyLicenseHeaders,
            "sourceHeaders": self.companySourceHeaders,
        }

    def emailBodyText(self):
        """生成纯文本邮件正文。"""
        if self.emailSignature:
            return self.emailMainBody + "\n\n\n" + self.emailSignature
        return self.emailMainBody

    def textToHtmlParagraphs(self, text):
        """把纯文本段落转换为 HTML 段落。"""
        paragraphs = []
        for paragraph in text.strip().split("\n\n"):
            lines = [html.escape(line) for line in paragraph.splitlines()]
            paragraphs.append("<p style=\"margin:0 0 14px 0;\">" + "<br>".join(lines) + "</p>")
        return "\n".join(paragraphs)

    def emailHtml(self):
        """生成带签名和内嵌 Logo 的 HTML 邮件正文。"""
        mainHtml = self.textToHtmlParagraphs(self.emailMainBody)
        websiteUrl = html.escape(self.websiteUrl)
        logoCid = html.escape(self.logoCid)
        return f"""<!doctype html>
<html>
<body style="font-family:Arial, Helvetica, sans-serif; font-size:14px; color:#111; line-height:1.45;">
{mainHtml}
<div style="margin-top:28px;">
  <p style="font-size:18px; font-weight:700; margin:0 0 22px 0;">Time2renew Support Team</p>
  <p style="margin:0 0 22px 0;"><strong>Website:</strong> <a href="{websiteUrl}" style="color:#111; text-decoration:underline;">time2renew</a></p>
  <img src="cid:{logoCid}" alt="Time2Renew" width="{self.logoWidth}" style="display:block; width:{self.logoWidth}px; height:auto; border:0;">
</div>
</body>
</html>"""

    def emailMessage(self, receiver):
        """构建一封 MIME 邮件，Logo 使用 CID 内嵌图片。"""
        if not self.senderEmail:
            raise ValueError("未配置发件邮箱")
        if not self.logoPath.exists():
            raise FileNotFoundError(f"邮件 Logo 图片不存在: {self.logoPath}")

        message = MIMEMultipart("related")
        message["From"] = formataddr((str(Header("Time2renew Support Team", "utf-8")), self.senderEmail))
        message["To"] = receiver
        message["Subject"] = Header(self.emailSubject, "utf-8").encode()

        alternative = MIMEMultipart("alternative")
        alternative.attach(MIMEText(self.emailBodyText(), "plain", "utf-8"))
        alternative.attach(MIMEText(self.emailHtml(), "html", "utf-8"))
        message.attach(alternative)

        with self.logoPath.open("rb") as imageFile:
            image = MIMEImage(imageFile.read(), _subtype="png")
        image.add_header("Content-ID", f"<{self.logoCid}>")
        image.add_header("Content-Disposition", "inline", filename=self.logoPath.name)
        message.attach(image)
        return message

    def smtpAuthCode(self):
        """返回 main.py 中固定维护的 SMTP 授权码。"""
        if self.smtpPassword == "__TIME2RENEW_SMTP_AUTH_CODE__":
            return ""
        return self.smtpPassword

    def runtimeText(self):
        """返回当前运行环境的中文名称。"""
        return "线上" if self.isOnline else "本机"

    def buildRecord(self, sourceRow, email):
        """把一行上游数据整理成一条邮件发送记录。"""
        rowData = sourceRow["rowData"]
        sourceType = sourceRow["sourceType"]
        rule = self.sourceRule(sourceType)
        return {
            "来源类型": sourceType,
            "运行环境": self.runtimeText(),
            "对象名称": self.getRowValue(rowData, rule["nameHeaders"]),
            "许可证号": self.getRowValue(rowData, rule["licenseHeaders"]),
            "邮箱": email,
            "邮件主题": self.emailSubject,
            "邮件正文": self.emailBodyText(),
            "采集状态": self.getRowValue(rowData, self.collectStatusHeaders),
            "搜索状态": self.getRowValue(rowData, self.searchStatusHeaders),
            "来源链接": self.combineValues(rowData, rule["sourceHeaders"]),
            "来源文件": sourceRow["filePath"].name,
            "来源行号": sourceRow["rowNumber"],
            "发送状态": "待发送",
            "发送结果": "",
            "失败原因": "",
            "备注": "",
        }

    def buildRecords(self, sourceFiles):
        """读取所有已选择文件并生成邮件发送记录。"""
        records = []
        warnings = []
        for sourceType, rawFilePath in sourceFiles:
            filePath = Path(str(rawFilePath))
            if not filePath.exists():
                warnings.append(f"文件不存在，已跳过: {filePath}")
                continue

            rowCount = 0
            emailCount = 0
            for sourceRow in self.readSourceRows(filePath, sourceType):
                rowCount += 1
                emailText = self.combineValues(sourceRow["rowData"], self.emailHeaders)
                emails = self.extractEmails(emailText)
                for email in emails:
                    emailCount += 1
                    records.append(self.buildRecord(sourceRow, email))
            self.log(f"{sourceType}文件读取完成: {rowCount} 行，生成邮件 {emailCount} 封")
        return records, warnings

    def sendRecords(self, records, executeSend):
        """根据开关真实发送邮件或只生成后台记录。"""
        summary = {
            "total": len(records),
            "emailTotal": len(records),
            "emailSent": 0,
            "emailFailed": 0,
        }

        if not executeSend:
            for record in records:
                record["发送状态"] = "待发送"
                record["发送结果"] = "已生成邮件记录，未执行发送"
            return summary

        if not records:
            return summary

        authCode = self.smtpAuthCode()
        if not authCode:
            for record in records:
                record["发送状态"] = "发送失败"
                record["发送结果"] = "邮件未发送"
                record["失败原因"] = "未配置 SMTP 授权码"
            summary["emailFailed"] = len(records)
            self.log("邮件发送失败: 未配置 SMTP 授权码")
            return summary

        server = smtplib.SMTP_SSL(self.smtpServer, self.smtpPort, timeout=30)
        try:
            server.login(self.smtpUser, authCode)
            for index, record in enumerate(records):
                receiver = record["邮箱"]
                try:
                    message = self.emailMessage(receiver)
                    server.sendmail(self.senderEmail, [receiver], message.as_bytes())
                    record["发送状态"] = "已发送"
                    record["发送结果"] = "邮件发送成功"
                    summary["emailSent"] += 1
                    self.log(f"邮件发送成功: {receiver}")
                except Exception as error:
                    record["发送状态"] = "发送失败"
                    record["发送结果"] = "邮件发送失败"
                    record["失败原因"] = str(error)
                    summary["emailFailed"] += 1
                    self.log(f"邮件发送失败: {receiver} - {error}")

                if self.sendWaitSecond > 0 and index < len(records) - 1:
                    time.sleep(self.sendWaitSecond)
        except Exception as error:
            for record in records:
                record["发送状态"] = "发送失败"
                record["发送结果"] = "邮件登录或连接失败"
                record["失败原因"] = str(error)
            summary["emailFailed"] = len(records)
            self.log(f"邮件登录或连接失败: {error}")
        finally:
            try:
                server.quit()
            except Exception:
                pass

        return summary

    def saveRecords(self, records):
        """保存后台邮件发送记录。"""
        self.outputFile.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Email Send Record"
        worksheet.append(self.outputHeaders)
        for record in records:
            worksheet.append([record.get(header, "") for header in self.outputHeaders])
        worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(str(self.outputFile))

    def run(self, executeSend=False, sourceFiles=None):
        """执行正式邮件推广流程。"""
        if sourceFiles is None:
            sourceFiles = self.selectedSourceFiles()

        self.log(f"运行环境: {self.runtimeText()}")
        records, warnings = self.buildRecords(sourceFiles)
        for warning in warnings:
            self.log(f"提示: {warning}")

        summary = self.sendRecords(records, executeSend)
        self.saveRecords(records)
        self.log(f"已保存后台邮件发送记录: {self.outputFile}")
        self.log(f"邮件任务数量: {len(records)}")
        return {
            "outputFile": str(self.outputFile),
            "records": records,
            "warnings": warnings,
            "summary": summary,
        }


if __name__ == "__main__":
    # 单文件调试配置：默认只生成后台发送记录，不真实发送邮件。
    config = {
        "includeCompany": True,
        "includePerson": True,
    }
    Main(config).run(executeSend=False)
