import warnings
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class Excel:
    def __init__(self, config, startDate, endDate):
        self.filePath = Path(config["file_path"])
        self.receiptFile = Path(config["receipt_file"]) if config["receipt_file"] else None
        self.startDate = startDate
        self.endDate = endDate
        self.filePath.mkdir(parents=True, exist_ok=True)
        warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

    # 沃尔玛表格文件处理
    def walmart(self):
        files = [file for file in self.filePath.glob("Walmart转账明细-*.xlsx")
                 if not file.name.startswith("~$") and "汇总完成" not in file.name]
        if not files:
            print("缺少Walmart转账明细文件")
            return

        download_file = max(files, key=lambda file: file.stat().st_mtime)
        walmart_data = pd.read_excel(download_file, usecols=["店铺", "付款周期结束时间", "金额"])
        walmart_data = walmart_data.dropna(subset=["付款周期结束时间"])
        walmart_data["金额"] = pd.to_numeric(walmart_data["金额"].astype(str)
                                                .str.replace(r"[^\d.-]", "", regex=True), errors="coerce")
        walmart_data = walmart_data.groupby(["店铺", "付款周期结束时间"], as_index=False)["金额"].sum()
        walmart_data["金额"] = walmart_data["金额"].round(2)
        walmart_data["平台"] = "沃尔玛"
        walmart_data["币种"] = "USD"
        summary_file = download_file.with_name(f"{download_file.stem}-汇总完成.xlsx")
        walmart_data.to_excel(summary_file, index=False)

        workbook = load_workbook(summary_file)
        worksheet = workbook.active
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = [28, 22, 14, 12, 10]
        for column, width in enumerate(widths, 1):
            worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = width
        border = Border(left=Side(style="thin", color="B7C9D6"), right=Side(style="thin", color="B7C9D6"),
                        top=Side(style="thin", color="B7C9D6"), bottom=Side(style="thin", color="B7C9D6"))
        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        colors = ["FFF200", "D9EAD3", "F9CB9C", "D9EAF7", "EADCF8"]
        stores = {}
        for row in range(2, worksheet.max_row + 1):
            store = str(worksheet.cell(row, 1).value)
            if store not in stores:
                stores[store] = colors[len(stores) % len(colors)]
            for column in range(1, 4):
                worksheet.cell(row, column).fill = PatternFill("solid", fgColor=stores[store])
            for column in range(4, 6):
                worksheet.cell(row, column).fill = PatternFill("solid", fgColor="E2F0D9")
            for column in range(1, 6):
                worksheet.cell(row, column).border = border
                worksheet.cell(row, column).alignment = Alignment(vertical="center")
            worksheet.cell(row, 2).number_format = "yyyy-mm-dd"
            worksheet.cell(row, 3).number_format = "0.00"
        workbook.save(summary_file)
        print(f"Walmart数据汇总完成：{summary_file}")
        return summary_file

    # 合并明细：处理已完成下载的4个表格文件(沃尔玛处理完成后的文件)-导出最后的对账结果
    def merge(self):
        platform_files = {
            "TikTok": "TikTok转账明细-*.xlsx",
            "eBay": "eBay转账明细-*.xlsx",
            "沃尔玛": "Walmart转账明细-*-汇总完成.xlsx",
            "亚马逊": "亚马逊转账明细-*.xlsx",
        }
        download_files = {}
        for platform, pattern in platform_files.items():
            files = [file for file in self.filePath.glob(pattern) if not file.name.startswith("~$")]
            if files:
                download_files[platform] = max(files, key=lambda file: file.stat().st_mtime)

        missing_files = [platform for platform in platform_files if platform not in download_files]
        if missing_files:
            print(f"缺少平台转账明细文件：{'、'.join(missing_files)}")
            return

        all_data = []
        for platform, platform_file in download_files.items():
            platform_data = pd.read_excel(platform_file, dtype=str)
            if platform == "TikTok":
                amount_column = "支付金额(兑换后)" if "支付金额(兑换后)" in platform_data else "支付金额（兑换后）"
                amount_text = platform_data[amount_column].fillna("").astype(str)
                bank_column = "银行信息" if "银行信息" in platform_data else "银行尾号"
                bank_tail = platform_data[bank_column].fillna("").str.extract(r"(\d{4})\D*$", expand=False).fillna("")
                merge_data = pd.DataFrame({
                    "店铺": platform_data["店铺"], "追踪编号": platform_data["支付ID"].fillna(""),
                    "转账时间": platform_data["支付成功时间"], "到账金额": amount_text,
                    "银行尾号": bank_tail, "平台": "TikTok", "币种": "USD",
                })
            elif platform == "eBay":
                amount_text = platform_data["支付金额"].fillna("").astype(str)
                merge_data = pd.DataFrame({
                    "店铺": platform_data["店铺"], "追踪编号": platform_data["发款编号"].fillna(""),
                    "转账时间": platform_data["支付日期"], "到账金额": amount_text,
                    "银行尾号": platform_data["账号后四位"].fillna(""), "平台": "eBay", "币种": "USD",
                })
            elif platform == "沃尔玛":
                amount_text = platform_data["金额"].fillna("").astype(str)
                merge_data = pd.DataFrame({
                    "店铺": platform_data["店铺"], "追踪编号": "",
                    "转账时间": platform_data["付款周期结束时间"], "到账金额": amount_text,
                    "银行尾号": "", "平台": "沃尔玛", "币种": "USD",
                })
            else:
                amount_text = platform_data["到账金额"].fillna("").astype(str)
                currency = platform_data["站点"].map({
                    "日本": "JPY", "美国": "USD", "英国": "GBP", "加拿大": "CAD", "墨西哥": "MXN",
                    "澳大利亚": "AUD", "印度": "INR", "德国": "EUR", "法国": "EUR",
                    "意大利": "EUR", "西班牙": "EUR",
                }).fillna("USD")
                merge_data = pd.DataFrame({
                    "店铺": platform_data["店铺"], "追踪编号": platform_data["追踪编号"],
                    "转账时间": platform_data["转账时间"], "到账金额": amount_text,
                    "银行尾号": platform_data["银行尾号"], "平台": "亚马逊", "币种": currency,
                })

            merge_data["追踪编号"] = merge_data["追踪编号"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)
            merge_data["银行尾号"] = merge_data["银行尾号"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)
            merge_data.loc[amount_text.str.contains("£", regex=False), "币种"] = "GBP"
            merge_data.loc[amount_text.str.contains("€", regex=False), "币种"] = "EUR"
            merge_data.loc[amount_text.str.contains("JP¥|JPY", regex=True), "币种"] = "JPY"
            merge_data.loc[amount_text.str.contains("CN¥|CNH|CNY", regex=True), "币种"] = "CNH"
            merge_data.loc[amount_text.str.contains(r"CA\$|CAD", regex=True), "币种"] = "CAD"
            merge_data["到账金额"] = pd.to_numeric(amount_text.str.replace(r"[^\d.-]", "", regex=True), errors="coerce")
            all_data.append(merge_data)

        result_data = pd.concat(all_data, ignore_index=True)
        result_data["转账时间"] = pd.to_datetime(result_data["转账时间"], format="mixed", errors="coerce")
        result_data = result_data.dropna(subset=["转账时间", "到账金额"])
        result_data = result_data.sort_values("转账时间", ascending=False)
        result_data["到账状态"] = "暂无匹配"

        columns = ["店铺", "追踪编号", "转账时间", "到账金额", "银行尾号", "平台", "币种", "到账状态"]
        merge_file = self.filePath / f"各平台回款明细表-{self.startDate[:4]}年{self.startDate[5:7]}月.xlsx"
        merge_data = result_data.copy()
        merge_data["转账时间"] = merge_data["转账时间"].dt.strftime("%Y年%m月%d日")
        merge_data[columns].to_excel(merge_file, index=False)
        result_files = [merge_file]

        if self.receiptFile and self.receiptFile.exists():
            statement_data = pd.read_excel(self.receiptFile, dtype=str)
            statement_data["收款时间"] = pd.to_datetime(statement_data["收款时间"], format="mixed", errors="coerce")
            statement_data["收款币种"] = statement_data["收款币种"].fillna("").str.strip().str.upper()
            statement_data["收款金额"] = pd.to_numeric(statement_data["收款金额"].fillna("")
                                                         .str.replace(r"[^\d.-]", "", regex=True), errors="coerce")
            statement_data["已匹配"] = False
            for row in result_data.index:
                transfer_date = result_data.at[row, "转账时间"].normalize()
                matched = statement_data[
                    (~statement_data["已匹配"])
                    & (statement_data["收款币种"] == result_data.at[row, "币种"])
                    & (statement_data["收款金额"].round(2) == round(result_data.at[row, "到账金额"], 2))
                    & (statement_data["收款时间"].dt.normalize() >= transfer_date)
                    & (statement_data["收款时间"].dt.normalize() <= transfer_date + pd.Timedelta(days=5))
                ]
                if not matched.empty:
                    statement_data.at[matched.index[0], "已匹配"] = True
                    result_data.at[row, "到账状态"] = "已到账"

            result_file = self.filePath / f"各平台回款明细表（对账结果）-{self.startDate[:4]}年{self.startDate[5:7]}月.xlsx"
            result_data["转账时间"] = result_data["转账时间"].dt.strftime("%Y年%m月%d日")
            result_data[columns].to_excel(result_file, index=False)
            result_files.append(result_file)

        for result_file in result_files:
            workbook = load_workbook(result_file)
            worksheet = workbook.active
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            widths = [28, 38, 18, 16, 22, 12, 10, 14]
            border = Border(left=Side(style="thin", color="A9C7B0"), right=Side(style="thin", color="A9C7B0"),
                            top=Side(style="thin", color="A9C7B0"), bottom=Side(style="thin", color="A9C7B0"))
            for column, width in enumerate(widths, 1):
                worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = width
            for row_number, row in enumerate(worksheet.iter_rows(), 1):
                worksheet.row_dimensions[row_number].height = 24 if row_number == 1 else 22
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="EAF4EA" if row_number % 2 == 0 else "FFFFFF")
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                    cell.font = Font(name="微软雅黑", size=10)
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor="C6E0B4")
                cell.font = Font(name="微软雅黑", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, 2).number_format = "@"
                worksheet.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
                worksheet.cell(row, 4).number_format = "0.00"
                worksheet.cell(row, 4).alignment = Alignment(horizontal="right", vertical="center")
                worksheet.cell(row, 5).number_format = "@"
                for column in range(6, 9):
                    worksheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
            workbook.save(result_file)
            print(f"各平台回款明细表已生成：{result_file}")
        return result_files[-1]
