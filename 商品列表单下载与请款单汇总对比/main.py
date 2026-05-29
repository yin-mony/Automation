from DrissionPage import ChromiumPage
from test import SaihuERPLogin
import time
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def run(download_dir=None):
    page = ChromiumPage()
    login = SaihuERPLogin(page)
    file_path = Path(download_dir) if download_dir else Path(r"C:\Users\admin\Desktop\采购单与商品单汇总对比")
    file_path.mkdir(parents=True, exist_ok=True)

    ok = login.login(prefer_entry_check=True)
    if not ok:
        print("登录失败，终止后续操作")
        return

    # 到这里就是已登录状态，直接继续操作 page
    # 先进行采购单下载
    page.ele('x://div/ul/li/span[text()="采购"]', timeout=8).click()
    time.sleep(1)
    page.ele('x://a[text()="采购单"]', timeout=8).click()
    time.sleep(1)
    host = page.ele('x://wujie-app', timeout=5)
    shadow = host.shadow_root
    shadow.ele('x://div[@title="导出"]').click()
    time.sleep(1)
    shadow.ele('x://span[text()="自动化检查采购单价"]', timeout=10).click()
    time.sleep(1)
    page.ele('x://span[text()="下载文件已完成"]', timeout=800).click()
    time.sleep(2)
    download = page.ele('x://span[contains(text(), "立即下载")]').click.to_download(save_path=file_path,timeout=20,rename="采购单下载")
    download.wait()
    print("采购单下载完成")
    time.sleep(2)

    # 进行商品列表单下载
    page.ele('x://div/ul/li/span[text()="商品"]', timeout=8).click()
    time.sleep(1)
    page.ele('x://a[text()="商品列表"]', timeout=8).click()
    time.sleep(1)
    # 点击导出按钮
    page.ele('x://div[@class="dc"]/div[@title="导出"]').click()
    time.sleep(1)
    page.ele('x://div[@aria-label="导出"]//span[text()="确定"]', timeout=10).click()
    time.sleep(1)
    page.ele('x://span[text()="下载文件已完成"]', timeout=800).click()
    time.sleep(2)
    download = page.ele('x://span[contains(text(), "立即下载")]').click.to_download(save_path=file_path,timeout=20,rename="商品列表单下载")
    download.wait()
    print("商品列表单下载完成")
    excel_file(file_path)


def excel_file(download_dir=None):
    base_dir = Path(download_dir) if download_dir else Path(r"C:\Users\admin\Desktop\采购单与商品单汇总对比")
    base_dir.mkdir(parents=True, exist_ok=True)
    # 打开下载完成的采购单与商品列表单文件
    purchase_file = base_dir / "采购单下载.xlsx"
    product_list_file = base_dir / "商品列表单下载.xlsx"
    purchase_df = pd.read_excel(purchase_file)
    product_list_df = pd.read_excel(product_list_file)
    print("采购单文件、商品列表单文件打开完成")

    # 采购单：筛选请款状态不为“未请款”
    purchase_df = purchase_df[purchase_df["请款状态"].astype(str).str.strip() != "未请款"].copy()
    print(f"采购单筛选后剩余 {len(purchase_df)} 行数据")

    # 商品列表单：仅提取 SKU、采购单价，并按采购单 SKU 过滤
    purchase_df["SKU"] = purchase_df["SKU"].fillna("").astype(str).str.strip()
    product_list_df["SKU"] = product_list_df["SKU"].fillna("").astype(str).str.strip()
    product_list_df = product_list_df[product_list_df["SKU"].isin(purchase_df["SKU"])].copy()
    product_list_df = product_list_df[["SKU", "采购单价"]].copy()
    print(f"商品列表单筛选后剩余 {len(product_list_df)} 行数据（仅SKU和采购单价）")

    # 同一 SKU 只保留一条商品列表单价格记录，避免重复 merge
    product_list_df = product_list_df.drop_duplicates(subset=["SKU"], keep="first")
    product_list_df = product_list_df.rename(
        columns={
            "SKU": "商品列表单SKU",
            "采购单价": "商品列表单采购单价",
        }
    )

    # 汇总结果：保留采购单筛选后的 SKU/单号/采购单价/请款状态
    purchase_export_df = purchase_df[["SKU", "采购单号", "采购单价", "请款状态"]].copy()
    export_df = purchase_export_df.merge(
        product_list_df,
        left_on="SKU",
        right_on="商品列表单SKU",
        how="left",
    )

    # 调整输出字段顺序（两边字段明确区分）
    export_df = export_df[
        ["SKU", "采购单号", "采购单价", "请款状态", "商品列表单SKU", "商品列表单采购单价"]
    ]

    # 高亮规则：采购单价 > 商品列表单采购单价（含商品列表单为空）
    purchase_price_series = pd.to_numeric(export_df["采购单价"], errors="coerce")
    product_price_series = pd.to_numeric(export_df["商品列表单采购单价"], errors="coerce")
    export_df["_需高亮"] = purchase_price_series.notna() & (
        product_price_series.isna() | (purchase_price_series > product_price_series)
    )

    # 输出单工作表文件
    output_file = base_dir / "采购单与商品单汇总对比.xlsx"
    export_df.drop(columns=["_需高亮"]).to_excel(output_file, index=False)

    # 对应行亮色标注（整行）
    wb = load_workbook(output_file)
    ws = wb.active
    highlight_fill = PatternFill(fill_type="solid", start_color="FFF59D", end_color="FFF59D")
    for idx, need_highlight in enumerate(export_df["_需高亮"].tolist(), start=2):
        if need_highlight:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=idx, column=col).fill = highlight_fill
    wb.save(output_file)

    print(f"已生成文件: {output_file}")

if __name__ == "__main__":
    run()
    # excel_file()