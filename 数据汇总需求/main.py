from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
import sys


def _find_column(headers, name):
    """按列名查找索引，charged_amount 兼容透视表列名如「求和项:charged_amount」。"""
    name_lower = name.lower()
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str.lower() == name_lower:
            return i
    if name_lower == 'charged_amount':
        for i, h in enumerate(headers):
            if h is not None and 'charged_amount' in str(h).lower():
                return i
    raise ValueError(f'未找到列「{name}」，当前表头: {[h for h in headers if h]}')


def _get_source_worksheet(wb):
    """始终从含原始明细的 Sheet1 读取，避免误读已生成的 Sheet2。"""
    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
        headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
        try:
            _find_column(headers, 'myp_order_id')
            _find_column(headers, 'msku')
            _find_column(headers, 'charged_amount')
            return ws, headers
        except ValueError:
            pass

    for name in wb.sheetnames:
        if name.lower() == 'sheet2':
            continue
        ws = wb[name]
        headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
        try:
            _find_column(headers, 'myp_order_id')
            _find_column(headers, 'msku')
            _find_column(headers, 'charged_amount')
            return ws, headers
        except ValueError:
            continue

    raise ValueError('未找到包含 myp_order_id、msku、charged_amount 的数据表（请使用原始导出表或 Sheet1）')


class Excel_file:
    """
    用法：
        excel = Excel_file("C:/data/表.xlsx")
        excel.process_multiple_files()
        excel = Excel_file(["C:/a.xlsx", "C:/b.xlsx"])
        excel.process_multiple_files()
    """

    def __init__(self, path):
        # path：单个 xlsx 路径，或路径列表
        if isinstance(path, (list, tuple)):
            self.paths = [str(p) for p in path if str(p).strip()]
        else:
            self.paths = [str(path)] if str(path).strip() else []

    def process_excel_file(self, file_path):
        wb = load_workbook(file_path)
        ws, headers = _get_source_worksheet(wb)

        data = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            data.append(list(row))

        myp_order_id_col = _find_column(headers, 'myp_order_id')
        msku_col = _find_column(headers, 'msku')
        charged_amount_col = _find_column(headers, 'charged_amount')

        current_order_id = None
        for row in data:
            order_id = row[myp_order_id_col]
            if order_id is not None and str(order_id).strip():
                current_order_id = order_id
            else:
                row[myp_order_id_col] = current_order_id

        order_groups = {}
        for row in data:
            order_id = row[myp_order_id_col]
            if order_id not in order_groups:
                order_groups[order_id] = []
            order_groups[order_id].append(row)

        def order_has_multiple_msku(rows):
            return len({r[msku_col] for r in rows}) > 1

        kept_order_ids = {
            oid for oid, rows in order_groups.items() if order_has_multiple_msku(rows)
        }
        filtered_data = [row for row in data if row[myp_order_id_col] in kept_order_ids]

        new_wb = Workbook()
        new_ws1 = new_wb.active
        new_ws1.title = 'Sheet1'
        new_ws1.append(headers)
        for row in filtered_data:
            new_ws1.append(row)

        new_ws2 = new_wb.create_sheet(title='sheet2')

        header_row = ['myp_order_id', 'msku', 'charged_amount']
        new_ws2.append(header_row)

        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        total_font = Font(bold=True)
        right_align = Alignment(horizontal='right')

        for col in range(1, len(header_row) + 1):
            new_ws2.cell(row=1, column=col).fill = header_fill

        overall_total = 0
        for order_id, rows in order_groups.items():
            if order_id not in kept_order_ids:
                continue

            msku_totals = {}
            msku_order = []
            for row in rows:
                msku = row[msku_col]
                amount = row[charged_amount_col] if row[charged_amount_col] else 0
                if msku not in msku_totals:
                    msku_totals[msku] = 0
                    msku_order.append(msku)
                msku_totals[msku] += amount

            first_row = True
            for msku in msku_order:
                total_amount = msku_totals[msku]
                if first_row:
                    new_ws2.append([order_id, msku, total_amount])
                    first_row = False
                else:
                    new_ws2.append(['', msku, total_amount])
                overall_total += total_amount

        new_ws2.append(['', '', ''])
        new_ws2.append(['总计', '', overall_total])
        total_row = new_ws2.max_row
        for col in range(1, 4):
            cell = new_ws2.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            if col == 3:
                cell.alignment = right_align

        new_wb.active = new_ws1
        new_wb.save(file_path)

        return len(filtered_data)

    def process_multiple_files(self, file_paths=None):
        if file_paths is None:
            file_paths = self.paths
        results = {}
        for file_path in file_paths:
            if os.path.exists(file_path):
                try:
                    count = self.process_excel_file(file_path)
                    results[file_path] = {'success': True, 'count': count}
                    print(f"处理完成: {file_path}")
                    print(f"记录数: {count}")
                except Exception as e:
                    err = str(e)
                    print(f"处理失败 {file_path}: {err}")
                    results[file_path] = {'success': False, 'error': err}
            else:
                print(f"文件不存在: {file_path}")
                results[file_path] = {'success': False, 'error': '文件不存在'}
        return results


if __name__ == "__main__":
    

    config = {
        "file_paths": [
            r"C:\RPA流程\数据汇总需求\test.xlsx",
        ],
    }

    # 命令行可覆盖：python main.py <文件1> [文件2 ...]
    if len(sys.argv) >= 2:
        config["file_paths"] = sys.argv[1:]

    if not config["file_paths"]:
        raise ValueError("请在 config 中填写 file_paths（至少一个 Excel 路径）。")

    excel = Excel_file(config["file_paths"])
    results = excel.process_multiple_files()
    for file_path, result in results.items():
        if result.get('success'):
            print(f"处理完成: {file_path}，记录数: {result['count']}")
        else:
            print(f"处理失败: {file_path}，原因: {result.get('error', '未知错误')}")
