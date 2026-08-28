import json
import os
import re
import time

from DrissionPage import ChromiumPage
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels


class Factory1688:
    """按图片在1688筛选真实工厂，并导出整理后的Excel结果。"""

    def __init__(self, config):
        """初始化浏览器、输入目录、输出目录和固定筛选规则。"""
        # 复用同一个浏览器实例完成首页上传、结果筛选和商品详情解析。
        self.page = ChromiumPage()

        # 输入目录存放待搜索图片，目录名称会作为材质写入结果表。
        self.folderPath = config.get("folderPath", "")

        # 输出目录为空时，结果表保存到图片目录的上一级目录。
        self.outputDir = config.get("outputDir", "")

        # 未指定文件名时使用固定的默认结果表名称。
        self.outputFileName = config.get("outputFileName", "1688筛选工厂结果.xlsx")

        # 等待人工调整主体范围并确认；超时后终止当前流程，防止使用错误选区。
        self.cropTimeout = config.get("cropTimeout", 300)

        # 1688首页用于发起以图搜款，推荐接口用于确认页面结果已经刷新。
        self.homeUrl = "https://www.1688.com/"
        self.recommendApiName = "mtop.relationrecommend.wirelessrecommend.recommend"

        # 公司名称包含以下经营性质关键词时，不作为目标工厂写入结果表。
        self.excludeKeywords = ("电子商务", "贸易", "商贸")

        # 数值越小优先级越高，导出时按P0、P1、P2依次排列。
        self.priorityOrder = {
            "P0": 0,
            "P1": 1,
            "P2": 2
        }

    def getTasks(self):
        """读取图片目录，并把目录名作为当前批次的材质名称。"""
        if not self.folderPath or not os.path.exists(self.folderPath):
            print("路径不存在")
            return []

        material = os.path.basename(self.folderPath)
        imageExtensions = (".jpg", ".jpeg", ".png", ".bmp", ".webp")

        imageFiles = [
            fileName for fileName in os.listdir(self.folderPath)
            if fileName.lower().endswith(imageExtensions)
        ]

        if not imageFiles:
            print("没有找到图片")
            return []

        tasks = []

        for imageName in imageFiles:
            tasks.append({
                "material": material,
                "imagePath": os.path.join(self.folderPath, imageName)
            })

        print(f"检测到材质：{material}")
        print(f"共找到 {len(tasks)} 张图片")

        return tasks

    def crop(self, page):
        """在搜索结果页确认主体框选，并等待框选后的结果刷新。"""
        cropButton = page.ele(
            'x://*[normalize-space(text())="框选主体"]',
            timeout=self.cropTimeout
        )

        if not cropButton:
            raise RuntimeError("搜索结果页未找到“框选主体”按钮")

        print("已打开主体裁剪界面，请手动调整范围并点击确认...")

        # 框选会独立刷新一次推荐结果，先消费这次请求，避免与后续筛选请求混淆。
        page.listen.start(self.recommendApiName)

        try:
            cropButton.click()
            packet = page.listen.wait(
                timeout=self.cropTimeout,
                fit_count=False
            )

            if not packet:
                raise TimeoutError(
                    f"等待人工确认主体框选超时：{self.cropTimeout}秒"
                )

        finally:
            page.listen.stop()

        print("检测到框选结果刷新，继续执行后续筛选")

    def search(self, task):
        """上传单张图片并抓取真实工厂认证后的商品和店铺信息。"""
        material = task["material"]
        imagePath = task["imagePath"]

        self.page.get(self.homeUrl)
        print("开始自动化批量操作...")

        uploadElement = self.page.ele("x://div[text()='以图搜款']")
        uploadElement.click.to_upload(imagePath)

        self.page.ele('x://*[text()="搜索图片"]').click()
        time.sleep(2)

        # 不覆盖self.page，防止下一次任务出现MixTab问题。
        searchPage = self.page.latest_tab

        self.crop(searchPage)
        searchPage.listen.start(self.recommendApiName)

        # 先勾选真实工厂认证，再用材质关键词缩小结果范围。
        searchPage.ele(
            'x://div/span[text()="真实工厂认证"]'
        ).click(by_js=True)

        searchPage.ele(
            'x://input[starts-with(@class, "searchFormInput")]'
        ).input(f"{material}\n", clear=True)

        time.sleep(2)

        for packet in searchPage.listen.steps():
            if "relationrecommend" not in packet.url:
                continue

            print("捕获接口")

            try:
                body = packet.response.body

                # 1688接口返回的是JSONP，需要去掉外层函数名后再解析。
                jsonText = re.sub(r"^.*?\(", "", body)
                jsonText = jsonText.rsplit(")", 1)[0]
                data = json.loads(jsonText)

                offers = data["data"]["data"]["OFFER"]["items"]
                print(f"共找到 {len(offers)} 个商品")

            except Exception as error:
                print("接口数据解析失败：", error)
                return []

            results = []

            # 复用一个详情页标签页，避免每个商品都新开页面。
            detailPage = self.page.new_tab("about:blank")

            for index, item in enumerate(offers, start=1):
                product = item.get("data", {})
                productUrl = product.get("linkUrl", "")

                shopInfo = product.get("shopAddition", {})
                shopUrl = shopInfo.get("shopLinkUrl", "")
                companyName = shopInfo.get("text", "").strip()

                print("----------------------")
                print(f"第 {index}/{len(offers)} 个商品")
                print("公司：", companyName)
                print("商品链接：", productUrl)
                print("店铺链接：", shopUrl)

                if not productUrl:
                    print("商品链接为空，跳过")
                    continue

                if any(keyword in companyName for keyword in self.excludeKeywords):
                    print("包含排除关键词，已排除")
                    continue

                try:
                    detailPage.get(productUrl)
                    time.sleep(1.5)

                    level = "P2"

                    # 店铺图标中的i1/i2/i4用于区分源头旗舰、实力商家和超级工厂。
                    icons = detailPage.eles('xpath://img[@alt="店铺图标"]')

                    for icon in icons:
                        src = (icon.attr("src") or "").lower()

                        if not src:
                            continue

                        print("店铺图标：", src)

                        if "/i4/" in src:
                            level = "P0"
                            break

                        if "/i1/" in src or "/i2/" in src:
                            level = "P1"

                    print("等级：", level)

                    results.append({
                        "材质": material,
                        "图片路径": imagePath,
                        "等级": level,
                        "公司名称": companyName,
                        "商品链接": productUrl,
                        "店铺链接": shopUrl
                    })

                except Exception as error:
                    print("详情页解析失败：", error)

            try:
                detailPage.close()
            except Exception:
                pass

            print(
                f"{os.path.basename(imagePath)}处理完成，"
                f"共获得{len(results)}条有效数据"
            )

            return results

        return []

    def export(self, tasks, allResults):
        """新建Excel结果表，并在同一流程内完成清洗、合并、排版和保存。"""
        print("\n========== 开始最终写入Excel ==========")

        groups = []

        for task in tasks:
            imagePath = task["imagePath"]

            # 只保留当前图片对应的数据，并再次过滤不符合工厂筛选要求的公司名称。
            rawResults = [
                item for item in allResults
                if item["图片路径"] == imagePath
                and not any(
                    keyword in item["公司名称"]
                    for keyword in self.excludeKeywords
                )
            ]

            uniqueResults = []
            seenKeys = set()

            for item in rawResults:
                uniqueKey = (
                    item["公司名称"].strip(),
                    item["店铺链接"].strip(),
                    item["商品链接"].strip()
                )

                if uniqueKey in seenKeys:
                    continue

                seenKeys.add(uniqueKey)
                uniqueResults.append(item)

            # 记录同等级、同公司、同门店第一次出现的位置，排序后仍尽量贴近1688原顺序。
            companyOrder = {}
            orderIndex = 0

            for item in uniqueResults:
                groupKey = (
                    item["等级"],
                    item["公司名称"],
                    item["店铺链接"]
                )

                if groupKey not in companyOrder:
                    companyOrder[groupKey] = orderIndex
                    orderIndex += 1

            uniqueResults.sort(
                key=lambda item: (
                    self.priorityOrder.get(item["等级"], 99),
                    companyOrder.get(
                        (item["等级"], item["公司名称"], item["店铺链接"]),
                        999999
                    )
                )
            )

            groups.append({
                "material": task["material"],
                "imagePath": imagePath,
                "results": uniqueResults
            })

            print(
                f"{os.path.basename(imagePath)}："
                f"原始{len(rawResults)}条，去重后{len(uniqueResults)}条"
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"

        headers = [
            "主图（自填）",
            "材质（自填）",
            "1688链接",
            "工厂名称",
            "门店链接",
            "优先级排序"
        ]
        thinSide = Side(style="thin", color="000000")
        thinBorder = Border(
            left=thinSide,
            right=thinSide,
            top=thinSide,
            bottom=thinSide
        )

        # 新建工作簿后手动设置表头、列宽和固定样式，替代原来的模板文件。
        sheet.append(headers)
        sheet.row_dimensions[1].height = 24

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thinBorder

        columnWidths = {
            "A": 18,
            "B": 18,
            "C": 42,
            "D": 30,
            "E": 38,
            "F": 12
        }

        for columnName, width in columnWidths.items():
            sheet.column_dimensions[columnName].width = width

        currentRow = 2

        for group in groups:
            material = group["material"]
            imagePath = group["imagePath"]
            results = group["results"]
            groupStartRow = currentRow

            if not results:
                sheet.cell(currentRow, 2, material)
                sheet.row_dimensions[currentRow].height = 90
                currentRow += 1
            else:
                for item in results:
                    sheet.cell(currentRow, 3, item["商品链接"])
                    sheet.cell(currentRow, 4, item["公司名称"])
                    sheet.cell(currentRow, 5, item["店铺链接"])
                    sheet.cell(currentRow, 6, item["等级"])
                    sheet.row_dimensions[currentRow].height = 30
                    currentRow += 1

            groupEndRow = currentRow - 1

            # A/B列按图片分组纵向合并，图片和材质只显示一次。
            if groupEndRow > groupStartRow:
                sheet.merge_cells(
                    start_row=groupStartRow,
                    start_column=1,
                    end_row=groupEndRow,
                    end_column=1
                )
                sheet.merge_cells(
                    start_row=groupStartRow,
                    start_column=2,
                    end_row=groupEndRow,
                    end_column=2
                )

            materialCell = sheet.cell(groupStartRow, 2, material)
            materialCell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # D/E列把同等级、同公司、同门店的多条商品链接聚合到一个视觉分组。
            relativeIndex = 0

            while relativeIndex < len(results):
                firstItem = results[relativeIndex]
                companyName = firstItem["公司名称"]
                shopUrl = firstItem["店铺链接"]
                level = firstItem["等级"]
                mergeStartRow = groupStartRow + relativeIndex
                mergeEndIndex = relativeIndex

                while mergeEndIndex + 1 < len(results):
                    nextItem = results[mergeEndIndex + 1]

                    if (
                        nextItem["公司名称"] == companyName
                        and nextItem["店铺链接"] == shopUrl
                        and nextItem["等级"] == level
                    ):
                        mergeEndIndex += 1
                    else:
                        break

                mergeEndRow = groupStartRow + mergeEndIndex

                if mergeEndRow > mergeStartRow:
                    sheet.merge_cells(
                        start_row=mergeStartRow,
                        start_column=4,
                        end_row=mergeEndRow,
                        end_column=4
                    )
                    sheet.merge_cells(
                        start_row=mergeStartRow,
                        start_column=5,
                        end_row=mergeEndRow,
                        end_column=5
                    )

                # 合并单元格的内容必须写在左上角单元格。
                sheet.cell(mergeStartRow, 4, companyName)
                sheet.cell(mergeStartRow, 5, shopUrl)
                relativeIndex = mergeEndIndex + 1

            try:
                absImagePath = os.path.abspath(imagePath)

                if os.path.exists(absImagePath):
                    picture = Image(absImagePath)

                    # A列宽18约等于126像素，保留左右各5像素并按原比例缩放。
                    columnWidthPixels = 126
                    maxWidthPixels = columnWidthPixels - 10

                    if picture.width > maxWidthPixels:
                        scale = maxWidthPixels / picture.width
                        picture.width = int(picture.width * scale)
                        picture.height = int(picture.height * scale)

                    horizontalOffset = (columnWidthPixels - picture.width) / 2
                    marker = AnchorMarker(
                        col=0,
                        colOff=pixels_to_EMU(horizontalOffset),
                        row=groupStartRow - 1,
                        rowOff=pixels_to_EMU(points_to_pixels(5))
                    )
                    picture.anchor = OneCellAnchor(
                        _from=marker,
                        ext=XDRPositiveSize2D(
                            pixels_to_EMU(picture.width),
                            pixels_to_EMU(picture.height)
                        )
                    )
                    sheet.add_image(picture)

            except Exception as error:
                print(f"图片插入失败：{imagePath}，{error}")

        lastDataRow = currentRow - 1

        if lastDataRow >= 2:
            for row in sheet.iter_rows(
                min_row=2,
                max_row=lastDataRow,
                min_col=1,
                max_col=6
            ):
                for cell in row:
                    cell.border = thinBorder
                    cell.alignment = Alignment(vertical="center")

            for row in range(2, lastDataRow + 1):
                sheet.cell(row, 3).font = Font(size=9)
                sheet.cell(row, 3).alignment = Alignment(
                    vertical="center",
                    wrap_text=False
                )
                sheet.cell(row, 4).alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )
                sheet.cell(row, 5).font = Font(size=9)
                sheet.cell(row, 5).alignment = Alignment(
                    vertical="center",
                    wrap_text=False
                )
                sheet.cell(row, 6).alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        if self.outputDir:
            outputDir = os.path.abspath(self.outputDir)
        else:
            outputDir = os.path.dirname(os.path.abspath(self.folderPath))

        os.makedirs(outputDir, exist_ok=True)
        outputPath = os.path.join(outputDir, self.outputFileName)
        workbook.save(outputPath)

        cleanCount = sum(len(group["results"]) for group in groups)

        print("\n========================================")
        print("Excel最终写入完成")
        print("文件：", outputPath)
        print("原始抓取：", len(allResults))
        print("清洗后写入：", cleanCount)
        print("========================================")

    def run(self):
        """执行图片读取、1688筛选、结果导出三个主流程。"""
        tasks = self.getTasks()

        if not tasks:
            print("没有需要处理的图片")
            return

        allResults = []

        for index, task in enumerate(tasks, start=1):
            print("\n========================================")
            print(
                f"开始处理第{index}/{len(tasks)}张图片："
                f"{os.path.basename(task['imagePath'])}"
            )
            print("材质：", task["material"])
            print("========================================")

            results = self.search(task)

            if results:
                allResults.extend(results)

            print(f"当前累计抓取：{len(allResults)}条")

        print("\n所有图片搜索完成，开始统一整理Excel...")
        self.export(tasks, allResults)


if __name__ == "__main__":
    config = {
        # 待搜索图片目录，目录名会作为材质名称。
        "folderPath": "./file/不锈钢镀金",
        # 结果表输出目录；留空时输出到图片目录的上一级目录。
        "outputDir": "",
        # 最终导出的Excel文件名。
        "outputFileName": "1688筛选工厂结果.xlsx",
        # 等待人工完成主体框选并确认的最长秒数。
        "cropTimeout": 300
    }

    factory = Factory1688(config)
    factory.run()
