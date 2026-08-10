
import time
import threading
import os
from DrissionPage import ChromiumPage,Chromium
import pandas as pd
import psutil
from YidekeLogin import Specification
from openpyxl import load_workbook
import re
class Automation():
    def __init__(self,config):
        self.username = config["username"]
        self.password = config['password']
        self.ip = config['ip']
        self.port = config['port']
        self.file = config['file']
        self.StorePassword = config['StorePassword']
        self.country = config['country']
        self.station = config['station']
        self.mode = config['mode']
        # self.mode = True

    def kill_edecker(self, exclude_pid):
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                pid = proc.info['pid']
                name = proc.info['name']
                if name and name.lower() == 'edecker.exe':
                    if pid != exclude_pid:
                        proc.kill()
            except:
                pass

    def run_edecker_automation(self, port=9222):
        """
        全部启动店铺
        :param port:
        :return:
        """
        browser = Chromium(port)
        tab = browser.latest_tab
        # buttons = tab.eles("t:button@@text()=访问")
        # for btn in buttons:
        #     btn.click()
        #     time.sleep(3)
        time.sleep(4)
        if self.station:
            tab.ele(
                f"x://div[contains(@class,'shop-item')][.//span[contains(text(),'{self.station}')] and .//div[text()='{self.ip}']]//button").click()
        else:
            tab.ele(f'x://div[text()="{self.ip}"]/following-sibling::button').click()

        time.sleep(7)

        self.kill_edecker(browser.process_id)
        time.sleep(1)
        tab.refresh()
        time.sleep(3)

    def start_edecker(self, ip: str, port: int):
        import subprocess
        from pathlib import Path

        base = Path.home() / "AppData/Local/eDecker6"
        exe_path = base / "Application/edecker.exe"
        profiles_path = base / "Profiles"

        print("EXE:", exe_path, exe_path.exists())
        print("Profiles dir exists:", profiles_path.exists())

        if not exe_path.exists():
            raise FileNotFoundError(f"找不到 exe: {exe_path}")

        if not profiles_path.exists():
            raise FileNotFoundError(f"找不到 profiles 目录: {profiles_path}")

        ip_dot = ip
        ip_underline = ip.replace('.', '_')

        all_profiles = list(profiles_path.iterdir())
        print("所有 profile:")
        for p in all_profiles:
            print(" -", p.name)

        candidates = [
            p for p in all_profiles
            if p.is_dir() and (ip_dot in p.name or ip_underline in p.name)
        ]

        if not candidates:
            raise Exception(f"未找到 IP={ip} 的 profile")

        latest = max(candidates, key=lambda p: p.stat().st_mtime)

        print("使用 profile:", latest)

        cmd = [
            str(exe_path),
            f'--user-data-dir={latest}',
            '--no-sandbox',
            f'--remote-debugging-port={port}'
        ]

        print("启动命令:")
        print(" ".join(cmd))

        try:
            subprocess.Popen(cmd, cwd=str(base))
            print("启动成功（已发起进程）")
        except Exception as e:
            print("启动失败:", e)
            raise

    # 抓取模式
    def Grabbing(self,page):
        from pathlib import Path
        export_suffixes = ('.csv', '.xlsx', '.xls')
        temp_suffixes = ('.crdownload', '.tmp')
        table_js = r"""
        return (function() {
          const rows = Array.from(document.querySelectorAll('table tbody tr'))
            .filter(row => (row.innerText || '').trim());
          const rowTexts = rows.slice(0, 8).map(row => (row.innerText || '').replace(/\s+/g, ' ').trim());
          return {rowCount: rows.length, signature: rowTexts.join(' | ')};
        })();
        """

        page.get('https://sellercentral.amazon.com/cu/case-lobby')

        table_state = None
        for _ in range(45):
            table_state = page.run_js(table_js)
            if table_state and table_state.get('rowCount', 0) > 0:
                break
            time.sleep(1)
        else:
            raise RuntimeError("管理支持问题单表格未加载")

        run_dir = Path(self.file).parent / "导出分页" / time.strftime("%Y%m%d_%H%M%S")
        run_dir.mkdir(parents=True, exist_ok=True)
        page.run_cdp('Browser.setDownloadBehavior', behavior='allow', downloadPath=str(run_dir), eventsEnabled=True)
        print(f"本次导出目录: {run_dir}")

        page.run_js(r"""
        return (function() {
          const dropdown = document.querySelector('[data-test-tag="results-per-page"] kat-dropdown');
          if (!dropdown) return false;
          dropdown.value = '50';
          dropdown.setAttribute('value', '50');
          const detail = {value: '50', selectedOption: {value: '50'}};
          dropdown.dispatchEvent(new CustomEvent('change', {detail, bubbles: true, composed: true}));
          dropdown.dispatchEvent(new CustomEvent('kat-change', {detail, bubbles: true, composed: true}));
          dropdown.dispatchEvent(new CustomEvent('katChange', {detail, bubbles: true, composed: true}));
          return true;
        })();
        """)
        for _ in range(20):
            table_state = page.run_js(table_js)
            page_size = page.run_js(
                "return document.querySelector('kat-pagination') && document.querySelector('kat-pagination').getAttribute('items-per-page')"
            )
            if page_size == '50' and table_state and table_state.get('rowCount', 0) > 0:
                print("已切换为每页 50 条")
                break
            time.sleep(1)

        exported_files = []
        seen_signatures = set()
        page_no = 1
        max_pages = 500

        while page_no <= max_pages:
            state = page.run_js(table_js)
            if not state or state.get('rowCount', 0) <= 0:
                print(f"第 {page_no} 页表格未加载，终止抓取")
                break

            signature = state.get('signature', '')
            if signature and signature in seen_signatures:
                print(f"第 {page_no} 页内容与已导出页面重复，停止翻页")
                break
            seen_signatures.add(signature)

            before = {p.name for p in run_dir.iterdir() if p.is_file()}
            rect = page.run_js(r"""
            return (function() {
              const host = document.querySelector('kat-button.export-report-button');
              if (!host || !host.shadowRoot) return null;
              const button = host.shadowRoot.querySelector('button');
              if (!button || button.disabled) return null;
              button.scrollIntoView({block: 'center', inline: 'center'});
              const rect = button.getBoundingClientRect();
              return {x: rect.left + rect.width / 2, y: rect.top + rect.height / 2, width: rect.width, height: rect.height};
            })();
            """)
            if not rect or rect.get('width', 0) <= 0 or rect.get('height', 0) <= 0:
                raise RuntimeError(f"第 {page_no} 页未找到导出报告按钮")

            page.run_cdp('Input.dispatchMouseEvent', type='mouseMoved', x=rect['x'], y=rect['y'], button='none')
            page.run_cdp('Input.dispatchMouseEvent', type='mousePressed', x=rect['x'], y=rect['y'], button='left', clickCount=1)
            page.run_cdp('Input.dispatchMouseEvent', type='mouseReleased', x=rect['x'], y=rect['y'], button='left', clickCount=1)

            new_file = None
            last_path = None
            last_size = -1
            stable_count = 0
            for _ in range(90):
                candidates = [
                    p for p in run_dir.iterdir()
                    if p.is_file() and p.name not in before and p.suffix.lower() in export_suffixes
                ]
                temp_files = [
                    p for p in run_dir.iterdir()
                    if p.is_file() and p.suffix.lower() in temp_suffixes
                ]
                candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                if candidates:
                    candidate = candidates[0]
                    size = candidate.stat().st_size
                    stable_count = stable_count + 1 if candidate == last_path and size == last_size and not temp_files else 0
                    last_path = candidate
                    last_size = size
                    if stable_count >= 2 and size > 0:
                        new_file = candidate
                        break
                time.sleep(1)

            if not new_file:
                raise RuntimeError(f"第 {page_no} 页导出失败，未检测到下载文件")

            exported_files.append(new_file)
            print(f"已导出第 {page_no} 页: {new_file.name}")

            next_rect = page.run_js(r"""
            return (function() {
              const pagination = document.querySelector('kat-pagination');
              if (!pagination || !pagination.shadowRoot) return null;
              const nav = pagination.shadowRoot.querySelector('[part="pagination-nav-right"]');
              if (!nav || /\bend\b/.test(nav.className || '')) return null;
              nav.scrollIntoView({block: 'center', inline: 'center'});
              const rect = nav.getBoundingClientRect();
              return {x: rect.left + rect.width / 2, y: rect.top + rect.height / 2, width: rect.width, height: rect.height};
            })();
            """)
            if not next_rect:
                break

            page.run_cdp('Input.dispatchMouseEvent', type='mouseMoved', x=next_rect['x'], y=next_rect['y'], button='none')
            page.run_cdp('Input.dispatchMouseEvent', type='mousePressed', x=next_rect['x'], y=next_rect['y'], button='left', clickCount=1)
            page.run_cdp('Input.dispatchMouseEvent', type='mouseReleased', x=next_rect['x'], y=next_rect['y'], button='left', clickCount=1)

            for _ in range(30):
                time.sleep(1)
                state = page.run_js(table_js)
                if state and state.get('signature') and state.get('signature') != signature:
                    break
            else:
                print("点击下一页后表格内容未变化，按末页处理")
                break
            page_no += 1

        print(f"共导出 {len(exported_files)} 个文件")

        frames = []
        for file_path in exported_files:
            try:
                if file_path.suffix.lower() == '.csv':
                    frame = pd.read_csv(file_path, encoding='utf-8-sig')
                else:
                    frame = pd.read_excel(file_path)
                if not frame.empty:
                    frames.append(frame)
            except Exception as e:
                print(f"读取 {file_path.name} 失败，跳过: {e}")

        if not frames:
            print("未拿到任何有效导出数据，未生成合并文件")
            return

        merged = pd.concat(frames, ignore_index=True)
        if '问题编号' in merged.columns:
            merged = merged.drop_duplicates(subset=['问题编号'], keep='first')
        elif '问题单编号' in merged.columns:
            merged = merged.drop_duplicates(subset=['问题单编号'], keep='first')
        else:
            merged = merged.drop_duplicates(keep='first')
        merged = merged.reset_index(drop=True)

        merged_path = Path(self.file).parent / "抓取合并结果.xlsx"
        merged.to_excel(merged_path, index=False)
        print(f"合并完成，共 {len(merged)} 行，已保存: {merged_path}")




    # 更新模式
    def Case(self,page):
        time.sleep(1)
        page.ele('x://div[@aria-label="语言"] | //div[@aria-label="Language"]').click()
        time.sleep(1.5)
        page.ele('x://div[text()="中文(简体)"]').click()
        time.sleep(7)


        # # 切换国家
        # page.ele('x://*[@href="/home"]/following-sibling::div').click()
        # time.sleep(3)
        # div = page.ele('x://*[text()="查看所有"]',timeout=5)
        # if div:
        #     div.click()
        #     time.sleep(5)
        # page.ele(f'x://span[contains(text(), "{self.country}")]').click()
        # time.sleep(0.78)
        # page.ele('x://kat-button[@label="选择账户"]').click()
        # time.sleep(8)
        # page.ele('x://div[@aria-label="语言"] | //div[@aria-label="Language"]').click()
        # time.sleep(1.5)
        # page.ele('x://*[text()="中文(简体)"]').click()
        # time.sleep(7)

        page.ele('x://span[text()="Help"] | //span[text()="帮助"]').click()
        time.sleep(0.78)
        page.ele('x://a[@href="/cu/case-lobby"]').click()  # 点击 Manage support cases
        time.sleep(2)


        df = pd.read_excel(self.file)

        # 转换为二维数组（跳过表头）
        two_d_array = df.values.tolist()

        for data in two_d_array:
            for _ in range(7):
                shadow_host = page.ele("x://kat-filter-bar")
                shadow_input = page.ele('x://*[@placeholder="按问题单编号或主题搜索"]')
                time.sleep(2)
                shadow_input.shadow_root('x://*[@placeholder="按问题单编号或主题搜索"]').clear()
                time.sleep(2)
                shadow_input.shadow_root('x://*[@placeholder="按问题单编号或主题搜索"]').input(str(data[1])+"\n")
                time.sleep(1.5)
                kat_button = shadow_host.shadow_root('x://kat-button').click(by_js=True)
                # kat_button.ele('x://*[text()="搜索"]').click()
                time.sleep(2)
                tr  = page.ele(f'x://tbody/tr//a[text()="{data[1]}"]',timeout=15)
                if not tr:
                    continue
                articles = page.ele("x://span[contains(., '问题单') and contains(., '共')]").text
                match = re.search(r'共\s*(\d+)\s*个', articles)
                number = match.group(1)

                if int(number) != 1:
                    page.refresh()
                    time.sleep(5)
                    continue
                else:
                    break

            status = page.ele('x://tbody/tr/td[2]').text  # 状态
            recentTime = page.ele('x://tbody/tr/td[5]').text  # 亚马逊最近回复

            if recentTime == data[4]:
                progress = "否"
            else:
                progress = "是"

            data[3] = status
            data[4] = recentTime
            data[5] = progress
        wb = load_workbook(self.file)
        ws = wb.active
        # 从第2行开始写回数据
        for idx, data in enumerate(two_d_array, start=2):
            ws.cell(row=idx, column=4, value=data[3])  # 状态
            ws.cell(row=idx, column=5, value=data[4])  # 最近回复日期
            ws.cell(row=idx, column=6, value=data[5])  # 有新进度

        wb.save(self.file)
        print(f"已更新 {len(two_d_array)} 条数据到 {self.file}")

    def Code(self):
        import time
        from pywinauto import Desktop
        time.sleep(1)

        desktop = Desktop(backend="uia")

        success = False

        for win in desktop.windows():
            try:
                for btn in win.descendants(control_type="Button"):
                    if "二步验证码服务" in btn.window_text():
                        btn.click_input()
                        time.sleep(1.5)

                        while True:
                            found = False

                            for b in win.descendants(control_type="Button"):
                                name = b.window_text()

                                if name == "填入验证码":
                                    b.click_input()
                                    time.sleep(1.5)
                                    b.click_input()
                                    found = True
                                    success = True
                                    break

                                elif name == "获取最新验证码":
                                    b.click_input()
                                    time.sleep(1)
                                    b.click_input()
                                    found = True
                                    break

                            if success:
                                break

                            if not found:
                                time.sleep(0.2)

                        break

                if success:
                    break

            except Exception as e:
                print(e)

        return success

    def Login(self,page):
        login = page.ele('x://input[@id="continue"]')
        SFA = page.ele('x://kat-input[@placeholder="Search for an account"]', timeout=5)
        if login:
            login.click()
            time.sleep(5)
            page.ele('x://input[@type="password"]').input(self.StorePassword, clear=True)
            time.sleep(0.78)
            page.ele('x://input[@id="signInSubmit"]').click()
            time.sleep(5)
            self.Code()  # 点击验证码插件
            time.sleep(0.78)
            page.ele('x://input[@type="submit"]').click()  # 填入验证码
            SFA = page.ele('x://*[@placeholder="Search for an account"]')
        password = page.ele('x://input[@type="password"]',timeout=5)
        if password:
            time.sleep(5)
            page.ele('x://input[@type="password"]').input(self.StorePassword, clear=True)
            time.sleep(0.78)
            page.ele('x://input[@id="signInSubmit"]').click()
            time.sleep(5)
            self.Code()  # 点击验证码插件
            time.sleep(0.78)
            page.ele('x://input[@type="submit"]').click()  # 填入验证码
            SFA = page.ele('x://*[@placeholder="Search for an account"]')
        if SFA:
            time.sleep(4)
            SFA.input("United States", by_js=True)
            time.sleep(0.78)
            page.ele('x://span[text()="United States"]').click()  # 选择美国
            time.sleep(0.78)
            page.ele('x://kat-button[@label="Select account"]').click()
            time.sleep(5)
            try:
                sub = page.ele('x://input[@type="submit"]').click()
                if sub:
                    page.ele('x://input[@type="password"]').input(self.StorePassword, clear=True)
                    time.sleep(0.78)
                    sub.click()
                    self.Code()  # 点击验证码插件
            except:
                pass


    def main(self):
        import os
        sp = Specification(self.username, self.password)  # 其他易得客
        time.sleep(5)
        sp.YidekeLogin()
        time.sleep(3)

        self.run_edecker_automation()  # 访问全部店铺
        time.sleep(2.5)
        for _ in range(7):
            self.start_edecker(self.ip, self.port)  # 启动指定易得客浏览器
            try:
                time.sleep(2)
                os.system("taskkill /f /im chrome.exe")
                time.sleep(2)
                page = ChromiumPage("127.0.0.1:" + str(self.port))  # 接管浏览器
                time.sleep(1.5)
                page.set.window.max()
                break
            except:
                page.quit()
                time.sleep(2)
                continue
        time.sleep(5)
        self.Login(page)
        time.sleep(1)

        if self.mode:
            self.Grabbing(page)
        # 模式选择
        else:
            self.Case(page)
        page.quit()
        # time.sleep(999999)



if __name__ == '__main__':
    config = {
        "username": os.getenv("YIDEKE_USERNAME", ""),
        "password": os.getenv("YIDEKE_PASSWORD", ""),
        "ip": os.getenv("YIDEKE_SHOP_IP", ""),
        "port": int(os.getenv("CASE_BROWSER_PORT", "9228")),
        "file": os.getenv("CASE_FILE_PATH", r"C:\RPA流程\Case状态抓取\file\工作簿.xlsx"),
        "StorePassword": os.getenv("AMAZON_STORE_PASSWORD", ""), # 店铺密码
        "country": "美国",  # 选择店铺国家
        "station": "",  # 选择站点，可以不填
        # "mode":"更新"#【更新、抓取】
        "mode":True
    }
    automation = Automation(config)
    automation.main()
