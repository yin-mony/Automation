"""Excel 读取与规格汇总表样式格式化。

read_excel：从多工作表提取「描述」「链接」并按 offer_id 全局去重。
excel_columns：对导出的规格汇总 sheet 做合并单元格、列宽与表头样式。
"""

import re

import pandas as pd
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelDF:
    """输入 Excel 解析与输出 Excel 格式化。"""

    def __init__(self, config):
        """config 可为文件路径字符串，或含 file_path 的字典。"""
        if isinstance(config, str):
            self.path = config
        else:
            self.path = config['file_path']
        self.desc_col = '描述'
        self.link_col = '链接'

    def read_excel(self):
        """读取所有工作表，提取「描述」「链接」列并按 offer_id 去重。

        跳过缺少必需列的工作表；同一 offer_id 跨表重复时只保留首次出现。
        返回 list[dict]，每项含 工作表、描述、链接。
        """
        data = []
        seen_offer_ids = set()  # 全局去重：跨 sheet 同一商品只采一次
        raw_count = 0
        skip_dup = 0

        excel_path = Path(self.path)
        if not excel_path.exists():
            raise FileNotFoundError(f'未找到 Excel 文件: {excel_path}')

        sheets = pd.read_excel(excel_path, sheet_name=None)

        for sheet_name, df in sheets.items():
            if self.desc_col not in df.columns or self.link_col not in df.columns:
                print(f'工作表「{sheet_name}」缺少「{self.desc_col}」或「{self.link_col}」列，已跳过')
                continue

            for _, row in df.iterrows():
                desc = row[self.desc_col]
                link = row[self.link_col]

                if pd.isna(desc) and pd.isna(link):
                    continue

                desc = '' if pd.isna(desc) else str(desc).strip()
                link = '' if pd.isna(link) else str(link).strip()

                if not desc and not link:
                    continue

                raw_count += 1
                text = str(link).strip()
                match = re.search(r'offer/(\d+)', text)
                if match:
                    offer_id = match.group(1)
                elif text.isdigit():
                    offer_id = text
                else:
                    offer_id = None
                # 无法解析 offer_id 的行仍保留（由 main.link_extract 再判）
                if offer_id and offer_id in seen_offer_ids:
                    skip_dup += 1
                    print(f'跳过重复 offer_id {offer_id}（{sheet_name}）')
                    continue
                if offer_id:
                    seen_offer_ids.add(offer_id)

                data.append({
                    '工作表': sheet_name,
                    self.desc_col: desc,
                    self.link_col: link,
                })

        if skip_dup:
            print(f'原始 {raw_count} 条，去重后 {len(data)} 条，跳过 {skip_dup} 条重复')

        for name in sheets:
            count = sum(1 for item in data if item['工作表'] == name)
            if count:
                print(f'工作表「{name}」: {count} 条')

        return data

    def excel_columns(self, ws, min_width=8, max_width=50):
        """格式化规格汇总 sheet：合并单元格、列宽、表头样式与边框。

        列序：1=offer_id, 2=链接, 3=描述, 4=规格信息, 5=长, 6=宽, 7=高, 8=重量, 9=原价, 10=优惠价
        """
        if ws.max_row < 2:
            return

        offer_cols = (1, 2, 3)  # offer_id / 链接 / 描述：同商品多 SKU 时纵向合并
        no_dim = '无具体尺寸信息'
        no_weight = '无重量信息'
        combined = '无具体尺寸/重量信息'

        # 样式常量
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(bold=True, color='FFFFFF')
        thin_side = Side(style='thin', color='D9D9D9')
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_align = Alignment(horizontal='center', vertical='center')
        offer_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        spec_align = Alignment(vertical='center', wrap_text=True)
        data_center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        phys_align = data_center_align

        # 1. 按连续相同 offer_id 纵向合并商品信息列
        block_start = 2
        prev_offer = ws.cell(row=2, column=1).value
        for row_idx in range(3, ws.max_row + 2):
            if row_idx <= ws.max_row:
                offer = ws.cell(row=row_idx, column=1).value
            else:
                offer = None
            if row_idx > ws.max_row or offer != prev_offer:
                block_end = row_idx - 1
                if block_end > block_start:
                    for col in offer_cols:
                        ws.merge_cells(
                            start_row=block_start,
                            start_column=col,
                            end_row=block_end,
                            end_column=col,
                        )
                        ws.cell(row=block_start, column=col).alignment = offer_align
                if row_idx <= ws.max_row:
                    block_start = row_idx
                    prev_offer = offer

        # 2. 长/宽/高/重量占位列横向合并
        for row_idx in range(2, ws.max_row + 1):
            length = ws.cell(row=row_idx, column=5).value
            width = ws.cell(row=row_idx, column=6).value
            height = ws.cell(row=row_idx, column=7).value
            weight = ws.cell(row=row_idx, column=8).value
            # 四项均为占位：合并 E:H，显示统一文案
            if length == width == height == no_dim and weight == no_weight:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=5,
                    end_row=row_idx,
                    end_column=8,
                )
                ws.cell(row=row_idx, column=5).value = combined
                ws.cell(row=row_idx, column=5).alignment = phys_align
            # 仅尺寸三项占位、重量有值：合并 E:G
            elif length == width == height == no_dim:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=5,
                    end_row=row_idx,
                    end_column=7,
                )
                ws.cell(row=row_idx, column=5).alignment = phys_align

        # 3. 按内容估算列宽（中文计 2 单位）
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                display_width = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, display_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(max_len + 2, min_width),
                max_width,
            )

        # 4. 表头样式与数据区边框、对齐
        max_col = ws.max_column
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                elif col_idx in offer_cols:
                    cell.alignment = offer_align
                elif col_idx == 4:
                    cell.alignment = spec_align
                else:
                    cell.alignment = data_center_align

        # 5. 冻结首行与表头筛选
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(max_col)}{ws.max_row}'


if __name__ == '__main__':
    # 本地调试：打印各表去重后的链接条数
    config = {
        'file_path': r"C:\Users\admin\Desktop\压体积包装_分类汇总分享.xlsx",
    }
    exceldf = ExcelDF(config)
    exceldf.read_excel()
