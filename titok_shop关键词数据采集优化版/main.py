import json
import os
import random
import re
import time
from urllib.parse import urljoin

from openpyxl import Workbook, load_workbook
from DrissionPage import ChromiumPage
from exceldf import ExcelDF
from TiTokshopLogin import TikTokPage


class TikTok:
    def __init__(self, page):
        """初始化浏览器页面、验证码处理器及采集配置。"""
        # 绑定浏览器页面对象
        self.page = page
        # 初始化验证码处理器
        self.captcha_handler = TikTokPage(page)
        # 导出目录与关键词 Excel 路径分开配置，互不混用
        self.file = config['file']
        self.file_path = config['file_path']
        # 商品页路由 JSON 的 script id
        self.ROUTER_ID = '__MODERN_ROUTER_DATA__'
        # 详情页 loader 路径 key
        self.PDP_LOADER = '(region)/pdp/(product_name_slug$)/(product_id)/page'
        # 两条链接之间的采集间隔范围（秒）
        self.LINK_GAP_MIN = 3
        self.LINK_GAP_MAX = 5
        # 验证码最大重试次数
        self.CAPTCHA_RETRY = 3
        # 验证码处理后的等待时间（秒）
        self.CAPTCHA_WAIT = 5
        # 等待页面加载商品数据的最长时间（秒）
        self.PAGE_WAIT = 15
        # 匹配 HTML 中 script 标签的正则
        self.SCRIPT_RE = re.compile(r'<script\b([^>]*)>(.*?)</script>', re.IGNORECASE | re.DOTALL)
        # 从 script 标签属性中提取 id 的正则
        self.SCRIPT_ID_RE = re.compile(r'\bid=["\']([^"\']+)["\']', re.IGNORECASE)

    def read_keywords(self):
        """从关键词 Excel 独立读取列表（路径与导出目录无关）。"""
        return ExcelDF({'file_path': self.file_path}).read_excel()

    def _output_path(self, filename):
        """根据 config['file'] 解析导出目录，创建目录并返回完整文件路径。"""
        # file 本身是目录时，直接作为导出目录
        if os.path.isdir(self.file):
            output_dir = self.file
        else:
            # file 是文件路径时，取其所在目录
            output_dir = os.path.dirname(os.path.abspath(self.file))
            if not output_dir:
                output_dir = os.path.dirname(os.path.abspath(__file__))
        # 确保导出目录存在
        os.makedirs(output_dir, exist_ok=True)
        return os.path.join(output_dir, filename)

    @staticmethod
    def _format_usd(amount):
        """将金额格式化为美元字符串，例如 15.95 -> $15.95。"""
        # 尝试转为浮点数并格式化为两位小数
        try:
            return f'${float(amount):.2f}'
        except (TypeError, ValueError):
            # 无法转换时原样返回，空值返回空字符串
            return f'${amount}' if amount not in (None, '') else ''

    def _blank_detail(self, product_id='', error=''):
        """返回空商品详情，解析失败或触发验证码时占位。"""
        # 返回统一结构的空商品详情字典
        return {
            'product_id': product_id,
            'name': '',
            'sold_count': '',
            'shop_name': '',
            'shop_rating': '',
            'product_rating': '',
            'review_count': '',
            'star_5_count': '',
            'star_4_count': '',
            'variant_options': {},
            'variant_option_names': [],
            'variant_option_count': 0,
            'shipping_status': '',
            'shipping_fee': '',
            'sale_price': '',
            'id_matched': False,
            'error': error,
        }

    def _solve_captcha(self, timeout=180):
        """处理验证码：先点关闭，不行再手动滑动。"""
        # 优先尝试点击关闭按钮
        if self.captcha_handler.try_close_captcha(self.page):
            return True
        # 关闭失败则等待用户手动完成滑动验证
        try:
            return self.captcha_handler.wait_manual_captcha(self.page, timeout=timeout)
        except TimeoutError:
            return False

    # 验证码
    def img_code(self, url=None, timeout=30, captcha_timeout=180):
        """无 url 时检测并处理验证码；有 url 时执行浏览器监听与验证码重试。"""
        # 无 url：仅检测当前页面是否有验证码
        if url is None:
            if self.captcha_handler.has_img_code(self.page):
                return self._solve_captcha(timeout=captcha_timeout)
            return True

        # 有 url：进入监听采集流程，记录验证码重试次数
        retry = 0

        while True:
            # 重启监听，访问商品页并等待数据包
            if self.page.listen.listening:
                self.page.listen.stop()
            self.page.listen.start(url)
            self.page.get(url)
            time.sleep(2)
            packet = self.page.listen.wait(timeout=timeout, fit_count=False)
            if self.page.listen.listening:
                self.page.listen.stop()

            # 未监听到数据包，直接返回错误
            if not packet:
                return {
                    'url': url,
                    'api_url': url,
                    'json_data': None,
                    'error': '未监听到该链接对应的数据包',
                }

            # 监听到多个包时取第一个
            if isinstance(packet, list):
                packet = packet[0]

            # 第一步：解析响应体中的 JSON script 标签
            item = {
                'url': url,
                'api_url': packet.url,
                'json_data': self.analysis_data(body=packet.response.body, parse_only=True),
                'error': '',
            }
            # 检查响应体和页面是否触发验证码
            resp_captcha = self.captcha_handler.is_captcha_response(item.get('json_data'))
            page_captcha = self.captcha_handler.has_img_code(self.page)

            # 无验证码，正常返回采集结果
            if not resp_captcha and not page_captcha:
                return item

            # 验证码重试次数超限，返回失败
            retry += 1
            if retry > self.CAPTCHA_RETRY:
                item['error'] = '触发滑动验证码，重试后仍未获取商品数据'
                return item

            # 尝试处理验证码
            print(f'  检测到滑动验证码（第 {retry}/{self.CAPTCHA_RETRY} 次），尝试点击关闭...')
            # 响应含验证码但页面未显示时，重新加载页面
            if resp_captcha and not page_captcha:
                self.page.get(url)
                time.sleep(2)

            if not self._solve_captcha(timeout=captcha_timeout):
                item['error'] = '触发滑动验证码，关闭失败且未完成手动验证'
                return item

            # 验证码处理完毕，等待页面恢复后再从 HTML 提取数据
            print(f'  验证码已处理，等待 {self.CAPTCHA_WAIT} 秒后继续获取数据...')
            time.sleep(self.CAPTCHA_WAIT)

            # 轮询当前页面 HTML，尝试直接提取商品数据
            t0 = time.time()
            from_page = None
            while time.time() - t0 < self.PAGE_WAIT:
                try:
                    html = self.page.html
                except Exception:
                    html = ''
                parsed = {
                    'url': url,
                    'api_url': self.page.url or url,
                    # 第一步：从页面 HTML 提取 JSON script
                    'json_data': self.analysis_data(body=html, parse_only=True),
                    'error': '',
                }
                # 第二步：解析商品业务字段
                parsed = self.analysis_data(parsed)
                if parsed.get('product_detail', {}).get('name'):
                    from_page = parsed
                    break
                time.sleep(1)

            # 从页面 HTML 成功提取到商品数据
            if from_page:
                print('  已从当前页面直接提取商品数据（无需重新请求）')
                return from_page

            # 页面暂无数据，刷新后再次尝试从 HTML 提取
            print('  当前页面暂无商品数据，尝试刷新页面...')
            self.page.refresh()
            time.sleep(3)

            t0 = time.time()
            while time.time() - t0 < 10:
                try:
                    html = self.page.html
                except Exception:
                    html = ''
                parsed = {
                    'url': url,
                    'api_url': self.page.url or url,
                    # 第一步：从页面 HTML 提取 JSON script
                    'json_data': self.analysis_data(body=html, parse_only=True),
                    'error': '',
                }
                # 第二步：解析商品业务字段
                parsed = self.analysis_data(parsed)
                if parsed.get('product_detail', {}).get('name'):
                    print('  刷新后已从当前页面提取商品数据')
                    return parsed
                time.sleep(1)

            # 刷新后仍无数据，重新进入监听循环
            print('  页面仍未加载商品数据，将重新监听请求...')

    def url(self):
        """按关键词搜索商品、加载全部列表，并导出商品链接 txt 文件。"""
        # 打开 TikTok Shop 首页
        self.page.get('https://shop.tiktok.com/us')
        time.sleep(2)

        # 统一关键词为列表格式
        keywords = self.read_keywords()
        all_urls = []

        for kw in keywords:
            # 输入关键词并校验搜索框内容是否一致
            while True:
                self.img_code()
                search_ele = self.page.ele('x://input[@aria-label="Search"]', timeout=20)
                search_ele.input(f'{kw}\n', clear=True)
                time.sleep(1)
                typed = (
                    self.page.ele('x://input[@aria-label="Search"]', timeout=10).attr('value') or ''
                ).strip()
                if typed.lower() == kw.strip().lower():
                    break
                print(f'搜索框关键词不一致，当前: {typed}，目标: {kw}')
                self.page.refresh()
                time.sleep(3)
            time.sleep(3)

            # 循环点击「查看更多」直到加载全部商品
            while True:
                self.img_code()
                if self.page.ele('x://span[text()="No more products"]', timeout=1):
                    print('已展示全部相关商品')
                    break
                button = self.page.ele('x://button[text()="查看更多"]', timeout=2)
                if not button:
                    print('已展示全部相关商品')
                    break
                button.click()
                time.sleep(2)

            # 从页面提取所有商品详情链接
            time.sleep(2)
            links = self.page.eles('x://div[@class=" max-h-51"]/a')
            links = [link for link in links if link.attr('href') and '/pdp/' in link.attr('href')]
            print(f'共获取到 {len(links)} 条商品链接')

            # 拼接完整 URL 并打印
            urls = []
            for i, link in enumerate(links, 1):
                link_url = urljoin('https://shop.tiktok.com', link.attr('href'))
                urls.append(link_url)
                print(f'{i}. {link_url}')

            # 按关键词导出链接 txt 文件
            safe_name = re.sub(r'[^\w\-]+', '_', kw.strip()).strip('_') or 'keywords'
            links_path = self._output_path(f'{safe_name}_商品链接.txt')
            with open(links_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(urls))
            print(f'商品链接已导出到: {links_path}')
            all_urls.extend(urls)

        return all_urls

    # 数据采集
    def data(self, limit=None, timeout=30):
        """读取商品链接文件，逐条监听采集 JSON 数据并导出数据与摘要文件。"""
        # 统一关键词为列表格式
        keywords = self.read_keywords()
        all_items = []

        for kw in keywords:
            # 定位当前关键词对应的链接文件
            safe_name = re.sub(r'[^\w\-]+', '_', kw.strip()).strip('_') or 'keywords'
            links_path = self._output_path(f'{safe_name}_商品链接.txt')

            if not os.path.exists(links_path):
                raise FileNotFoundError(f'未找到文件: {links_path}')

            # 从 txt 文件中提取所有 HTTP 链接（全量链接列表）
            with open(links_path, 'r', encoding='utf-8') as f:
                content = f.read()
            links = re.findall(r'https?://[^\s<>"\']+', content)
            links = [(link or '').strip().rstrip('.,;') for link in links]

            # 读取已有采集结果，过滤出尚未采集的新链接
            data_path = self._output_path(f'{safe_name}_商品数据.json')
            old_results = []
            if os.path.exists(data_path):
                with open(data_path, 'r', encoding='utf-8') as f:
                    old_results = json.load(f)
            collected = {(i.get('url') or '').strip().rstrip('.,;') for i in old_results}
            new_links = [link for link in links if link not in collected]
            print(
                f'全量链接 {len(links)} 条，已采集 {len(collected)} 条，'
                f'本次增量采集 {len(new_links)} 条'
            )

            # 限制采集条数（调试用，仅作用于增量链接）
            if limit:
                new_links = new_links[:limit]

            new_results = []

            if new_links:
                # 预热浏览器会话，避免首次请求被拦截
                print('预热会话：访问 TikTok Shop 首页...')
                self.page.get('https://shop.tiktok.com/us')
                time.sleep(2)
                self.img_code()

                # 逐条监听尚未采集的商品链接
                for i, url in enumerate(new_links, 1):
                    print(f'[{i}/{len(new_links)}] 正在监听: {url}')
                    try:
                        item = self.img_code(url, timeout=timeout)
                        # 第二步：从 json_data 提取商品业务字段
                        item = self.analysis_data(item)
                        new_results.append(item)
                        detail = item.get('product_detail', {})

                        # 打印成功采集的商品信息
                        if detail.get('name'):
                            print(
                                f"  商品: {detail.get('name')}\n"
                                f"  已售: {detail.get('sold_count')} | 售价: {detail.get('sale_price', '')} | "
                                f"店铺: {detail.get('shop_name')} | 店铺评分: {detail.get('shop_rating')}\n"
                                f"  商品评分: {detail.get('product_rating')}| 商品评论总数: ({detail.get('review_count')}) | "
                                f"5星: {detail.get('star_5_count')} | 4星: {detail.get('star_4_count')}\n"
                                f"  运费: {detail.get('shipping_status', '')} {detail.get('shipping_fee', '')} | "
                                f"规格选项数: {detail.get('variant_option_count', 0)} | "
                                f"选项名称: {detail.get('variant_option_names', [])}"
                            )
                            if not detail.get('id_matched'):
                                print(f"  校验失败: {detail.get('error')}")
                        # 有 JSON 但未提取到商品字段
                        elif item['json_data'] and item['json_data'].get('script_count'):
                            print(
                                f"  已提取 {item['json_data']['script_count']} 个 application/json 标签，"
                                f"监听地址: {item['api_url']}"
                            )
                            if detail.get('error'):
                                print(f"  字段提取失败: {detail.get('error')}")
                        else:
                            print(f"  未获取到数据: {item['error']}")
                    except Exception as e:
                        # 单条采集异常，记录错误后继续下一条
                        print(f'  监听失败: {e}')
                        item = {
                            'url': url,
                            'api_url': url,
                            'json_data': None,
                            'error': str(e),
                        }
                        item = self.analysis_data(item)
                        new_results.append(item)
                    finally:
                        # 停止监听并在两条链接之间随机等待
                        if self.page.listen.listening:
                            self.page.listen.stop()
                        if i < len(new_links):
                            delay = random.uniform(self.LINK_GAP_MIN, self.LINK_GAP_MAX)
                            print(f'  等待 {delay:.1f} 秒后继续下一条...')
                            time.sleep(delay)
            else:
                print('无新增链接，跳过采集，沿用已有数据')

            # 合并旧数据与本次增量，导出 JSON
            seen = {(i.get('url') or '').strip().rstrip('.,;') for i in old_results}
            results = list(old_results)
            for item in new_results:
                url = (item.get('url') or '').strip().rstrip('.,;')
                if url and url not in seen:
                    seen.add(url)
                    results.append(item)

            with open(data_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            print(f'JSON 数据已导出到: {data_path}')

            summary = []
            for item in results:
                detail = item.get('product_detail', {})
                summary.append({
                    'url': item.get('url', ''),
                    'api_url': item.get('api_url', ''),
                    'product_id': detail.get('product_id', ''),
                    'name': detail.get('name', ''),
                    'sold_count': detail.get('sold_count', ''),
                    'shop_name': detail.get('shop_name', ''),
                    'shop_rating': detail.get('shop_rating', ''),
                    'product_rating': detail.get('product_rating', ''),
                    'review_count': detail.get('review_count', ''),
                    'star_5_count': detail.get('star_5_count', ''),
                    'star_4_count': detail.get('star_4_count', ''),
                    'variant_options': detail.get('variant_options', {}),
                    'variant_option_names': detail.get('variant_option_names', []),
                    'variant_option_count': detail.get('variant_option_count', 0),
                    'shipping_status': detail.get('shipping_status', ''),
                    'shipping_fee': detail.get('shipping_fee', ''),
                    'sale_price': detail.get('sale_price', ''),
                    'id_matched': detail.get('id_matched', False),
                    'error': detail.get('error') or item.get('error', ''),
                })

            summary_path = self._output_path(f'{safe_name}_商品摘要.json')
            with open(summary_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
            print(f'商品摘要已导出到: {summary_path}')

            all_items.extend(results)

        return all_items

    # json数据分析
    def analysis_data(self, item=None, body=None, parse_only=False):
        """两步解析：第一步从 HTML/响应体提取 JSON script；第二步从 json_data 提取商品字段。"""
        # 第一步：从 HTML/响应体中提取 application/json script 标签
        if parse_only or (body is not None and item is None):
            # 根据 body 类型获取 HTML 字符串
            if isinstance(body, dict):
                if 'application_json_scripts' in body:
                    return body
                html = body.get('raw_body', '')
            elif isinstance(body, str):
                html = body
            else:
                return {'application_json_scripts': [], 'script_count': 0}

            # 遍历 HTML 中所有 script 标签，提取 JSON 内容
            scripts = []
            if isinstance(html, str):
                for index, (attrs, content) in enumerate(
                    self.SCRIPT_RE.findall(html), 1
                ):
                    if 'application/json' not in attrs.lower():
                        continue

                    id_match = self.SCRIPT_ID_RE.search(attrs)
                    script_id = id_match.group(1) if id_match else f'script_{index}'
                    text = content.strip()

                    try:
                        data = json.loads(text)
                    except json.JSONDecodeError:
                        data = {'raw_text': text}

                    scripts.append({
                        'id': script_id,
                        'data': data,
                    })

            return {
                'application_json_scripts': scripts,
                'script_count': len(scripts),
            }

        # 第二步：从 item 的 json_data 中解析商品业务字段
        # 从 URL 中提取目标商品 ID
        url = item.get('url', '')
        api_url = item.get('api_url', '')
        match = re.search(r'/(\d{10,})(?:\?|$)', url)
        url_pid = str(item.get('product_id') or (match.group(1) if match else ''))

        # 响应体含验证码，返回空详情
        json_data = item.get('json_data')
        if self.captcha_handler.is_captcha_response(json_data):
            item['product_detail'] = self._blank_detail(
                product_id=url_pid,
                error=item.get('error') or '触发滑动验证码，未完成验证',
            )
            return item

        # 从 script 列表中查找路由 JSON（__MODERN_ROUTER_DATA__）
        router = None
        if isinstance(json_data, dict):
            for script in json_data.get('application_json_scripts', []):
                if script.get('id') == self.ROUTER_ID:
                    router = script.get('data')
                    break

        if not router:
            item['product_detail'] = self._blank_detail(
                product_id=url_pid,
                error='未找到 __MODERN_ROUTER_DATA__ 数据',
            )
            return item

        # 从路由数据中获取详情页 loader 和 product_info 组件
        loader = router.get('loaderData', {}).get(self.PDP_LOADER, {})
        blocks = loader.get('page_config', {}).get('components_map', [])
        info_block = next(
            (b for b in blocks if b.get('component_type') == 'product_info'),
            None,
        )
        if not info_block:
            item['product_detail'] = self._blank_detail(
                product_id=url_pid,
                error='未找到 product_info 组件',
            )
            return item

        # 拆分各业务子模块：商品、评价、店铺、星级分布
        block = info_block.get('component_data', {})
        pinfo = block.get('product_info', {})
        model = pinfo.get('product_model', {})
        reviews = pinfo.get('review_model', {})
        shop = block.get('shop_info', {})
        stars = (
            block.get('review_info', {})
            .get('review_ratings', {})
            .get('rating_result', {})
        )

        # 提取规格选项：按属性分组 + 扁平名称列表
        variants = {}
        opt_names = []
        for prop in model.get('sale_properties', []):
            prop_name = prop.get('property_name') or f"property_{prop.get('property_id', '')}"
            names = [
                v.get('property_value_name', '').strip()
                for v in prop.get('property_values', [])
                if v.get('property_value_name')
            ]
            if not names:
                continue
            variants[prop_name] = {'names': names, 'count': len(names)}
            opt_names.extend(names)

        # 校验 JSON 中的商品 ID 是否与 URL 一致
        pid = str(model.get('product_id', ''))
        pid_ok = pid and pid == url_pid and pid in url and pid in api_url

        # 提取运费信息
        ship_status, ship_fee = '', ''
        promo = pinfo.get('promotion_model', {})
        ships = promo.get('promotion_logistic_list', [])
        if ships:
            ship = ships[0]
            fee = ship.get('shippingFee', {}) or {}
            fee_val = fee.get('real_price', '')
            if ship.get('freeShipping') or str(fee_val) in ('0', '0.00'):
                ship_status, ship_fee = '免费运送', self._format_usd('0')
            else:
                ship_status, ship_fee = '有运费', self._format_usd(fee_val)

        # 提取最低售价
        sale_price = ''
        price_raw = promo.get('promotion_product_price', {}).get('min_price', {}).get('sale_price_decimal', '')
        if price_raw:
            sale_price = self._format_usd(price_raw)

        # 组装商品详情字典
        detail = {
            'product_id': pid,
            'name': model.get('name', ''),
            'sold_count': model.get('sold_count', ''),
            'sale_price': sale_price,
            'shop_name': shop.get('shop_name', ''),
            'shop_rating': shop.get('shop_rating', ''),
            'product_rating': reviews.get('product_overall_score', ''),
            'review_count': reviews.get('product_review_count', ''),
            'star_5_count': stars.get('5', ''),
            'star_4_count': stars.get('4', ''),
            'variant_options': variants,
            'variant_option_names': opt_names,
            'variant_option_count': len(opt_names),
            'shipping_status': ship_status,
            'shipping_fee': ship_fee,
            'id_matched': pid_ok,
            'error': '' if pid_ok else 'product_id 与当前监听地址不匹配',
        }

        # 仅有 Default 规格时标记为无变体
        opts = detail.get('variant_option_names', [])
        if isinstance(opts, list):
            opts_clean = [str(n).strip() for n in opts if str(n).strip()]
            if opts_clean and all(n == 'Default' for n in opts_clean):
                detail['variant_option_names'] = '该商品无变体'
                detail['variant_option_count'] = 0
        elif str(opts).strip() == 'Default':
            detail['variant_option_names'] = '该商品无变体'
            detail['variant_option_count'] = 0

        # 无评价时替换为友好提示文案
        rv_count = str(detail.get('review_count', '')).strip()
        if rv_count in ('0', '0.0'):
            detail['review_count'] = '该商品无任何评价'
            detail['product_rating'] = '该商品没有星级'

        item['product_detail'] = detail
        return item

    def main(self):
        """采集商品数据并按关键词分 sheet 导出 Excel 表格。"""
        # 执行数据采集，获取全部结果（含历史 + 增量）
        results = self.data()

        # Excel 表头定义
        headers = [
            '商品标题',
            '完整商品链接',
            '已售出',
            '商品售价',
            '店铺名',
            '店铺评分',
            '商品星级',
            '商品评论总数',
            '五星评论数量',
            '四星评论数量',
            '运费状态',
            '订单运费',
            '变体名称',
            '变体总数',
        ]

        keywords = self.read_keywords()
        if os.path.isdir(self.file):
            excel_path = os.path.join(self.file, 'TikTok_Shop_竞品数据.xlsx')
        else:
            excel_path = self.file

        # 加载已有 Excel，保留不在本次 config 中的旧 sheet
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        # 仅更新当前 config 中各关键词对应的 sheet
        for kw in keywords:
            safe_name = re.sub(r'[^\w\-]+', '_', kw.strip()).strip('_') or 'keywords'
            links_path = self._output_path(f'{safe_name}_商品链接.txt')
            link_set = set()
            if os.path.exists(links_path):
                with open(links_path, 'r', encoding='utf-8') as f:
                    link_set = {
                        (link or '').strip().rstrip('.,;')
                        for link in re.findall(r'https?://[^\s<>"\']+', f.read())
                    }

            sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', kw.strip()).strip() or '未命名关键词'
            sheet_name = sheet_name[:31]
            kw_items = [
                item for item in results
                if not link_set or (item.get('url') or '').strip().rstrip('.,;') in link_set
            ]

            # 重写当前关键词 sheet，其他 sheet 不动
            if sheet_name in wb.sheetnames:
                idx = wb.sheetnames.index(sheet_name)
                wb.remove(wb[sheet_name])
                ws = wb.create_sheet(sheet_name, idx)
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            for item in kw_items:
                detail = item.get('product_detail', {})
                opts = detail.get('variant_option_names', [])
                if isinstance(opts, list):
                    opts_str = ', '.join(opts)
                else:
                    opts_str = str(opts or '')
                ws.append([
                    detail.get('name', ''),
                    item.get('url', ''),
                    detail.get('sold_count', ''),
                    detail.get('sale_price', ''),
                    detail.get('shop_name', ''),
                    detail.get('shop_rating', ''),
                    detail.get('product_rating', ''),
                    detail.get('review_count', ''),
                    detail.get('star_5_count', ''),
                    detail.get('star_4_count', ''),
                    detail.get('shipping_status', ''),
                    detail.get('shipping_fee', ''),
                    opts_str,
                    detail.get('variant_option_count', ''),
                ])
            print(f'已写入工作表: {sheet_name}（共 {len(kw_items)} 条）')

        # 全新工作簿且无任何 sheet 时创建占位页
        if not wb.sheetnames:
            ws = wb.create_sheet('暂无数据')
            ws.append(headers)

        wb.save(excel_path)
        print(f'表格已导出到: {excel_path}（已保留其他历史 sheet）')
        return results

    def run(self):
        """一键执行：获取链接 -> 采集数据 -> 导出 Excel。"""
        # 第 1 步：搜索关键词并导出商品链接
        print('=== 第1步：获取商品链接 ===')
        self.url()
        # 第 2 步：逐条采集商品数据（data 内部会打印进度）
        print('=== 第2步：采集商品数据 ===')
        # 第 3 步：将采集结果导出为 Excel
        print('=== 第3步：导出 Excel 表格 ===')
        return self.main()


if __name__ == '__main__':
    # file_path：关键词来源 Excel；file：链接/JSON/导出 Excel 目录
    config = {
        # 初始关键词表格文件
        'file_path': r'C:\Users\admin\Desktop\tiktok-竞品信息抓取.xlsx',
        # 导出目录
        'file': r'D:\Desktop',
    }
    # 启动浏览器并执行完整采集流程
    page = ChromiumPage()
    tiktok = TikTok(page=page)
    tiktok.run()
