from DrissionPage import ChromiumPage,Chromium

from openpyxl import load_workbook
from urllib.parse import urljoin
import time
import os
import re


class TikTok:
    def __init__(self):
        self.file = config['files']

    def has_img_code(self, page):
        title = ''
        html = ''
        try:
            title = page.title.lower()
        except Exception:
            pass
        try:
            html = page.html.lower()
        except Exception:
            pass
        keywords = (
            'security check',
            'verify to continue',
            'drag the puzzle piece',
            'drag the puzzle piece into place',
        )
        return any(k in title or k in html for k in keywords)

    def wait_img_code(self, page, timeout=180):
        """监测页面是否出现滑动验证码，出现则等待人工完成。"""
        if not self.has_img_code(page):
            return True
        print('检测到滑动验证码，请手动完成滑动...')
        start_time = time.time()
        while time.time() - start_time < timeout:
            if not self.has_img_code(page):
                print('人工验证码已通过，继续执行')
                return True
            time.sleep(2)
        raise TimeoutError('等待人工完成验证码超时')

    def main(self):
        page = ChromiumPage()
        page.get('https://shop.tiktok.com/us')

        time.sleep(2)
        # 关键词
        codes = "Hand Sanitizer Spray"
        while True:
            self.wait_img_code(page)
            search_ele = page.ele('x://input[@aria-label="Search"]', timeout=20)
            search_ele.input(f'{codes}\n', clear=True)
            time.sleep(1)
            current_keyword = (page.ele('x://input[@aria-label="Search"]', timeout=10).attr('value') or '').strip()
            if current_keyword.lower() == codes.strip().lower():
                break
            print(f'搜索框关键词不一致，当前: {current_keyword}，目标: {codes}')
            page.refresh()
            time.sleep(3)
        time.sleep(3)

        while True:
            self.wait_img_code(page)
            if page.ele('x://span[text()="No more products"]', timeout=1):
                print("已展示全部相关商品")
                break
            button = page.ele('x://button[text()="查看更多"]', timeout=2)
            if not button:
                print("已展示全部相关商品")
                break
            button.click()
            time.sleep(2)

        time.sleep(2)
        links = page.eles('x://div[@class=" max-h-51"]/a')
        links = [link for link in links if link.attr('href') and '/pdp/' in link.attr('href')]
        print(f'共获取到 {len(links)} 条商品链接')
        for i, link in enumerate(links, 1):
            url = urljoin('https://shop.tiktok.com', link.attr('href'))
            print(f'{i}. {url}')
            tab = Chromium()
            tab.new_tab(url=url)
            print('已创建并进入当前页面')
            # page.listen.start(url)

        data = []

        #     query = 'source=ecommerce_searchresult&enter_method=feed_list_search_word&first_entrance=ecommerce_mall&first_entrance_position=search&first_entrance_tt_scene=seo'
        #     for i, link in enumerate(links[:5], 1):
        #         login.img_code()
        #         # 商品链接地址
        #         base_url = urljoin('https://shop.tiktok.com', link.attr('href'))
        #         # 拼接完成后的完整商品链接地址
        #         url = f'{base_url}&{query}' if '?' in base_url else f'{base_url}?{query}'
        #         print(f'{i}. {url}')
        #         item_data = {
        #             '关键词': keyword,
        #             '链接': url,
        #             '销售店铺名称': '',
        #             '完整商品标题': '',
        #             '商品话题标签': [],
        #             '售卖单价': '',
        #             '运费（Free shipping=0 美元包邮）': '',
        #             '产品星级': '',
        #             '评价总数': '',
        #             '五星 / 四星分项评价数量': '',
        #             '累计已售件数': '',
        #             '全系列变体': ''
        #         }
        #
        #         tab = None
        #         try:
        #             tab = page.new_tab(url=url)
        #             self.wait_img_code(tab)
        #             time.sleep(1)
        #
        #             # 获取销售店铺
        #             store = tab.ele('x://a[contains(text(), "销售店铺")]', timeout=10)
        #             if store:
        #                 shop_name = re.sub(r'^销售店铺\s*', '', store.text.strip())
        #                 item_data['销售店铺名称'] = shop_name
        #                 print(f'店铺名: {shop_name}')
        #             else:
        #                 print('未找到销售店铺')
        #
        #             # 获取完整商品标题
        #             title = tab.ele('x://h1/span[contains(@class, "H2-Semibold")]', timeout=10)
        #             if title:
        #                 shop_title = title.text.strip()
        #                 item_data['完整商品标题'] = shop_title
        #                 print(f'商品标题: {shop_title}')
        #             else:
        #                 shop_title = ''
        #                 print('未找到商品标题')
        #
        #             # 获取商品话题标签
        #             topic = re.findall(r'#\S+', shop_title)
        #             if topic:
        #                 item_data['商品话题标签'] = topic
        #                 print(f'商品标签: {topic}')
        #             else:
        #                 item_data['商品话题标签'] = "该商品未携带话题标签"
        #
        #             # 获取售卖单价
        #             price = tab.ele('x://span[contains(@class, "items-baseline")]/span', timeout=25)
        #             shop_price = price.text.strip() if price else ''
        #             item_data['售卖单价'] = shop_price
        #             print("售卖单价: " + shop_price)
        #             # 获取运费（Free shipping=0 美元包邮）
        #             free_shipping = tab.ele('x://span[contains(text(), "免费运送")]', timeout=10)
        #             if free_shipping:
        #                 item_data['运费（Free shipping=0 美元包邮）'] = free_shipping.text + '$0.00'
        #             else:
        #                 order_shipping = tab.ele('x://span[contains(text(), "此订单运费为")]', timeout=10)
        #                 if order_shipping:
        #                     item_data['运费（Free shipping=0 美元包邮）'] = order_shipping.text
        #             # 产品星级
        #             review_section = tab.ele('x://div[@id="pdp-review-section"]', timeout=10)
        #             if review_section:
        #                 review_section.scroll.to_see(center=True)
        #                 time.sleep(1)
        #                 rating = tab.ele('x://div/span[contains(@class, "H2-Semibold mr-2")]', timeout=25)
        #                 # rating.scroll.to_see(center=True)
        #                 item_data['产品星级'] = rating.text if rating else ''
        #                 time.sleep(1)
        #             else:
        #                 item_data['产品星级'] = "该产品没有相关星级信息"
        #                 print("该产品没有相关星级信息")
        #             # 获取评价总数
        #             review_total = tab.ele('x://div[contains(@class, "H2-Semibold")]', timeout=25)
        #             item_data['评价总数'] = review_total.text if review_total else ''
        #             if review_total:
        #                 time.sleep(1)
        #                 # 五星 / 四星分项评价数量
        #                 five_star_count = tab.ele('x://div[div/div[text()="5"]]//div[contains(@class,"H3-Regular")]',
        #                                           timeout=25)
        #                 # item_data['五星评价数量'] = five_star_count.text if five_star_count else ''
        #                 time.sleep(1)
        #                 four_star_count = tab.ele('x://div[div/div[text()="4"]]//div[contains(@class,"H3-Regular")]',
        #                                           timeout=25)
        #                 # item_data['四星评价数量'] = four_star_count.text if four_star_count else ''
        #                 # 合计数量:五星+四星
        #                 item_data['五星 / 四星分项评价数量'] = int(four_star_count.text) + int(five_star_count.text)
        #                 time.sleep(1)
        #             else:
        #                 item_data['评价总数'] = "该商品没有相关评价信息"
        #                 print('该商品没有相关评价信息')
        #
        #             # 累计已售件数
        #             sold_count = tab.ele('x://span[contains(@class,"H3-Regular")][contains(.,"已售")]', timeout=25)
        #             item_data['累计已售件数'] = sold_count.text if sold_count else ''
        #
        #             # 全系列变体
        #             if tab.ele('x://span[text() = "Scent"]', timeout=25):
        #                 variant = tab.eles('x://div[contains(@class,"border-solid")]/span', timeout=25)
        #                 variant_count = len(variant)
        #                 # 存放成字符串
        #                 variant_text = [ele.text.strip() for ele in variant if ele.text.strip()]
        #                 item_data['全系列变体'] = ', '.join(variant_text)
        #                 print(variant_count)
        #             else:
        #                 item_data['全系列变体'] = "该商品无任何变体信息"
        #                 print("该商品无任何变体信息")
        #
        #             data.append(item_data)
        #             print(item_data)
        #         except Exception as e:
        #             print(f'商品采集失败，跳过: {url}，原因: {e}')
        #         finally:
        #             if tab:
        #                 tab.close()
        # print(data)
        # return data

    # 根据main函数处理完成 返回的data数据，做新建工作表处理
    def new_excel_file(self):
        data = self.main()
        wb = load_workbook(self.file)

        if not data:
            sheet_name = '暂无数据'
            index = 1
            while sheet_name in wb.sheetnames:
                suffix = f'_{index}'
                sheet_name = f'暂无数据'[:31 - len(suffix)] + suffix
                index += 1
            ws = wb.create_sheet(sheet_name)
            ws.append(['暂无数据'])
            wb.save(self.file)
            print(f'已创建工作表: {sheet_name}')
            return data

        grouped_data = {}
        for item in data:
            keyword = str(item.get('关键词') or '未命名关键词').strip()
            grouped_data.setdefault(keyword, []).append(item)

        def get_sheet_name(value):
            value = re.sub(r'[\[\]\:\*\?\/\\]', '_', value).strip() or '未命名关键词'
            return value[:31]

        for keyword, items in grouped_data.items():
            sheet_name = get_sheet_name(keyword)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
            if not headers:
                headers = [key for key in items[0].keys() if key != '关键词']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            for item in items:
                row = []
                for key in headers:
                    value = '' if key == '关键词' else item.get(key, '')
                    if isinstance(value, list):
                        value = ', '.join(value)
                    row.append(value)
                ws.append(row)
            print(f'已写入工作表: {sheet_name}')

        wb.save(self.file)
        return data

    def run(self):
        self.main()
        # self.new_excel_file()


if __name__ == '__main__':
    config = {
        "files": r"C:\Users\admin\Desktop\tiktok-竞品信息抓取.xlsx",
    }
    tiktok = TikTok()
    tiktok.run()
