import json
import html
import os
import re
import smtplib
import time
from calendar import monthrange
from datetime import datetime
from email.header import Header
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from html import unescape
from pathlib import Path
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen
from openpyxl import Workbook, load_workbook
from email_util import EmailUtil

try:
    from DrissionPage import ChromiumPage
except Exception:
    ChromiumPage = None



# Trac 公司+个人合作推广主流程
class Main:
    # 初始化配置、固定状态、文件名和搜索参数。
    def __init__(self, config=None):
        # baseDir：当前子项目目录，保证从任意工作目录启动都能找到本机配置、file 和 output。
        self.baseDir = Path(__file__).resolve().parent
        # localConfigPath：本机明文配置文件路径，只在本机使用，不提交到远端仓库。
        localConfigPath = self.baseDir / "run_config.local.json"
        # localConfig：本机配置默认空字典，文件不存在时不影响程序启动。
        localConfig = {}
        # 如果本机存在 run_config.local.json，就先读取里面的 SerpApi Key 和 SMTP 授权码。
        if localConfigPath.exists():
            # read_text：读取本机 JSON 配置内容。
            localText = localConfigPath.read_text(encoding="utf-8")
            # json.loads：把本机 JSON 配置转换为字典。
            localData = json.loads(localText)
            # isinstance：只接受 JSON 对象，避免误把数组或字符串当作配置。
            if isinstance(localData, dict):
                # localConfig：保存本机配置，后续会和 GUI 配置合并。
                localConfig = localData
        # config：外部传入配置；GUI 会传入用户填写项，命令行调试可传空。
        config = dict(config or {})
        # localConfig.update：GUI/命令行传入配置优先级更高，本机配置主要补充密钥字段。
        localConfig.update(config)
        # config：合并后的运行配置。
        config = localConfig
        # self.config：项目唯一默认配置来源，不再单独维护默认配置函数。
        self.config = {
            # isOnline：运行环境标记；False 表示本机，True 表示线上。
            "isOnline": bool(config.get("isOnline", False)),
            # outputDir：最终结果表、断点和缓存所在目录；不能用于存放内置底表。
            "outputDir": str(config.get("outputDir") or "output"),
            # rawFileName：file 内置目录中的 TREC 未清洗全量底表文件名，仅保留和邮件附件使用。
            "rawFileName": str(config.get("rawFileName", "初始总量数据未清洗.xlsx")),
            # cleanFileName：file 内置目录中的已清洗初始表文件名，当前搜索流程只读取它。
            "cleanFileName": str(config.get("cleanFileName", "已获取到的初始总数据.xlsx")),
            # companyResultFileName：公司模式最终导出表文件名。
            "companyResultFileName": str(config.get("companyResultFileName", "已完成搜索匹配的公司联系信息数据.xlsx")),
            # personResultFileName：个人模式最终导出表文件名。
            "personResultFileName": str(config.get("personResultFileName", "已完成搜索匹配的个人联系信息数据.xlsx")),
            # expireMonths：个人模式只处理未来多少个月内到期的数据。
            "expireMonths": int(config.get("expireMonths", 6)),
            # serpapiUrl：SerpApi Google 搜索接口地址。
            "serpapiUrl": str(config.get("serpapiUrl", "https://serpapi.com/search")),
            # serpapiKey：SerpApi Key 不进入 GUI 和 run_config.json；正式运行从环境变量 TREC_SERPAPI_KEY 读取。
            "serpapiKey": str(config.get("serpapiKey") or os.getenv("TREC_SERPAPI_KEY", "")),
            # sendEmail：是否在流程结束后发送固定数据表附件。
            "sendEmail": bool(config.get("sendEmail", False)),
            # email：收件邮箱；开启邮件发送时必填。
            "email": str(config.get("email", "")),
            # sender_email：固定发件邮箱，不交给用户在 GUI 中修改。
            "sender_email": str(config.get("sender_email", "1974419863@qq.com")),
            # smtp_auth_code：结果附件 SMTP 授权码不进 GUI；正式运行从环境变量 TREC_RESULT_SMTP_AUTH_CODE 读取。
            "smtp_auth_code": str(config.get("smtp_auth_code") or os.getenv("TREC_RESULT_SMTP_AUTH_CODE", "")),
            # emailSubject：邮件标题。
            "emailSubject": str(config.get("emailSubject", "自动化_TREC公司+个人合作推广数据")),
            # promotionExecuteSend：推广邮件单一开关；False 只生成后台记录，True 真实发送。
            "promotionExecuteSend": bool(config.get("promotionExecuteSend", False)),
            # promotionSenderEmail：推广邮件固定发件邮箱，GUI 只展示不可修改。
            "promotionSenderEmail": "info@time2renew.com",
            # promotionSmtpAuthCode：推广邮件 SMTP 授权码不进 GUI；优先读本机配置，其次读环境变量。
            "promotionSmtpAuthCode": str(config.get("promotionSmtpAuthCode") or os.getenv("TREC_PROMO_SMTP_AUTH_CODE", "")),
        }
        # dataDir：内置数据文件目录，固定为当前子项目下的 file，不允许从 GUI 或配置文件修改。
        self.dataDir = self.baseDir / "file"
        # outputDir：最终结果表、断点和缓存目录，不再和 file 内置数据目录混用。
        outputDir = Path(self.config["outputDir"])
        self.outputDir = outputDir if outputDir.is_absolute() else self.baseDir / outputDir
        
        # 输出目录禁止使用 file，避免覆盖或污染内置底表目录。
        if self.outputDir.resolve() == self.dataDir.resolve() or self.outputDir.name.lower() == self.dataDir.name.lower():
            raise ValueError("outputDir 不能设置为内置数据目录 file，请改为 output 或其他输出目录。")
        # rawFile：未清洗底表文件名，主要用于邮件附件。
        self.rawFile = self.config["rawFileName"]
        # cleanFile：已清洗初始表文件名，主流程必须读取它。
        self.cleanFile = self.config["cleanFileName"]
        # companyFile：公司模式导出文件名。
        self.companyFile = self.config["companyResultFileName"]
        # personFile：个人模式导出文件名。
        self.personFile = self.config["personResultFileName"]
        # months：个人模式到期月份筛选值。
        self.months = self.config["expireMonths"]
        # serpUrl：SerpApi 请求地址。
        self.serpUrl = self.config["serpapiUrl"]
        # serpKey：SerpApi 请求密钥。
        self.serpKey = self.config["serpapiKey"].strip()
        # online：当前运行环境标记。
        self.online = self.config["isOnline"]
        # sendNotice：结果附件邮件发送开关。
        self.sendNotice = self.config["sendEmail"]
        # promoSend：推广邮件单一开关，False 只生成后台记录，True 真实发送。
        self.promoSend = self.config["promotionExecuteSend"]
        # promoLogFile：推广发送记录固定文件名。
        self.promoLogFile = "邮件发送记录.xlsx"
        # promoFrom：推广邮件发件邮箱，只做展示和流程内部使用。
        self.promoFrom = self.config["promotionSenderEmail"]
        # runMode：GUI 当前运行状态；只在运行时传入，不写入默认配置。
        self.mode = str(config.get("runMode", "company")).strip().lower()
        # batch：程序固定每次处理对象数。
        self.batch = 10
        # retry：SerpApi 内部固定重试次数。
        self.retry = 3
        # pageLimit：二级页面最大读取字节数，避免单页过大拖慢。
        self.pageLimit = 400000
        # page：二级页面抓取使用 DP 浏览器，默认启动或复用一个 ChromiumPage。
        self.page = config.get("page")
        # pageWait：二级页面打开后的固定等待秒数，避免页面未加载完就提取。
        self.pageWait = 5
        # ua：SerpApi 和普通二级页面请求头。
        self.ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36"
        # companyProg：公司模式后台断点文件。
        self.companyProg = "serpapi_company_search_progress.json"
        # companyCache：公司模式后台缓存文件。
        self.companyCache = "serpapi_company_contacts_cache.json"
        # personProg：个人模式后台断点文件。
        self.personProg = "serpapi_person_search_progress.json"
        # personCache：个人模式后台缓存文件。
        self.personCache = "serpapi_person_contacts_cache.json"

    # 归档旧版 TREC 网站采集流程，当前版本不再执行。
    def trecArchive(self):

        # 旧流程保留说明：
        # 1. 使用 DrissionPage 打开 https://www.trec.texas.gov/。
        # 2. 进入 License Holder Search 页面，选择 Active 状态。
        # 3. 监听 collections/licenses/documents/search 接口。
        # 4. 读取 TREC 列表接口 hits 数据，按 detailId 去重后导出“初始总量数据未清洗.xlsx”。
        # 5. 再读取 Texas Open Data，把挂靠公司、过期日期、地区等字段融合到清洗初始表。
        # 当前版本已经内置清洗后的全量表“已获取到的初始总数据.xlsx”，所以不再重复执行网站采集。
        return None

    # 搜索并解析一个对象，包含 SerpApi 请求、二级页面抓取和联系方式提取。
    def search(self, query):
        txt = lambda value: "" if value is None else str(value).strip()
        if not self.serpKey:
            return {
                "emails": [], "phones": [], "sourceUrls": [], "detailUrls": [],
                "facebookUrls": [], "resultCount": 0, "errorReason": "SerpApi Key 为空",
            }

        params = {
            "engine": "google",
            "q": query,
            "api_key": self.serpKey,
            "hl": "en",
            "gl": "us",
            "start": 0,
        }
        url = self.serpUrl + "?" + urlencode(params)
        request = Request(url, headers={"User-Agent": self.ua, "Accept": "application/json"})

        data = {}
        lastError = ""
        for attempt in range(1, self.retry + 1):
            try:
                response = urlopen(request, timeout=60).read().decode("utf-8")
                data = json.loads(response)
                break
            except Exception as error:
                lastError = str(error)
                print("SerpApi 请求失败，准备重试:", attempt, "/", self.retry, lastError)

        if not data:
            return {
                "emails": [], "phones": [], "sourceUrls": [], "detailUrls": [],
                "facebookUrls": [], "resultCount": 0, "errorReason": lastError or "SerpApi 请求失败",
            }

        organicResults = data.get("organic_results") or []
        emails = []
        phones = []
        sourceUrls = []
        detailUrls = []
        facebookUrls = []

        emailPattern = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
        phonePattern = re.compile(r"\+?1?[\s.-]?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}")
        blockedHosts = ["facebook.com", "instagram.com", "linkedin.com", "youtube.com", "google.com"]
        blockedExts = [".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".zip", ".mp4", ".mp3"]
        badEmailNames = {
            "admin", "administrator", "info", "information", "contact", "contactus", "support", "help",
            "hello", "office", "mail", "email", "team", "sales", "marketing", "service", "services",
            "customerservice", "customer.service", "webmaster", "postmaster", "abuse", "privacy",
            "legal", "compliance", "noreply", "no-reply", "donotreply", "do-not-reply", "reception",
            "receptionist", "billing", "accounting", "hr", "careers", "jobs", "press", "media",
        }
        badEmailTlds = (".gov", ".mil", ".edu")
        badEmailWords = (
            "army", "navy", "airforce", "marines", "military", "defense", "dod", "usaf",
            "usarmy", "uscg", "police", "sheriff", "school", "k12", "isd",
        )

        for item in organicResults:
            title = txt(item.get("title"))
            link = txt(item.get("link"))
            snippet = txt(item.get("snippet"))
            richText = json.dumps(item.get("rich_snippet") or "", ensure_ascii=False)
            searchText = " ".join([title, link, snippet, richText])

            for email in emailPattern.findall(searchText):
                email = email.lower().strip("._-")
                emailName = email.split("@", 1)[0] if "@" in email else ""
                emailDomain = email.split("@", 1)[1] if "@" in email else ""
                badEmail = (
                    not email
                    or ".." in email
                    or emailName in badEmailNames
                    or emailDomain.endswith(badEmailTlds)
                    or any(word in emailDomain for word in badEmailWords)
                )
                if not badEmail and email not in emails:
                    emails.append(email)

            for phone in phonePattern.findall(searchText):
                digits = re.sub(r"\D", "", phone)
                if len(digits) == 11 and digits.startswith("1"):
                    digits = digits[1:]
                if len(digits) == 10:
                    phoneText = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
                    badPhone = len(set(digits)) == 1 or digits[3:6] == "555" or digits in {"1234567890", "0123456789", "5555555555"}
                    if not badPhone and phoneText not in phones:
                        phones.append(phoneText)

            if not link:
                continue

            if link not in sourceUrls:
                sourceUrls.append(link)
            parsedUrl = urlparse(link)
            facebookHost = parsedUrl.netloc.lower().endswith("facebook.com")
            facebookPath = parsedUrl.path.lower()
            facebookBlocked = ["/login", "/search", "/share", "/groups/", "/events/", "/posts/", "/photos/", "/videos/"]
            if facebookHost and facebookPath and facebookPath != "/" and not any(item in facebookPath for item in facebookBlocked) and link not in facebookUrls:
                facebookUrls.append(link)

            detailText = ""
            host = parsedUrl.netloc.lower()
            path = parsedUrl.path.lower()
            if parsedUrl.scheme in ["http", "https"] and host:
                canRead = True
                if any(host == item or host.endswith("." + item) for item in blockedHosts):
                    canRead = False
                if any(path.endswith(item) for item in blockedExts):
                    canRead = False
                if canRead:
                    try:
                        if self.page is None:
                            if ChromiumPage is None:
                                raise RuntimeError("当前环境未安装 DrissionPage，无法启动浏览器抓取二级页面")
                            print("启动浏览器用于二级页面抓取")
                            self.page = ChromiumPage()

                        print("浏览器打开二级页面:", link)
                        self.page.get(link)
                        time.sleep(self.pageWait)

                        body = self.page.ele("tag:body", timeout=3)
                        if body:
                            detailText = txt(body.text)
                        else:
                            pageHtml = txt(getattr(self.page, "html", ""))
                            detailText = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", pageHtml)
                            detailText = re.sub(r"(?is)<br\s*/?>", " ", detailText)
                            detailText = re.sub(r"(?is)<[^>]+>", " ", detailText)
                            detailText = re.sub(r"\s+", " ", unescape(detailText))
                    except Exception as error:
                        print("二级页面抓取失败:", link, txt(error)[:120])

            if detailText:
                if link not in detailUrls:
                    detailUrls.append(link)
                for email in emailPattern.findall(detailText):
                    email = email.lower().strip("._-")
                    emailName = email.split("@", 1)[0] if "@" in email else ""
                    emailDomain = email.split("@", 1)[1] if "@" in email else ""
                    badEmail = (
                        not email
                        or ".." in email
                        or emailName in badEmailNames
                        or emailDomain.endswith(badEmailTlds)
                        or any(word in emailDomain for word in badEmailWords)
                    )
                    if not badEmail and email not in emails:
                        emails.append(email)
                for phone in phonePattern.findall(detailText):
                    digits = re.sub(r"\D", "", phone)
                    if len(digits) == 11 and digits.startswith("1"):
                        digits = digits[1:]
                    if len(digits) == 10:
                        phoneText = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
                        badPhone = len(set(digits)) == 1 or digits[3:6] == "555" or digits in {"1234567890", "0123456789", "5555555555"}
                        if not badPhone and phoneText not in phones:
                            phones.append(phoneText)

        return {
            "emails": emails,
            "phones": phones,
            "sourceUrls": sourceUrls,
            "detailUrls": detailUrls,
            "facebookUrls": facebookUrls,
            "resultCount": len(organicResults),
            "errorReason": txt(data.get("error")),
        }

    # 公司模式：筛选挂靠公司，搜索公司联系方式。
    def company(self, rows):
        txt = lambda value: "" if value is None else str(value).strip()
        key = lambda value: re.sub(r"\s+", " ", txt(value)).lower()
        items = []
        itemMap = {}
        for row in rows:
            if key(row.get("status")) != "active":
                continue
            name = txt(row.get("company_name") or row.get("relation_name"))
            if not name:
                continue

            itemKey = key(name)
            if itemKey not in itemMap:
                itemMap[itemKey] = {
                    "companyKey": itemKey,
                    "companyName": name,
                    "companyCode": txt(row.get("company_code")),
                    "sampleAgent": txt(row.get("name")),
                    "regionText": txt(row.get("region_text")),
                    "city": txt(row.get("city")),
                    "state": txt(row.get("state")),
                    "agentCount": 0,
                }
                items.append(itemMap[itemKey])

            itemMap[itemKey]["agentCount"] += 1

        outFile = self.outputDir / self.companyFile
        cacheFile = self.outputDir / self.companyCache
        progFile = self.outputDir / self.companyProg
        results = []
        if cacheFile.exists():
            try:
                oldData = json.loads(cacheFile.read_text(encoding="utf-8"))
                results = oldData if isinstance(oldData, list) else []
            except Exception:
                results = []
        badNames = {
            "admin", "administrator", "info", "information", "contact", "contactus", "support", "help",
            "hello", "office", "mail", "email", "team", "sales", "marketing", "service", "services",
            "customerservice", "customer.service", "webmaster", "postmaster", "abuse", "privacy",
            "legal", "compliance", "noreply", "no-reply", "donotreply", "do-not-reply", "reception",
            "receptionist", "billing", "accounting", "hr", "careers", "jobs", "press", "media",
        }
        badTlds = (".gov", ".mil", ".edu")
        badWords = ("army", "navy", "airforce", "marines", "military", "defense", "dod", "usaf", "usarmy", "uscg", "police", "sheriff", "school", "k12", "isd")
        mailRe = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
        phoneRe = re.compile(r"\+?1?[\s.-]?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}")
        for rec in results:
            mails = []
            for mail in mailRe.findall(txt(rec.get("emails"))):
                mail = mail.lower().strip("._-")
                mailName = mail.split("@", 1)[0] if "@" in mail else ""
                mailDomain = mail.split("@", 1)[1] if "@" in mail else ""
                badMail = not mail or ".." in mail or mailName in badNames or mailDomain.endswith(badTlds) or any(word in mailDomain for word in badWords)
                if not badMail and mail not in mails:
                    mails.append(mail)
            rec["emails"] = "; ".join(mails)
            cleanPhones = []
            for phone in phoneRe.findall(txt(rec.get("phones"))):
                digits = re.sub(r"\D", "", phone)
                if len(digits) == 11 and digits.startswith("1"):
                    digits = digits[1:]
                badPhone = len(digits) != 10 or len(set(digits)) == 1 or digits[3:6] == "555" or digits in {"1234567890", "0123456789", "5555555555"}
                phoneText = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}" if len(digits) == 10 else ""
                if not badPhone and phoneText not in cleanPhones:
                    cleanPhones.append(phoneText)
            rec["phones"] = "; ".join(cleanPhones)
            rec["contactStatus"] = "已找到联系方式" if rec["emails"] or rec["phones"] else "未找到联系方式"
        if cacheFile.exists():
            cacheFile.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        doneKeys = {item.get("companyKey") for item in results if item.get("companyKey")}

        start = 0
        if progFile.exists():
            try:
                prog = json.loads(progFile.read_text(encoding="utf-8"))
                start = int(prog.get("nextIndex", 0) or 0)
            except Exception:
                start = 0

        end = min(len(items), start + self.batch)
        print("公司模式待搜索公司数量:", len(items))
        print("公司模式本次固定处理数量:", self.batch)
        print("公司模式本次范围:", start + 1, "-", end)

        headers = [
            "companyName", "companyCode", "agentCount", "sampleAgent", "regionText", "query",
            "emails", "phones", "facebookUrls", "sourceUrls", "detailUrls", "serpapiResultCount",
            "detailPageCount", "contactStatus", "searchStatus", "errorReason",
        ]
        zh = {
            "companyName": "公司名称",
            "companyCode": "挂靠许可证号",
            "agentCount": "关联经纪人数量",
            "sampleAgent": "样本经纪人",
            "regionText": "地区信息",
            "query": "搜索词",
            "emails": "邮箱",
            "phones": "电话",
            "facebookUrls": "Facebook主页候选",
            "sourceUrls": "SerpApi来源链接",
            "detailUrls": "已抓取二级页面链接",
            "serpapiResultCount": "SerpApi结果数",
            "detailPageCount": "二级页面抓取数",
            "contactStatus": "采集状态",
            "searchStatus": "搜索状态",
            "errorReason": "错误原因",
        }

        for index in range(start, end):
            item = items[index]
            itemKey = item.get("companyKey", "")
            name = item.get("companyName", "")

            if itemKey in doneKeys:
                print("公司已在缓存中，跳过:", name)
                progFile.parent.mkdir(parents=True, exist_ok=True)
                progFile.write_text(
                    json.dumps({"nextIndex": index + 1, "total": len(items), "lastName": name}, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                continue

            area = item.get("city") or item.get("state") or "Texas"
            query = f'"{name}" real estate contact email phone {area}'
            print("SerpApi 搜索公司:", index + 1, "/", len(items), query)
            found = self.search(query)
            error = txt(found.get("errorReason"))
            status = "失败" if error else "已完成"

            rec = {
                "companyKey": itemKey,
                "companyName": name,
                "companyCode": item.get("companyCode", ""),
                "agentCount": item.get("agentCount", 0),
                "sampleAgent": item.get("sampleAgent", ""),
                "regionText": item.get("regionText", ""),
                "query": query,
                "emails": "; ".join(found.get("emails", [])),
                "phones": "; ".join(found.get("phones", [])),
                "facebookUrls": "; ".join(found.get("facebookUrls", [])),
                "sourceUrls": "; ".join(found.get("sourceUrls", [])),
                "detailUrls": "; ".join(found.get("detailUrls", [])),
                "serpapiResultCount": found.get("resultCount", 0),
                "detailPageCount": len(found.get("detailUrls", [])),
                "contactStatus": "已找到联系方式" if found.get("emails") or found.get("phones") else "未找到联系方式",
                "searchStatus": status,
                "errorReason": error,
            }

            results.append(rec)
            doneKeys.add(itemKey)
            cacheFile.parent.mkdir(parents=True, exist_ok=True)
            cacheFile.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
            progFile.write_text(
                json.dumps({"nextIndex": index + 1, "total": len(items), "lastName": name}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "公司联系信息"
            worksheet.append([zh.get(header, header) for header in headers])
            for row in results:
                worksheet.append([row.get(header, "") for header in headers])
            worksheet.auto_filter.ref = worksheet.dimensions
            outFile.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(outFile))
            print("结果已导出:", outFile, "数量:", len(results))

        if results:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "公司联系信息"
            worksheet.append([zh.get(header, header) for header in headers])
            for row in results:
                worksheet.append([row.get(header, "") for header in headers])
            worksheet.auto_filter.ref = worksheet.dimensions
            outFile.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(outFile))
            print("结果已导出:", outFile, "数量:", len(results))

        print("公司模式完成，本次累计结果:", len(results))
        return results

    # 个人模式：筛选无挂靠且快到期个人，搜索个人联系方式。
    def person(self, rows):
        txt = lambda value: "" if value is None else str(value).strip()
        key = lambda value: re.sub(r"\s+", " ", txt(value)).lower()
        today = datetime.today().date()
        monthNo = today.month - 1 + self.months
        endYear = today.year + monthNo // 12
        endMonth = monthNo % 12 + 1
        endDayNo = min(today.day, monthrange(endYear, endMonth)[1])
        endDay = today.replace(year=endYear, month=endMonth, day=endDayNo)
        items = []
        seen = set()

        for row in rows:
            if key(row.get("status")) != "active":
                continue
            if txt(row.get("company_name")):
                continue

            expDate = None
            expText = txt(row.get("expiration_date"))
            for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"]:
                try:
                    expDate = datetime.strptime(expText, fmt).date()
                    break
                except ValueError:
                    pass
            if not expDate or expDate < today or expDate > endDay:
                continue

            itemKey = txt(row.get("uid")) or key(row.get("name") + row.get("code", ""))
            if not itemKey or itemKey in seen:
                continue
            seen.add(itemKey)

            items.append({
                "personKey": itemKey,
                "uid": txt(row.get("uid")),
                "code": txt(row.get("code")),
                "name": txt(row.get("name")),
                "url": txt(row.get("url")),
                "status": txt(row.get("status")),
                "expirationDate": txt(row.get("expiration_date")),
                "licenseType": txt(row.get("license_type") or row.get("record_type")),
                "city": txt(row.get("city")),
                "county": txt(row.get("county")),
                "state": txt(row.get("state")),
                "zipCode": txt(row.get("zip_code")),
                "regionText": txt(row.get("region_text")),
            })

        outFile = self.outputDir / self.personFile
        cacheFile = self.outputDir / self.personCache
        progFile = self.outputDir / self.personProg
        results = []
        if cacheFile.exists():
            try:
                oldData = json.loads(cacheFile.read_text(encoding="utf-8"))
                results = oldData if isinstance(oldData, list) else []
            except Exception:
                results = []
        badNames = {
            "admin", "administrator", "info", "information", "contact", "contactus", "support", "help",
            "hello", "office", "mail", "email", "team", "sales", "marketing", "service", "services",
            "customerservice", "customer.service", "webmaster", "postmaster", "abuse", "privacy",
            "legal", "compliance", "noreply", "no-reply", "donotreply", "do-not-reply", "reception",
            "receptionist", "billing", "accounting", "hr", "careers", "jobs", "press", "media",
        }
        badTlds = (".gov", ".mil", ".edu")
        badWords = ("army", "navy", "airforce", "marines", "military", "defense", "dod", "usaf", "usarmy", "uscg", "police", "sheriff", "school", "k12", "isd")
        mailRe = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
        phoneRe = re.compile(r"\+?1?[\s.-]?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}")
        for rec in results:
            mails = []
            for mail in mailRe.findall(txt(rec.get("emails"))):
                mail = mail.lower().strip("._-")
                mailName = mail.split("@", 1)[0] if "@" in mail else ""
                mailDomain = mail.split("@", 1)[1] if "@" in mail else ""
                badMail = not mail or ".." in mail or mailName in badNames or mailDomain.endswith(badTlds) or any(word in mailDomain for word in badWords)
                if not badMail and mail not in mails:
                    mails.append(mail)
            rec["emails"] = "; ".join(mails)
            cleanPhones = []
            for phone in phoneRe.findall(txt(rec.get("phones"))):
                digits = re.sub(r"\D", "", phone)
                if len(digits) == 11 and digits.startswith("1"):
                    digits = digits[1:]
                badPhone = len(digits) != 10 or len(set(digits)) == 1 or digits[3:6] == "555" or digits in {"1234567890", "0123456789", "5555555555"}
                phoneText = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}" if len(digits) == 10 else ""
                if not badPhone and phoneText not in cleanPhones:
                    cleanPhones.append(phoneText)
            rec["phones"] = "; ".join(cleanPhones)
            rec["contactStatus"] = "已找到联系方式" if rec["emails"] or rec["phones"] else "未找到联系方式"
        if cacheFile.exists():
            cacheFile.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        doneKeys = {item.get("personKey") for item in results if item.get("personKey")}

        start = 0
        if progFile.exists():
            try:
                prog = json.loads(progFile.read_text(encoding="utf-8"))
                start = int(prog.get("nextIndex", 0) or 0)
            except Exception:
                start = 0

        end = min(len(items), start + self.batch)
        print("个人模式待搜索个人数量:", len(items))
        print("个人模式本次固定处理数量:", self.batch)
        print("个人模式本次范围:", start + 1, "-", end)

        headers = [
            "name", "code", "uid", "url", "status", "expirationDate", "licenseType", "city",
            "county", "state", "zipCode", "regionText", "query", "emails", "phones",
            "facebookUrls", "sourceUrls", "detailUrls", "serpapiResultCount", "detailPageCount",
            "contactStatus", "searchStatus", "errorReason",
        ]
        zh = {
            "name": "姓名",
            "code": "许可证号",
            "uid": "详情ID",
            "url": "详情页链接",
            "status": "状态",
            "expirationDate": "过期日期",
            "licenseType": "许可证类型",
            "city": "城市",
            "county": "县/郡",
            "state": "州",
            "zipCode": "邮编",
            "regionText": "地区信息",
            "query": "搜索词",
            "emails": "邮箱",
            "phones": "电话",
            "facebookUrls": "Facebook主页候选",
            "sourceUrls": "SerpApi来源链接",
            "detailUrls": "已抓取二级页面链接",
            "serpapiResultCount": "SerpApi结果数",
            "detailPageCount": "二级页面抓取数",
            "contactStatus": "采集状态",
            "searchStatus": "搜索状态",
            "errorReason": "错误原因",
        }

        for index in range(start, end):
            item = items[index]
            itemKey = item.get("personKey", "")
            name = item.get("name", "")

            if itemKey in doneKeys:
                print("个人已在缓存中，跳过:", name)
                progFile.parent.mkdir(parents=True, exist_ok=True)
                progFile.write_text(
                    json.dumps({"nextIndex": index + 1, "total": len(items), "lastName": name}, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                continue

            area = item.get("city") or item.get("state") or "Texas"
            query = f'"{name}" "real estate agent" contact email phone {area}'
            print("SerpApi 搜索个人:", index + 1, "/", len(items), query)
            found = self.search(query)
            error = txt(found.get("errorReason"))
            status = "失败" if error else "已完成"

            rec = {
                "personKey": itemKey,
                "name": name,
                "code": item.get("code", ""),
                "uid": item.get("uid", ""),
                "url": item.get("url", ""),
                "status": item.get("status", ""),
                "expirationDate": item.get("expirationDate", ""),
                "licenseType": item.get("licenseType", ""),
                "city": item.get("city", ""),
                "county": item.get("county", ""),
                "state": item.get("state", ""),
                "zipCode": item.get("zipCode", ""),
                "regionText": item.get("regionText", ""),
                "query": query,
                "emails": "; ".join(found.get("emails", [])),
                "phones": "; ".join(found.get("phones", [])),
                "facebookUrls": "; ".join(found.get("facebookUrls", [])),
                "sourceUrls": "; ".join(found.get("sourceUrls", [])),
                "detailUrls": "; ".join(found.get("detailUrls", [])),
                "serpapiResultCount": found.get("resultCount", 0),
                "detailPageCount": len(found.get("detailUrls", [])),
                "contactStatus": "已找到联系方式" if found.get("emails") or found.get("phones") else "未找到联系方式",
                "searchStatus": status,
                "errorReason": error,
            }

            results.append(rec)
            doneKeys.add(itemKey)
            cacheFile.parent.mkdir(parents=True, exist_ok=True)
            cacheFile.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
            progFile.write_text(
                json.dumps({"nextIndex": index + 1, "total": len(items), "lastName": name}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "个人联系信息"
            worksheet.append([zh.get(header, header) for header in headers])
            for row in results:
                worksheet.append([row.get(header, "") for header in headers])
            worksheet.auto_filter.ref = worksheet.dimensions
            outFile.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(outFile))
            print("结果已导出:", outFile, "数量:", len(results))

        if results:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "个人联系信息"
            worksheet.append([zh.get(header, header) for header in headers])
            for row in results:
                worksheet.append([row.get(header, "") for header in headers])
            worksheet.auto_filter.ref = worksheet.dimensions
            outFile.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(outFile))
            print("结果已导出:", outFile, "数量:", len(results))

        print("个人模式完成，本次累计结果:", len(results))
        return results

    # 推广邮件流程：读取公司/个人结果表，生成发送记录，并按开关发送推广邮件。
    def promoMail(self):
        txt = lambda value: "" if value is None else str(value).strip()
        # sourceFiles：固定读取当前项目 outputDir 下两个模式的最终结果表。
        sourceFiles = [
            ("公司", self.outputDir / self.companyFile),
            ("个人", self.outputDir / self.personFile),
        ]

        # recordPath：推广邮件发送记录表，固定写入 outputDir，不写入 file 内置数据目录。
        recordPath = self.outputDir / self.promoLogFile

        # fixedText：固定列名配置，兼容中文表头和少量历史英文表头。
        emailHeaders = ("邮箱", "email", "emails", "office email", "office_email")
        collectStatusHeaders = ("采集状态", "contact_status")
        searchStatusHeaders = ("搜索状态", "search_status")
        companyNameHeaders = ("公司名称", "company_name")
        companyLicenseHeaders = ("挂靠许可证号", "公司许可证号", "许可证号", "code")
        companySourceHeaders = ("SerpApi来源链接", "已抓取二级页面链接", "Facebook主页候选", "source_urls", "detail_urls")
        personNameHeaders = ("姓名", "name")
        personLicenseHeaders = ("许可证号", "code")
        personSourceHeaders = ("HAR来源链接", "Google来源链接", "SerpApi来源链接", "已抓取二级页面链接", "source_urls", "detail_urls")

        # outputHeaders：后台发送记录表头，和 trec推广发送 项目的记录格式保持一致。
        outputHeaders = [
            "来源类型", "运行环境", "对象名称", "许可证号", "邮箱", "邮件主题", "邮件正文",
            "采集状态", "搜索状态", "来源链接", "来源文件", "来源行号",
            "发送状态", "发送结果", "失败原因", "备注",
        ]

        # senderEmail：推广邮件固定发件邮箱，只在 GUI 展示，不允许在 GUI 修改。
        senderEmail = txt(self.promoFrom)

        # smtpServer：推广邮件固定 SMTP 服务器，不进入初始配置和 GUI。
        smtpServer = "smtp.qiye.aliyun.com"

        # smtpPort：推广邮件固定 SMTP SSL 端口，不进入初始配置和 GUI。
        smtpPort = 465

        # smtpUser：推广邮件固定 SMTP 登录账号，和固定发件邮箱保持一致。
        smtpUser = senderEmail

        # smtpPassword：推广邮件 SMTP 授权码不进 GUI；优先读本机配置，其次读环境变量。
        smtpPassword = txt(self.config.get("promotionSmtpAuthCode") or os.getenv("TREC_PROMO_SMTP_AUTH_CODE", ""))

        # emailSubject：推广邮件主题，写在流程函数中，便于和推广流程一起维护。
        emailSubject = "Partner with us on agent CE renewals"

        # emailMainBody：推广邮件正文主体，写在流程函数中，不放进初始化配置。
        emailMainBody = (
            "Hi,\n"
            "Your agents' renewal season is coming up. We offer a full 18-hour "
            "TREC-approved CE package at $49.99 -- probably the lowest price they'll "
            "find. TREC Provider #11011-CEP.\n\n"
            "Here's the deal: for every agent in your office who uses our package, "
            "I'll send you 20% back as a referral fee. No complicated setup -- just "
            "a straight split.\n\n"
            "Your agents get a solid course at a great price. You get an easy way "
            "to help your team save while putting something back in your pocket.\n\n"
            "Want to see the quality first? Let me know if you're open to a quick chat.\n\n"
            "Best,\n"
            "Qian Yi\n"
            "Ho"
        )

        # emailSignature：推广邮件签名文字，写在流程函数中。
        emailSignature = "Time2renew Support Team\n\nWebsite: time2renew"

        # websiteUrl：推广邮件 HTML 签名中的官网链接。
        websiteUrl = "https://time2renew.com"

        # logoPath：推广邮件内嵌 Logo，固定读取当前项目 file 内置目录。
        logoPath = self.dataDir / "time2renew-logo.png"

        # logoCid：推广邮件 HTML 里引用 Logo 的 CID。
        logoCid = "time2renew-logo"

        # logoWidth：推广邮件 Logo 展示宽度。
        logoWidth = 180

        # sendWaitSecond：真实发送推广邮件时每封之间的等待秒数。
        sendWaitSecond = 10.0

        # headerCleanPattern：表头匹配时忽略大小写、空格、下划线和常见中英文标点。
        headerCleanPattern = re.compile(r"[ _\-：:（）()]")

        # emailPattern：只从邮箱列里识别标准邮箱地址。
        emailPattern = re.compile(
            r"(?<![A-Za-z0-9._%+-])"
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
            r"(?![A-Za-z0-9._%+-])"
        )
        badEmailNames = {
            "admin", "administrator", "info", "information", "contact", "contactus", "support", "help",
            "hello", "office", "mail", "email", "team", "sales", "marketing", "service", "services",
            "customerservice", "customer.service", "webmaster", "postmaster", "abuse", "privacy",
            "legal", "compliance", "noreply", "no-reply", "donotreply", "do-not-reply", "reception",
            "receptionist", "billing", "accounting", "hr", "careers", "jobs", "press", "media",
        }
        badEmailTlds = (".gov", ".mil", ".edu")
        badEmailWords = (
            "army", "navy", "airforce", "marines", "military", "defense", "dod", "usaf",
            "usarmy", "uscg", "police", "sheriff", "school", "k12", "isd",
        )

        # emailHeaderKeys：邮箱列候选表头的标准化集合。
        emailHeaderKeys = {headerCleanPattern.sub("", txt(header).lower()) for header in emailHeaders}

        # collectStatusKeys：采集状态列候选表头的标准化集合。
        collectStatusKeys = {headerCleanPattern.sub("", txt(header).lower()) for header in collectStatusHeaders}

        # searchStatusKeys：搜索状态列候选表头的标准化集合。
        searchStatusKeys = {headerCleanPattern.sub("", txt(header).lower()) for header in searchStatusHeaders}

        # sourceRuleMap：公司和个人结果表的对象名、许可证号、来源链接列规则。
        sourceRuleMap = {
            "公司": {
                "nameKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in companyNameHeaders},
                "licenseKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in companyLicenseHeaders},
                "sourceKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in companySourceHeaders},
            },
            "个人": {
                "nameKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in personNameHeaders},
                "licenseKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in personLicenseHeaders},
                "sourceKeys": {headerCleanPattern.sub("", txt(header).lower()) for header in personSourceHeaders},
            },
        }

        # emailBody：推广邮件纯文本正文。
        emailBody = emailMainBody + "\n\n\n" + emailSignature if emailSignature else emailMainBody

        # mainHtml：推广邮件正文 HTML 段落。
        htmlParagraphs = []
        for paragraph in emailMainBody.strip().split("\n\n"):
            lines = [html.escape(line) for line in paragraph.splitlines()]
            htmlParagraphs.append("<p style=\"margin:0 0 14px 0;\">" + "<br>".join(lines) + "</p>")
        mainHtml = "\n".join(htmlParagraphs)

        # logoHtml：Logo 存在时使用 CID 内嵌到 HTML 邮件里。
        safeWebsiteUrl = html.escape(websiteUrl)
        safeLogoCid = html.escape(logoCid)
        logoHtml = ""
        if logoPath.exists():
            logoHtml = (
                f"<img src=\"cid:{safeLogoCid}\" alt=\"Time2Renew\" width=\"{logoWidth}\" "
                f"style=\"display:block; width:{logoWidth}px; height:auto; border:0;\">"
            )

        # emailHtml：推广邮件 HTML 正文。
        emailHtml = f"""<!doctype html>
<html>
<body style="font-family:Arial, Helvetica, sans-serif; font-size:14px; color:#111; line-height:1.45;">
{mainHtml}
<div style="margin-top:28px;">
  <p style="font-size:18px; font-weight:700; margin:0 0 22px 0;">Time2renew Support Team</p>
  <p style="margin:0 0 22px 0;"><strong>Website:</strong> <a href="{safeWebsiteUrl}" style="color:#111; text-decoration:underline;">time2renew</a></p>
  {logoHtml}
</div>
</body>
</html>"""

        # buildRecords：读取公司/个人结果表，并按邮箱生成发送记录。
        records = []
        warnings = []
        for sourceType, filePath in sourceFiles:
            if not filePath.exists():
                warnings.append(f"文件不存在，已跳过: {filePath}")
                continue

            rowCount = 0
            emailCount = 0
            rule = sourceRuleMap[sourceType]
            workbook = load_workbook(str(filePath), read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                rowIterator = worksheet.iter_rows(values_only=True)
                rawHeaders = next(rowIterator, None)
                if not rawHeaders:
                    continue

                # headerKeys：把当前结果表表头标准化，后面直接按标准化表头取值。
                headerKeys = [
                    headerCleanPattern.sub("", txt(header).lower())
                    for header in rawHeaders
                ]

                for rowNumber, values in enumerate(rowIterator, start=2):
                    rowCount += 1
                    rowData = {}
                    for index, headerKey in enumerate(headerKeys):
                        if not headerKey:
                            continue
                        value = values[index] if index < len(values) else ""
                        rowData[headerKey] = txt(value)

                    emailText = "; ".join(value for key, value in rowData.items() if key in emailHeaderKeys and value)
                    emails = []
                    seenEmails = set()
                    for match in emailPattern.findall(emailText):
                        receiver = match.lower().strip("._-")
                        emailName = receiver.split("@", 1)[0] if "@" in receiver else ""
                        emailDomain = receiver.split("@", 1)[1] if "@" in receiver else ""
                        badEmail = (
                            not receiver
                            or ".." in receiver
                            or emailName in badEmailNames
                            or emailDomain.endswith(badEmailTlds)
                            or any(word in emailDomain for word in badEmailWords)
                        )
                        if badEmail or receiver in seenEmails:
                            continue
                        emails.append(receiver)
                        seenEmails.add(receiver)

                    objectName = next((value for key, value in rowData.items() if key in rule["nameKeys"] and value), "")
                    licenseCode = next((value for key, value in rowData.items() if key in rule["licenseKeys"] and value), "")
                    collectStatus = next((value for key, value in rowData.items() if key in collectStatusKeys and value), "")
                    searchStatus = next((value for key, value in rowData.items() if key in searchStatusKeys and value), "")
                    sourceLinks = "; ".join(value for key, value in rowData.items() if key in rule["sourceKeys"] and value)

                    for receiver in emails:
                        emailCount += 1
                        records.append({
                            "来源类型": sourceType,
                            "运行环境": "线上" if self.online else "本机",
                            "对象名称": objectName,
                            "许可证号": licenseCode,
                            "邮箱": receiver,
                            "邮件主题": emailSubject,
                            "邮件正文": emailBody,
                            "采集状态": collectStatus,
                            "搜索状态": searchStatus,
                            "来源链接": sourceLinks,
                            "来源文件": filePath.name,
                            "来源行号": rowNumber,
                            "发送状态": "待发送",
                            "发送结果": "",
                            "失败原因": "",
                            "备注": "",
                        })
            finally:
                workbook.close()
            print(f"{sourceType}推广结果读取完成: {rowCount} 行，生成邮件 {emailCount} 封")

        # sendSummary：统计本次推广邮件流程数量。
        sendSummary = {
            "enabled": True,
            "emailTotal": len(records),
            "emailSent": 0,
            "emailFailed": 0,
            "recordFile": str(recordPath),
        }

        # warnings：把缺少文件等提示输出到日志，但不阻断已有文件的处理。
        for warning in warnings:
            print("提示:", warning)

        # promoSend：False 只生成后台记录表，True 才真实发送推广邮件。
        if not self.promoSend:
            for record in records:
                record["发送状态"] = "待发送"
                record["发送结果"] = "已生成邮件记录，未执行发送"
        elif not smtpPassword:
            for record in records:
                record["发送状态"] = "发送失败"
                record["发送结果"] = "邮件未发送"
                record["失败原因"] = "未配置推广邮件 SMTP 授权码"
            sendSummary["emailFailed"] = len(records)
            print("推广邮件发送失败: 未配置 SMTP 授权码")
        elif records:
            server = None
            try:
                server = smtplib.SMTP_SSL(smtpServer, smtpPort, timeout=30)
                server.login(smtpUser, smtpPassword)
                for index, record in enumerate(records):
                    receiver = record["邮箱"]
                    try:
                        # message：按当前收件人构建一封推广邮件。
                        if not senderEmail:
                            raise ValueError("未配置推广邮件发件邮箱")
                        message = MIMEMultipart("related")
                        message["From"] = formataddr((str(Header("Time2renew Support Team", "utf-8")), senderEmail))
                        message["To"] = receiver
                        message["Subject"] = Header(emailSubject, "utf-8").encode()

                        # alternative：同时提供纯文本和 HTML 两种正文，提升邮件客户端兼容性。
                        alternative = MIMEMultipart("alternative")
                        alternative.attach(MIMEText(emailBody, "plain", "utf-8"))
                        alternative.attach(MIMEText(emailHtml, "html", "utf-8"))
                        message.attach(alternative)

                        # logoPath：Logo 文件存在时内嵌到 HTML 邮件签名中。
                        if logoPath.exists():
                            with logoPath.open("rb") as imageFile:
                                image = MIMEImage(imageFile.read(), _subtype="png")
                            image.add_header("Content-ID", f"<{logoCid}>")
                            image.add_header("Content-Disposition", "inline", filename=logoPath.name)
                            message.attach(image)

                        server.sendmail(senderEmail, [receiver], message.as_bytes())
                        record["发送状态"] = "已发送"
                        record["发送结果"] = "邮件发送成功"
                        sendSummary["emailSent"] += 1
                        print("推广邮件发送成功:", receiver)
                    except Exception as error:
                        record["发送状态"] = "发送失败"
                        record["发送结果"] = "邮件发送失败"
                        record["失败原因"] = str(error)
                        sendSummary["emailFailed"] += 1
                        print("推广邮件发送失败:", receiver, error)

                    if sendWaitSecond > 0 and index < len(records) - 1:
                        time.sleep(sendWaitSecond)
            except Exception as error:
                for record in records:
                    record["发送状态"] = "发送失败"
                    record["发送结果"] = "邮件登录或连接失败"
                    record["失败原因"] = str(error)
                sendSummary["emailFailed"] = len(records)
                print("推广邮件登录或连接失败:", error)
            finally:
                if server:
                    try:
                        server.quit()
                    except Exception:
                        pass

        # saveRecords：无论是否真实发送，都保存后台发送记录，便于复核和继续处理。
        recordPath.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Email Send Record"
        worksheet.append(outputHeaders)
        for record in records:
            worksheet.append([record.get(header, "") for header in outputHeaders])
        worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(str(recordPath))

        print("推广邮件发送记录已导出:", recordPath)
        print("推广邮件任务数量:", len(records))
        return sendSummary

    # 主函数：函数之间的调用
    # 读取清洗表，并按当前模式调用对应函数。
    def main(self):
        txt = lambda value: "" if value is None else str(value).strip()
        self.outputDir.mkdir(parents=True, exist_ok=True)
        dataPath = self.dataDir / self.cleanFile

        if not dataPath.exists():
            print("未找到已清洗初始数据表，流程停止:", dataPath)
            print("当前版本不会自动重新采集 TREC，请先确认内置 file 目录中存在清洗后的全量表。")
            return

        print("运行环境:", "线上" if self.online else "本机")
        print("读取清洗初始数据:", dataPath)
        print("当前运行模式:", "个人模式" if self.mode == "person" else "公司模式")

        # 读取清洗初始表：字段映射直接留在主流程，避免额外 readRows 包装。
        fieldMap = {
            "详情ID": "uid",
            "许可证号": "code",
            "姓名": "name",
            "详情页链接": "url",
            "状态": "status",
            "过期日期": "expiration_date",
            "列表许可证类型": "license_type",
            "城市": "city",
            "县/郡": "county",
            "州": "state",
            "邮编": "zip_code",
            "地址": "address",
            "地区信息": "region_text",
            "挂靠许可证号": "company_code",
            "挂靠公司名称": "company_name",
            "许可证类型": "record_type",
            "关联名称": "relation_name",
        }
        workbook = load_workbook(str(dataPath), read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rowIter = worksheet.iter_rows(values_only=True)
            rawHeaders = next(rowIter, [])
            fields = [fieldMap.get(txt(header), txt(header)) for header in rawHeaders]

            rows = []
            for cells in rowIter:
                row = {}
                for index, field in enumerate(fields):
                    if not field:
                        continue
                    row[field] = txt(cells[index] if index < len(cells) else "")
                rows.append(row)
        finally:
            workbook.close()

        companyRows = []
        personRows = []
        if self.mode == "person":
            personRows = self.person(rows)
        else:
            companyRows = self.company(rows)

        promo = self.promoMail()
        summary = {
            "companyResults": len(companyRows),
            "personResults": len(personRows),
            "environment": "线上" if self.online else "本机",
            "runMode": self.mode,
            "promotionEmailTotal": promo.get("emailTotal", 0),
            "promotionEmailSent": promo.get("emailSent", 0),
            "promotionEmailFailed": promo.get("emailFailed", 0),
            "promotionRecordFile": promo.get("recordFile", ""),
        }
        print(json.dumps(summary, ensure_ascii=False))
        self.sendReport(summary)

    # 结果通知邮件是否发送
    def sendReport(self, summary):
        """按配置发送邮件；未开启邮件时只打印提示。"""
        if not self.sendNotice:
            print("邮件发送未启用，跳过邮件通知。")
            return True

        return EmailUtil.deliverOutputs(
            config=self.config,
            dataDir=self.dataDir,
            outputDir=self.outputDir,
            summary=summary,
        )

    # 执行
    def run(self):
        """主流程入口，只负责调用 main。"""
        self.main()


if __name__ == "__main__":
    # config：单文件调试配置，正式 GUI 会从 run.py 传入。
    config = {
        # runMode：调试时可填 company 或 person。
        "runMode": "company",
    }
    Main(config).run()
