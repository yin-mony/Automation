"""TREC 官方数据、搜索缓存、额度记录和审核结果的统一存储。"""

import hashlib
import json
import sqlite3
import threading
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook


class Data:
    """使用一个 SQLite 数据库管理项目运行状态和 Excel 导出。"""

    def __init__(self, path):
        """初始化数据库路径、TREC 接口和固定字段。"""
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.lock = threading.RLock()
        self.trecUrl = "https://data.texas.gov/resource/s7ft-44qi.json"
        self.trecFields = [
            "license_type", "license_number", "full_name", "suffix", "status",
            "original_license_date", "license_expiration_date", "designated_supervisor_flag",
            "county", "related_license_type", "related_license_number",
            "related_license_full_name", "related_license_suffix", "related_license_start_date",
            "agency_identifier", "key_name", "first_name", "middle_name", "last_name", "updated",
        ]
        self.initialize()

    def nowText(self):
        """返回带时区的当前时间文本。"""
        return datetime.now().astimezone().isoformat(timespec="seconds")

    def normalizeKey(self, value):
        """把名称压缩为空格统一的小写键。"""
        return " ".join(str(value or "").strip().lower().split())

    def openDb(self):
        """创建一个启用字典行访问的 SQLite 连接。"""
        connection = sqlite3.connect(self.path, timeout=30)
        connection.row_factory = sqlite3.Row
        return connection

    def initialize(self):
        """创建 TREC、缓存、额度、结果和审核表。"""
        schema = """
        PRAGMA journal_mode=WAL;
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS trec_licenses (
            license_number TEXT PRIMARY KEY,
            license_type TEXT,
            full_name TEXT,
            suffix TEXT,
            status TEXT,
            original_license_date TEXT,
            license_expiration_date TEXT,
            designated_supervisor_flag TEXT,
            county TEXT,
            related_license_type TEXT,
            related_license_number TEXT,
            related_license_full_name TEXT,
            related_license_suffix TEXT,
            related_license_start_date TEXT,
            agency_identifier TEXT,
            key_name TEXT,
            first_name TEXT,
            middle_name TEXT,
            last_name TEXT,
            updated TEXT,
            synced_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_trec_status ON trec_licenses(status);
        CREATE INDEX IF NOT EXISTS idx_trec_related ON trec_licenses(related_license_number);
        CREATE INDEX IF NOT EXISTS idx_trec_expiration ON trec_licenses(license_expiration_date);
        CREATE TABLE IF NOT EXISTS search_cache (
            query_hash TEXT PRIMARY KEY,
            mode TEXT NOT NULL,
            object_key TEXT NOT NULL,
            query TEXT NOT NULL,
            response_json TEXT,
            result_json TEXT,
            search_id TEXT,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_cache_object ON search_cache(mode, object_key);
        CREATE TABLE IF NOT EXISTS contact_results (
            mode TEXT NOT NULL,
            object_key TEXT NOT NULL,
            object_name TEXT NOT NULL,
            license_number TEXT,
            result_json TEXT NOT NULL,
            status TEXT NOT NULL,
            review_status TEXT NOT NULL DEFAULT 'pending',
            updated_at TEXT NOT NULL,
            PRIMARY KEY(mode, object_key)
        );
        CREATE TABLE IF NOT EXISTS quota_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query_hash TEXT NOT NULL,
            mode TEXT NOT NULL,
            object_key TEXT NOT NULL,
            category TEXT NOT NULL,
            success INTEGER NOT NULL,
            counted INTEGER NOT NULL,
            search_id TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_quota_created ON quota_events(created_at);
        CREATE TABLE IF NOT EXISTS review_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mode TEXT NOT NULL,
            object_key TEXT NOT NULL,
            object_name TEXT NOT NULL,
            source_url TEXT,
            confidence INTEGER NOT NULL,
            reason TEXT,
            contacts_json TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(mode, object_key, source_url)
        );
        CREATE TABLE IF NOT EXISTS mail_actions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            recipient TEXT NOT NULL,
            action TEXT NOT NULL,
            object_key TEXT,
            object_name TEXT,
            subject_hash TEXT,
            status TEXT NOT NULL,
            detail TEXT,
            created_at TEXT NOT NULL,
            UNIQUE(recipient, action)
        );
        CREATE INDEX IF NOT EXISTS idx_mail_recipient ON mail_actions(recipient);
        """
        with self.lock:
            connection = self.openDb()
            try:
                connection.executescript(schema)
                connection.commit()
            finally:
                connection.close()

    def getMeta(self, key, default = ""):
        """读取一项运行元数据。"""
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute("SELECT value FROM metadata WHERE key = ?", (key,)).fetchone()
            finally:
                connection.close()
        return str(row["value"]) if row else default

    def setMeta(self, key, value):
        """新增或覆盖一项运行元数据。"""
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    "INSERT INTO metadata(key, value) VALUES(?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    (key, str(value)),
                )
                connection.commit()
            finally:
                connection.close()

    def upsertTrec(self, rows):
        """按许可证号批量写入或更新官方 TREC 数据。"""
        if not rows:
            return
        placeholders = ", ".join("?" for _ in self.trecFields)
        updates = ", ".join(f"{field}=excluded.{field}" for field in self.trecFields[1:])
        sql = (
            f"INSERT INTO trec_licenses({', '.join(self.trecFields)}, synced_at) "
            f"VALUES({placeholders}, ?) ON CONFLICT(license_number) DO UPDATE SET "
            f"{updates}, synced_at=excluded.synced_at"
        )
        stamp = self.nowText()
        values = [
            tuple(str(row.get(field, "") or "") for field in self.trecFields) + (stamp,)
            for row in rows
        ]
        with self.lock:
            connection = self.openDb()
            try:
                connection.executemany(sql, values)
                connection.commit()
            finally:
                connection.close()

    def trecCount(self):
        """返回本地 TREC 许可证总数。"""
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute("SELECT COUNT(*) AS count FROM trec_licenses").fetchone()
            finally:
                connection.close()
        return int(row["count"])

    def trecRows(self):
        """按许可证号读取全部 TREC 数据。"""
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(
                    "SELECT * FROM trec_licenses ORDER BY license_number"
                ).fetchall()
            finally:
                connection.close()
        return [dict(row) for row in rows]

    def companyCandidates(self):
        """按挂靠公司去重汇总 Active 经纪人。"""
        sql = """
        SELECT
            COALESCE(
                NULLIF(TRIM(related_license_number), ''),
                LOWER(TRIM(related_license_full_name))
            ) AS objectKey,
            related_license_full_name AS name,
            related_license_number AS licenseNumber,
            MIN(county) AS county,
            COUNT(*) AS agentCount,
            MIN(full_name) AS sampleAgent
        FROM trec_licenses
        WHERE LOWER(status) = 'active'
          AND TRIM(COALESCE(related_license_full_name, '')) <> ''
        GROUP BY LOWER(TRIM(related_license_full_name)), related_license_number
        ORDER BY related_license_full_name
        """
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(sql).fetchall()
            finally:
                connection.close()
        return [dict(row) for row in rows]

    def parseDate(self, value):
        """兼容解析 TREC 常见日期格式。"""
        text = str(value or "").strip()
        for formatText in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(text[:19], formatText).date()
            except ValueError:
                continue
        return None

    def personCandidates(self, expireMonths):
        """筛选无挂靠且在指定月份内到期的 Active 个人。"""
        today = date.today()
        monthIndex = today.month - 1 + max(1, int(expireMonths))
        endYear = today.year + monthIndex // 12
        endMonth = monthIndex % 12 + 1
        endDay = min(today.day, monthrange(endYear, endMonth)[1])
        cutoff = date(endYear, endMonth, endDay)
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(
                    """
                    SELECT * FROM trec_licenses
                    WHERE LOWER(status) = 'active'
                      AND LOWER(COALESCE(license_type, '')) NOT LIKE '%company%'
                      AND TRIM(COALESCE(related_license_full_name, '')) = ''
                    ORDER BY full_name, license_number
                    """
                ).fetchall()
            finally:
                connection.close()

        output = []
        for raw in rows:
            row = dict(raw)
            expiration = self.parseDate(row.get("license_expiration_date"))
            if not expiration or expiration < today or expiration > cutoff:
                continue
            objectKey = row.get("agency_identifier") or self.normalizeKey(
                f"{row.get('full_name', '')} {row.get('license_number', '')}"
            )
            output.append({
                "objectKey": objectKey,
                "name": row.get("full_name", ""),
                "licenseNumber": row.get("license_number", ""),
                "agencyIdentifier": row.get("agency_identifier", ""),
                "status": row.get("status", ""),
                "expirationDate": row.get("license_expiration_date", ""),
                "licenseType": row.get("license_type", ""),
                "county": row.get("county", ""),
                "relatedLicenseName": row.get("related_license_full_name", ""),
            })
        return output

    def queryHash(self, query, mode, objectKey):
        """为搜索对象和搜索词生成稳定缓存键。"""
        value = json.dumps([mode, objectKey, query], ensure_ascii=False, separators=(",", ":"))
        return hashlib.sha256(value.encode("utf-8")).hexdigest()

    def searchCache(self, queryHash):
        """读取一次完整的搜索缓存记录。"""
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute(
                    "SELECT * FROM search_cache WHERE query_hash = ?", (queryHash,)
                ).fetchone()
            finally:
                connection.close()
        return dict(row) if row else None

    def saveSearchPayload(
        self,
        queryHash,
        mode,
        objectKey,
        query,
        payload,
        status = "fetched",
    ):
        """先保存 SerpApi 原始响应，支持中断后继续提取。"""
        stamp = self.nowText()
        searchId = str((payload.get("search_metadata") or {}).get("id") or "")
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    """
                    INSERT INTO search_cache(
                        query_hash, mode, object_key, query, response_json, result_json,
                        search_id, status, created_at, updated_at
                    ) VALUES(?, ?, ?, ?, ?, '', ?, ?, ?, ?)
                    ON CONFLICT(query_hash) DO UPDATE SET
                        response_json=excluded.response_json,
                        search_id=excluded.search_id,
                        status=excluded.status,
                        updated_at=excluded.updated_at
                    """,
                    (
                        queryHash, mode, objectKey, query,
                        json.dumps(payload, ensure_ascii=False), searchId, status, stamp, stamp,
                    ),
                )
                connection.commit()
            finally:
                connection.close()

    def saveSearchResult(self, queryHash, result, status = "complete"):
        """保存经过网页补充后的最终搜索结果。"""
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    "UPDATE search_cache SET result_json = ?, status = ?, updated_at = ? "
                    "WHERE query_hash = ?",
                    (json.dumps(result, ensure_ascii=False), status, self.nowText(), queryHash),
                )
                connection.commit()
            finally:
                connection.close()

    def saveContact(
        self,
        mode,
        objectKey,
        objectName,
        licenseNumber,
        result,
        status = "complete",
    ):
        """按公司或个人唯一键保存最终联系方式。"""
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    """
                    INSERT INTO contact_results(
                        mode, object_key, object_name, license_number, result_json, status, updated_at
                    ) VALUES(?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(mode, object_key) DO UPDATE SET
                        object_name=excluded.object_name,
                        license_number=excluded.license_number,
                        result_json=excluded.result_json,
                        status=excluded.status,
                        updated_at=excluded.updated_at
                    """,
                    (
                        mode, objectKey, objectName, licenseNumber,
                        json.dumps(result, ensure_ascii=False), status, self.nowText(),
                    ),
                )
                connection.commit()
            finally:
                connection.close()

    def hasContact(self, mode, objectKey):
        """判断一个对象是否已经完成联系方式采集。"""
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute(
                    "SELECT 1 FROM contact_results WHERE mode = ? AND object_key = ? "
                    "AND status = 'complete'",
                    (mode, objectKey),
                ).fetchone()
            finally:
                connection.close()
        return bool(row)

    def contactResults(self, mode = None):
        """读取并展开公司或个人结果 JSON。"""
        sql = "SELECT * FROM contact_results"
        params = ()
        if mode:
            sql += " WHERE mode = ?"
            params = (mode,)
        sql += " ORDER BY updated_at"
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(sql, params).fetchall()
            finally:
                connection.close()
        output = []
        for raw in rows:
            row = dict(raw)
            result = json.loads(str(row.get("result_json") or "{}"))
            record = {
                "mode": row.get("mode", ""),
                "objectKey": row.get("object_key", ""),
                "objectName": row.get("object_name", ""),
                "licenseNumber": row.get("license_number", ""),
                "status": row.get("status", ""),
                "reviewStatus": row.get("review_status", "pending"),
                "updatedAt": row.get("updated_at", ""),
            }
            record.update(result)
            output.append(record)
        return output

    def hasMailAction(self, recipient, action):
        """判断邮箱是否已经创建过草稿或完成过真实发送。"""
        email = str(recipient or "").strip().lower()
        actions = ("draft", "send") if action == "draft" else ("send",)
        placeholders = ", ".join("?" for _ in actions)
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute(
                    f"SELECT 1 FROM mail_actions WHERE recipient = ? "
                    f"AND action IN ({placeholders}) AND status = 'success'",
                    (email, *actions),
                ).fetchone()
            finally:
                connection.close()
        return bool(row)

    def recordMailAction(
        self,
        recipient,
        action,
        objectKey,
        objectName,
        subjectHash,
        detail="",
    ):
        """持久记录成功草稿或真实发送，作为跨运行去重依据。"""
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    """
                    INSERT INTO mail_actions(
                        recipient, action, object_key, object_name, subject_hash,
                        status, detail, created_at
                    ) VALUES(?, ?, ?, ?, ?, 'success', ?, ?)
                    ON CONFLICT(recipient, action) DO UPDATE SET
                        object_key=excluded.object_key,
                        object_name=excluded.object_name,
                        subject_hash=excluded.subject_hash,
                        status='success',
                        detail=excluded.detail,
                        created_at=excluded.created_at
                    """,
                    (
                        str(recipient or "").strip().lower(),
                        action,
                        objectKey,
                        objectName,
                        subjectHash,
                        detail,
                        self.nowText(),
                    ),
                )
                connection.commit()
            finally:
                connection.close()

    def mailActionSummary(self):
        """返回草稿、真实发送和唯一收件人数。"""
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(
                    "SELECT action, COUNT(*) AS count FROM mail_actions "
                    "WHERE status = 'success' GROUP BY action"
                ).fetchall()
                uniqueRow = connection.execute(
                    "SELECT COUNT(DISTINCT recipient) AS count FROM mail_actions "
                    "WHERE status = 'success'"
                ).fetchone()
            finally:
                connection.close()
        counts = {str(row["action"]): int(row["count"]) for row in rows}
        return {
            "drafts": counts.get("draft", 0),
            "sent": counts.get("send", 0),
            "uniqueRecipients": int(uniqueRow["count"] if uniqueRow else 0),
        }

    def recordQuota(
        self,
        queryHash,
        mode,
        objectKey,
        category,
        success,
        counted,
        searchId = "",
    ):
        """记录每一次 SerpApi 请求是否可能消耗额度。"""
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    """
                    INSERT INTO quota_events(
                        query_hash, mode, object_key, category, success, counted, search_id, created_at
                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        queryHash, mode, objectKey, category, int(success), int(counted),
                        searchId, self.nowText(),
                    ),
                )
                connection.commit()
            finally:
                connection.close()

    def quotaCount(
        self,
        category = None,
        dayOnly = False,
        mode = None,
    ):
        """统计今日或本月已计费的 SerpApi 请求数。"""
        clauses = ["counted = 1"]
        params = []
        if category:
            clauses.append("category = ?")
            params.append(category)
        if mode:
            clauses.append("mode = ?")
            params.append(mode)
        if dayOnly:
            clauses.append("substr(created_at, 1, 10) = ?")
            params.append(date.today().isoformat())
        else:
            clauses.append("substr(created_at, 1, 7) = ?")
            params.append(date.today().strftime("%Y-%m"))
        with self.lock:
            connection = self.openDb()
            try:
                row = connection.execute(
                    f"SELECT COUNT(*) AS count FROM quota_events WHERE {' AND '.join(clauses)}",
                    params,
                ).fetchone()
            finally:
                connection.close()
        return int(row["count"])

    def saveReview(
        self,
        mode,
        objectKey,
        objectName,
        sourceUrl,
        confidence,
        reason,
        contacts,
    ):
        """新增或更新一条等待人工确认的联系方式。"""
        stamp = self.nowText()
        with self.lock:
            connection = self.openDb()
            try:
                connection.execute(
                    """
                    INSERT INTO review_items(
                        mode, object_key, object_name, source_url, confidence, reason,
                        contacts_json, status, created_at, updated_at
                    ) VALUES(?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)
                    ON CONFLICT(mode, object_key, source_url) DO UPDATE SET
                        confidence=excluded.confidence,
                        reason=excluded.reason,
                        contacts_json=excluded.contacts_json,
                        updated_at=excluded.updated_at
                    """,
                    (
                        mode, objectKey, objectName, sourceUrl, int(confidence), reason,
                        json.dumps(contacts, ensure_ascii=False), stamp, stamp,
                    ),
                )
                connection.commit()
            finally:
                connection.close()

    def reviewItems(self, status = "pending"):
        """读取指定状态或全部人工审核记录。"""
        sql = "SELECT * FROM review_items"
        params = ()
        if status != "all":
            sql += " WHERE status = ?"
            params = (status,)
        sql += " ORDER BY created_at"
        with self.lock:
            connection = self.openDb()
            try:
                rows = connection.execute(sql, params).fetchall()
            finally:
                connection.close()
        output = []
        for raw in rows:
            row = dict(raw)
            contacts = json.loads(str(row.pop("contacts_json") or "{}"))
            row["contacts"] = contacts
            row["emailText"] = "; ".join(contacts.get("emails") or [])
            row["phoneText"] = "; ".join(contacts.get("phones") or [])
            output.append(row)
        return output

    def setReviewStatus(self, itemId, status):
        """更新一条审核状态并同步到最终结果。"""
        return self.setReviewStatuses([itemId], status)

    def setReviewStatuses(self, itemIds, status):
        """在同一事务中批量更新审核状态和对应结果。"""
        if status not in {"pending", "approved", "rejected"}:
            raise ValueError("无效的审核状态")
        ids = list(dict.fromkeys(int(itemId) for itemId in itemIds))
        if not ids:
            return 0
        stamp = self.nowText()
        updated = 0
        with self.lock:
            connection = self.openDb()
            try:
                for itemId in ids:
                    item = connection.execute(
                        "SELECT mode, object_key FROM review_items WHERE id = ?", (itemId,)
                    ).fetchone()
                    if not item:
                        continue
                    connection.execute(
                        "UPDATE review_items SET status = ?, updated_at = ? WHERE id = ?",
                        (status, stamp, itemId),
                    )
                    connection.execute(
                        "UPDATE contact_results SET review_status = ?, updated_at = ? "
                        "WHERE mode = ? AND object_key = ?",
                        (status, stamp, item["mode"], item["object_key"]),
                    )
                    updated += 1
                connection.commit()
            finally:
                connection.close()
        return updated

    def summary(self):
        """返回 GUI 概览需要的本地统计数据。"""
        with self.lock:
            connection = self.openDb()
            try:
                pending = connection.execute(
                    "SELECT COUNT(*) AS count FROM review_items WHERE status = 'pending'"
                ).fetchone()
                company = connection.execute(
                    "SELECT COUNT(*) AS count FROM contact_results WHERE mode = 'company'"
                ).fetchone()
                person = connection.execute(
                    "SELECT COUNT(*) AS count FROM contact_results WHERE mode = 'person'"
                ).fetchone()
            finally:
                connection.close()
        return {
            "trecCount": self.trecCount(),
            "reviewPending": int(pending["count"]),
            "companyResults": int(company["count"]),
            "personResults": int(person["count"]),
            "lastSync": self.getMeta("trec_last_sync_at"),
        }

    def syncTrec(
        self,
        full = False,
        pageSize = 50000,
        checkpoint=None,
        log=None,
        progress=None,
    ):
        """从 Texas Open Data 直连同步全量或增量 TREC 数据。"""
        checkpoint = checkpoint or (lambda: None)
        log = log or (lambda message: None)
        progress = progress or (lambda current, total, message: None)
        previous = self.getMeta("trec_last_updated")
        incremental = bool(previous and not full)
        offset = 0
        fetched = 0
        newest = previous
        log("开始直连同步 Texas Open Data 官方 TREC 数据。")
        while True:
            checkpoint()
            params = {
                "$limit": int(pageSize),
                "$offset": offset,
                "$order": "updated,license_number" if incremental else "license_number",
            }
            if incremental:
                params["$where"] = f"updated > '{previous}'"
            request = Request(
                self.trecUrl + "?" + urlencode(params),
                headers={"Accept": "application/json", "User-Agent": "TREC-Automation/3.0"},
            )
            response = urlopen(request, timeout=90)
            rows = json.loads(response.read().decode("utf-8"))
            if not isinstance(rows, list):
                raise RuntimeError("TREC 官方数据返回格式异常")
            if not rows:
                break
            self.upsertTrec(rows)
            fetched += len(rows)
            offset += len(rows)
            for row in rows:
                updated = str(row.get("updated") or "")
                if updated > newest:
                    newest = updated
            progress(fetched, 0, f"已同步 {fetched:,} 条 TREC 数据")
            log(f"TREC 数据页同步完成：累计 {fetched:,} 条。")
            if len(rows) < pageSize:
                break
        if newest:
            self.setMeta("trec_last_updated", newest)
        self.setMeta("trec_last_sync_at", self.nowText())
        return {
            "mode": "incremental" if incremental else "full",
            "fetched": fetched,
            "total": self.trecCount(),
            "lastUpdated": newest,
        }

    def exportTrec(self, dataDir, checkpoint=None, progress=None):
        """导出兼容旧流程的未清洗和已清洗中文底表。"""
        checkpoint = checkpoint or (lambda: None)
        progress = progress or (lambda current, total, message: None)
        dataPath = Path(dataDir)
        dataPath.mkdir(parents=True, exist_ok=True)
        rawPath = dataPath / "初始总量数据未清洗.xlsx"
        cleanPath = dataPath / "已获取到的初始总数据.xlsx"
        cleanHeaders = [
            "详情ID", "许可证号", "姓名", "详情页链接", "状态", "过期日期",
            "列表许可证类型", "城市", "县/郡", "州", "邮编", "地址", "地区信息",
            "挂靠许可证号", "挂靠公司名称", "许可证类型", "关联名称",
        ]
        rawBook = Workbook(write_only=True)
        rawSheet = rawBook.create_sheet("TREC Raw")
        rawSheet.append(self.trecFields)
        cleanBook = Workbook(write_only=True)
        cleanSheet = cleanBook.create_sheet("TREC Clean")
        cleanSheet.append(cleanHeaders)
        rows = self.trecRows()
        for index, row in enumerate(rows, start=1):
            checkpoint()
            rawSheet.append([row.get(field, "") for field in self.trecFields])
            relatedName = str(row.get("related_license_full_name") or "")
            relatedNumber = str(row.get("related_license_number") or "")
            county = str(row.get("county") or "")
            cleanSheet.append([
                row.get("agency_identifier", ""), row.get("license_number", ""),
                row.get("full_name", ""), "", row.get("status", ""),
                row.get("license_expiration_date", ""), row.get("license_type", ""),
                "", county, "TX", "", "", "; ".join(value for value in (county, "TX") if value),
                relatedNumber, relatedName, row.get("license_type", ""), relatedName,
            ])
            if index % 10000 == 0:
                progress(index, len(rows), f"正在导出底表：{index:,} 条")
        rawBook.save(rawPath)
        cleanBook.save(cleanPath)
        return rawPath, cleanPath

    def exportResults(self, mode, outputDir, config):
        """按原中文文件名导出公司或个人最终结果。"""
        rows = self.contactResults(mode)
        if mode == "company":
            fileName = str(
                config.get("companyResultFileName") or "已完成搜索匹配的公司联系信息数据.xlsx"
            )
            headers = [
                ("objectName", "公司名称"), ("licenseNumber", "挂靠许可证号"),
                ("agentCount", "关联经纪人数量"), ("sampleAgent", "样本经纪人"),
                ("county", "地区信息"), ("query", "搜索词"), ("emails", "邮箱"),
                ("phones", "电话"), ("facebookUrls", "Facebook主页链接"),
                ("sourceUrls", "SerpApi来源链接"), ("detailUrls", "已抓取二级页面链接"),
                ("resultCount", "SerpApi结果数"), ("contactStatus", "采集状态"),
                ("reviewStatus", "审核状态"), ("pageErrors", "错误原因"),
            ]
            sheetTitle = "公司联系信息"
        else:
            fileName = str(
                config.get("personResultFileName") or "已完成搜索匹配的个人联系信息数据.xlsx"
            )
            headers = [
                ("objectName", "姓名"), ("licenseNumber", "许可证号"),
                ("agencyIdentifier", "详情ID"), ("status", "状态"),
                ("expirationDate", "过期日期"), ("licenseType", "许可证类型"),
                ("county", "县/郡"), ("query", "搜索词"), ("emails", "邮箱"),
                ("phones", "电话"), ("facebookUrls", "Facebook主页链接"),
                ("sourceUrls", "SerpApi来源链接"), ("detailUrls", "已抓取二级页面链接"),
                ("resultCount", "SerpApi结果数"), ("contactStatus", "采集状态"),
                ("reviewStatus", "审核状态"), ("pageErrors", "错误原因"),
            ]
            sheetTitle = "个人联系信息"

        path = Path(outputDir) / fileName
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheetTitle
        sheet.append([label for _, label in headers])
        for row in rows:
            values = []
            for key, _ in headers:
                value = row.get(key, "")
                if isinstance(value, list):
                    value = "; ".join(str(item) for item in value)
                values.append(value)
            sheet.append(values)
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(path)
        return path
